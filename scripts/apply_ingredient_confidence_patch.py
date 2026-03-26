#!/usr/bin/env python3
"""Apply confidence patch CSV to an ingredient workbook copy."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to apply workbook confidence patches. Install it in the local Python environment first."
    ) from exc


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def load_patch_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def build_header_index(sheet) -> dict[str, int]:
    header = [normalize_text(cell.value) for cell in sheet[1]]
    return {name: index + 1 for index, name in enumerate(header)}


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply ingredient confidence patch CSV to a workbook copy.")
    parser.add_argument("--ingredient-xlsx", required=True, help="Source ingredient workbook")
    parser.add_argument("--patch-csv", required=True, help="Apply-ready confidence patch CSV")
    parser.add_argument("--out-xlsx", required=True, help="Path to write the patched workbook copy")
    parser.add_argument("--out-report-json", help="Optional path for apply report JSON")
    args = parser.parse_args()

    workbook_path = Path(args.ingredient_xlsx).expanduser().resolve()
    patch_path = Path(args.patch_csv).expanduser().resolve()
    out_xlsx = Path(args.out_xlsx).expanduser().resolve()

    workbook = load_workbook(workbook_path)
    if "Dictionary" not in workbook.sheetnames:
        raise SystemExit("Workbook missing required 'Dictionary' sheet.")
    sheet = workbook["Dictionary"]
    header_index = build_header_index(sheet)

    required_columns = {"record_id", "confidence"}
    missing_columns = [column for column in required_columns if column not in header_index]
    if missing_columns:
        raise SystemExit(f"Workbook Dictionary sheet is missing required columns: {', '.join(missing_columns)}")

    patch_rows = load_patch_rows(patch_path)
    row_by_record_id: dict[str, int] = {}
    for row_number in range(2, sheet.max_row + 1):
        record_id = normalize_text(sheet.cell(row=row_number, column=header_index["record_id"]).value)
        if record_id:
            row_by_record_id[record_id] = row_number

    applied: list[str] = []
    skipped_missing_record: list[str] = []
    skipped_conflicts: list[dict[str, str]] = []

    for patch in patch_rows:
        record_id = normalize_text(patch.get("record_id"))
        if not record_id:
            continue
        row_number = row_by_record_id.get(record_id)
        if not row_number:
            skipped_missing_record.append(record_id)
            continue

        confidence_cell = sheet.cell(row=row_number, column=header_index["confidence"])
        current_confidence = normalize_text(confidence_cell.value)
        expected_confidence = normalize_text(patch.get("existing_confidence"))

        if current_confidence != expected_confidence:
            skipped_conflicts.append(
                {
                    "record_id": record_id,
                    "current_confidence": current_confidence,
                    "expected_confidence": expected_confidence,
                }
            )
            continue

        confidence_cell.value = normalize_text(patch.get("patch_confidence"))
        applied.append(record_id)

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_xlsx)

    report = {
        "source_workbook": str(workbook_path),
        "patch_csv": str(patch_path),
        "out_workbook": str(out_xlsx),
        "patch_row_count": len(patch_rows),
        "applied_count": len(applied),
        "skipped_missing_record_count": len(skipped_missing_record),
        "skipped_conflict_count": len(skipped_conflicts),
        "skipped_missing_record_ids": skipped_missing_record,
        "skipped_conflicts": skipped_conflicts,
    }

    rendered = json.dumps(report, ensure_ascii=True, indent=2)
    if args.out_report_json:
        report_path = Path(args.out_report_json).expanduser().resolve()
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(rendered + "\n", encoding="utf-8")
    print(rendered)


if __name__ == "__main__":
    main()
