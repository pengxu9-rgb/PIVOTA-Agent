#!/usr/bin/env python3
"""Build deterministic parser note proposals for ingredient workbook rows."""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to propose workbook parser note backfills. Install it in the local Python environment first."
    ) from exc


PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]


BOTANICAL_NOTE = "Botanical INCI often appears with a shorter common-language alias on PDPs."
MARKETING_NOTE = "Contains marketing/trade/common shorthand aliases; prefer canonical match when available."
LEGACY_NOTE = "Legacy term kept for backfill and historical label coverage."
CROSS_MARKET_NOTE = "Normalize cross-market label forms to canonical reference name."
BOTANICAL_MARKETING_NOTE = (
    "Botanical INCI often appears with a shorter common-language alias on PDPs. "
    "Contains marketing/trade/common shorthand aliases; prefer canonical match when available."
)
PEPTIDE_NOTE = "Keep peptide chain numbers and hyphenation during parsing; PDPs often vary separators."
ETHOXYLATE_NOTE = "Keep PEG/ethoxylate numbers and slash tokens during parsing; PDPs often vary separators."
SEPARATOR_NOTE = "Keep numeric prefixes, hyphenation, and slash tokens during parsing; PDPs often vary separators."
FERMENT_NOTE = "Ferment names may appear in shortened PDP form on PDPs; prefer canonical match when available."

PLANT_PART_TOKENS = {
    "extract",
    "oil",
    "juice",
    "butter",
    "leaf",
    "flower",
    "fruit",
    "seed",
    "kernel",
    "root",
    "bark",
    "stem",
    "bran",
    "wood",
    "powder",
    "water",
    "resin",
    "gum",
    "flour",
}


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def parse_multi(value: Any) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split(";") if part and part.strip()]


def resolve_sheet_name(workbook_path: Path, requested_sheet: str | None = None) -> str:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    if requested_sheet:
        if requested_sheet not in sheet_names:
            raise SystemExit(f"Workbook sheet not found: {requested_sheet}")
        return requested_sheet
    for candidate in PREFERRED_INGREDIENT_SHEETS:
        if candidate in sheet_names:
            return candidate
    raise SystemExit(
        "Workbook is missing a supported ingredient sheet. Expected one of: "
        + ", ".join(PREFERRED_INGREDIENT_SHEETS)
    )


def read_records(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(cell is not None and normalize_text(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        records.append(dict(zip(header, padded)))
    return records


def normalize_key_like(value: str) -> str:
    text = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    return "".join(ch for ch in text if ch.isalnum())


def same_normalized(left: str, right: str) -> bool:
    return bool(left and right and normalize_key_like(left) == normalize_key_like(right))


def has_material_display_alias(record: dict[str, Any], canonical: str) -> bool:
    display_name = normalize_text(record.get("canonical_display_name"))
    return bool(display_name and not same_normalized(display_name, canonical))


def has_material_label_variants(record: dict[str, Any], canonical: str) -> bool:
    for column in ["us_label_name", "eu_label_name", "us_label_variants", "eu_label_variants"]:
        for value in parse_multi(record.get(column)) if column.endswith("variants") else [record.get(column)]:
            text = normalize_text(value)
            if text and not same_normalized(text, canonical):
                return True
    return False


def has_cross_market_difference(record: dict[str, Any]) -> bool:
    us_name = normalize_text(record.get("us_label_name"))
    eu_name = normalize_text(record.get("eu_label_name"))
    return bool(us_name and eu_name and not same_normalized(us_name, eu_name))


def looks_botanical(record: dict[str, Any], canonical: str) -> bool:
    family = normalize_text(record.get("ingredient_family")).casefold()
    if family == "plant_extract":
        return True
    lower = canonical.casefold()
    tokens = set(re.findall(r"[a-z]+", lower))
    return len(tokens & PLANT_PART_TOKENS) >= 1 and len(lower.split()) >= 3


def looks_ferment(canonical: str) -> bool:
    return "ferment" in canonical.casefold()


def looks_peptide(record: dict[str, Any], canonical: str) -> bool:
    family = normalize_text(record.get("ingredient_family")).casefold()
    return family == "peptide" or "peptide" in canonical.casefold()


def looks_ethoxylate_or_surfactant(record: dict[str, Any], canonical: str) -> bool:
    family = normalize_text(record.get("ingredient_family")).casefold()
    lower = canonical.casefold()
    if family != "surfactant":
        return False
    return bool(
        re.search(r"\b(peg|ppg|laureth|oleth|ceteareth|polysorbate)\b", lower)
        or "/" in lower
        or re.search(r"\bglucoside\b", lower)
    )


def looks_separator_sensitive(canonical: str) -> bool:
    lower = canonical.casefold()
    return bool(re.search(r"\d", lower) and any(token in lower for token in ["-", "/", ","]))


def looks_legacy_abbreviation(canonical: str) -> bool:
    upper = canonical.strip().upper()
    return upper in {"PABA", "BHA", "BHT"}


def dedupe_keep_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        text = normalize_text(value)
        if not text:
            continue
        key = text.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(text)
    return out


def build_note_proposal(record: dict[str, Any]) -> tuple[str, str, str, list[str], str]:
    canonical = normalize_text(record.get("canonical_inci_name"))
    reasons: list[str] = []

    cross_market = has_cross_market_difference(record)
    display_alias = has_material_display_alias(record, canonical)
    label_alias = has_material_label_variants(record, canonical)
    botanical = looks_botanical(record, canonical)
    ferment = looks_ferment(canonical)
    peptide = looks_peptide(record, canonical)
    surfactant = looks_ethoxylate_or_surfactant(record, canonical)
    separator_sensitive = looks_separator_sensitive(canonical)
    legacy_abbrev = looks_legacy_abbreviation(canonical)

    if cross_market:
        reasons.append("cross_market_label_difference")
    if display_alias or label_alias:
        reasons.append("material_alias_or_shorthand_present")
    if botanical:
        reasons.append("botanical_or_latin_inci")
    if ferment:
        reasons.append("ferment_name")
    if peptide:
        reasons.append("peptide_family")
    if surfactant:
        reasons.append("ethoxylate_or_surfactant_separator_pattern")
    if separator_sensitive:
        reasons.append("numeric_separator_sensitive")
    if legacy_abbrev:
        reasons.append("legacy_abbreviation")

    if botanical and (display_alias or label_alias):
        return BOTANICAL_MARKETING_NOTE, "botanical_marketing", "high", reasons, "no"
    if cross_market:
        return CROSS_MARKET_NOTE, "cross_market", "high", reasons, "no"
    if botanical:
        return BOTANICAL_NOTE, "botanical", "high", reasons, "no"
    if peptide:
        return PEPTIDE_NOTE, "peptide_numeric", "high", reasons, "no"
    if surfactant:
        return ETHOXYLATE_NOTE, "ethoxylate_surfactant", "high", reasons, "no"
    if legacy_abbrev:
        return LEGACY_NOTE, "legacy_abbreviation", "high", reasons, "no"
    if ferment and (display_alias or label_alias):
        return FERMENT_NOTE, "ferment_shorthand", "medium", reasons, "yes"
    if display_alias or label_alias:
        return MARKETING_NOTE, "marketing_shorthand", "medium", reasons, "yes"
    if separator_sensitive:
        return SEPARATOR_NOTE, "separator_sensitive", "medium", reasons, "yes"
    return "", "manual_review_required", "none", reasons, "yes"


def build_proposals(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    proposals: list[dict[str, Any]] = []
    for record in records:
        if normalize_text(record.get("notes_for_parser")):
            continue
        canonical = normalize_text(record.get("canonical_inci_name"))
        suggested_note, template_name, confidence, reasons, needs_manual = build_note_proposal(record)
        proposals.append(
            {
                "record_id": normalize_text(record.get("record_id")),
                "canonical_inci_name": canonical,
                "canonical_display_name": normalize_text(record.get("canonical_display_name")),
                "ingredient_family": normalize_text(record.get("ingredient_family")),
                "existing_notes_for_parser": normalize_text(record.get("notes_for_parser")),
                "suggested_notes_for_parser": suggested_note,
                "proposal_template": template_name,
                "proposal_confidence": confidence,
                "proposal_reasons": "; ".join(dedupe_keep_order(reasons)),
                "needs_manual_review": needs_manual,
            }
        )
    return proposals


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "record_id",
        "canonical_inci_name",
        "canonical_display_name",
        "ingredient_family",
        "existing_notes_for_parser",
        "suggested_notes_for_parser",
        "proposal_template",
        "proposal_confidence",
        "proposal_reasons",
        "needs_manual_review",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Propose deterministic parser note backfills for the ingredient workbook.")
    parser.add_argument("--ingredient-xlsx", required=True, help="Path to the ingredient reference workbook")
    parser.add_argument("--sheet-name", help="Optional workbook sheet name")
    parser.add_argument("--out-json", help="Optional path to write JSON summary")
    parser.add_argument("--out-csv", help="Optional path to write proposal CSV")
    args = parser.parse_args()

    workbook_path = Path(args.ingredient_xlsx).expanduser().resolve()
    sheet_name = resolve_sheet_name(workbook_path, args.sheet_name)
    records = read_records(workbook_path, sheet_name)
    proposals = build_proposals(records)

    summary = {
        "workbook": str(workbook_path),
        "sheet_name": sheet_name,
        "proposal_row_count": len(proposals),
        "high_confidence_count": sum(row["proposal_confidence"] == "high" for row in proposals),
        "medium_confidence_count": sum(row["proposal_confidence"] == "medium" for row in proposals),
        "manual_review_count": sum(row["needs_manual_review"] == "yes" for row in proposals),
        "proposals": proposals,
    }

    rendered = json.dumps(summary, ensure_ascii=True, indent=2)
    if args.out_json:
        out_json = Path(args.out_json).expanduser().resolve()
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(rendered + "\n", encoding="utf-8")
    if args.out_csv:
        write_csv(Path(args.out_csv).expanduser().resolve(), proposals)
    print(rendered)


if __name__ == "__main__":
    main()
