#!/usr/bin/env python3
"""Build a decision-ready review packet from Signal_Review_Queue rows."""

from __future__ import annotations

import argparse
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build ingredient signal review packets. Install it in the local Python environment first."
    ) from exc


QUEUE_FIELDS = [
    "priority_score",
    "recommended_bucket",
    "recommended_action",
    "triage_reason",
    "raw_token",
    "normalized_token",
    "unmatched_count",
    "sku_row_count",
    "full_inci_count",
    "key_count",
    "product_only_count",
    "top_categories",
    "example_brands",
    "example_products",
    "example_urls",
    "in_current_master_like",
]

PACKET_FIELDS = [
    "priority_score",
    "raw_token",
    "normalized_token",
    "unmatched_count",
    "sku_row_count",
    "full_inci_count",
    "key_count",
    "top_categories",
    "example_brands",
    "example_products",
    "example_urls",
    "suggested_resolution",
    "suggested_signal_bucket",
    "suggested_signal_key",
    "suggestion_confidence",
    "resolution_rationale",
    "decision",
    "approved_signal_bucket",
    "approved_signal_key",
    "reviewer_notes",
]

ACID_FAMILY_PATTERNS = [
    re.compile(r"^\s*aha\s*$", re.IGNORECASE),
    re.compile(r"^\s*bha\s*$", re.IGNORECASE),
    re.compile(r"^\s*pha\s*$", re.IGNORECASE),
]

FAMILY_PATTERNS = [
    re.compile(r"\bceramides?\b", re.IGNORECASE),
    re.compile(r"\bpeptides?\b", re.IGNORECASE),
    re.compile(r"\bcollagen\b", re.IGNORECASE),
    re.compile(r"\blipid(?:s)?\b", re.IGNORECASE),
    re.compile(r"\bretinoid(?:s)?\b", re.IGNORECASE),
]

MARKETING_PATTERNS = [
    re.compile(r"\bcomplex\b", re.IGNORECASE),
    re.compile(r"\bblend\b", re.IGNORECASE),
    re.compile(r"\bantioxidants?\b", re.IGNORECASE),
]


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    return "".join(ch for ch in raw if ch.isalnum())


def slug_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    raw = re.sub(r"[^a-z0-9]+", "_", raw)
    return raw.strip("_")


def read_queue_rows(path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Queue workbook missing required sheet '{sheet_name}'.")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    missing = [field for field in QUEUE_FIELDS if field not in header]
    if missing:
        raise SystemExit(f"Queue sheet missing required columns: {', '.join(missing)}")
    out: list[dict[str, str]] = []
    for row in rows[1:]:
        if not any(normalize_text(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        out.append({header[index]: normalize_text(padded[index]) for index in range(len(header))})
    return out


def classify_signal(token: str) -> tuple[str, str, str, str]:
    text = normalize_text(token)
    lowered = text.casefold()

    if any(pattern.search(text) for pattern in ACID_FAMILY_PATTERNS):
        bucket = "acid_family_signal"
        return (
            "route_to_signal_dict",
            bucket,
            slug_key(text),
            "Short acid-family umbrella term; should stay out of canonical ingredient rows.",
        )

    if any(pattern.search(text) for pattern in FAMILY_PATTERNS):
        bucket = "ingredient_family_signal"
        return (
            "route_to_signal_dict",
            bucket,
            slug_key(text),
            "Broad ingredient-family term; treat as signal/family grouping rather than canonical ingredient.",
        )

    if any(pattern.search(text) for pattern in MARKETING_PATTERNS) or "™" in text or "®" in text:
        bucket = "marketing_or_blend_signal"
        return (
            "route_to_signal_dict",
            bucket,
            slug_key(text),
            "Looks like a marketing/blend umbrella term rather than a single canonical ingredient.",
        )

    return (
        "route_to_signal_dict",
        "needs_signal_review",
        slug_key(text) or normalize_key(text),
        "Looks signal-like in current triage, but still needs explicit reviewer confirmation on bucketing.",
    )


def confidence_for_bucket(bucket: str) -> str:
    if bucket in {"acid_family_signal", "ingredient_family_signal"}:
        return "high"
    if bucket == "marketing_or_blend_signal":
        return "medium"
    return "low"


def build_packet_rows(queue_rows: list[dict[str, str]]) -> list[dict[str, str]]:
    packet_rows: list[dict[str, str]] = []
    for row in queue_rows:
        resolution, signal_bucket, signal_key, rationale = classify_signal(row.get("raw_token", ""))
        packet_rows.append(
            {
                "priority_score": normalize_text(row.get("priority_score")),
                "raw_token": normalize_text(row.get("raw_token")),
                "normalized_token": normalize_text(row.get("normalized_token")) or normalize_key(row.get("raw_token")),
                "unmatched_count": normalize_text(row.get("unmatched_count")),
                "sku_row_count": normalize_text(row.get("sku_row_count")),
                "full_inci_count": normalize_text(row.get("full_inci_count")),
                "key_count": normalize_text(row.get("key_count")),
                "top_categories": normalize_text(row.get("top_categories")),
                "example_brands": normalize_text(row.get("example_brands")),
                "example_products": normalize_text(row.get("example_products")),
                "example_urls": normalize_text(row.get("example_urls")),
                "suggested_resolution": resolution,
                "suggested_signal_bucket": signal_bucket,
                "suggested_signal_key": signal_key,
                "suggestion_confidence": confidence_for_bucket(signal_bucket),
                "resolution_rationale": rationale,
                "decision": "",
                "approved_signal_bucket": "",
                "approved_signal_key": "",
                "reviewer_notes": "",
            }
        )
    return packet_rows


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    import csv

    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, sheet_name: str, rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(PACKET_FIELDS)
    for row in rows:
        sheet.append([row.get(field, "") for field in PACKET_FIELDS])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a decision-ready review packet from Signal_Review_Queue rows.")
    parser.add_argument("--queue-xlsx", required=True, help="Ingredient unmatched priority queue workbook")
    parser.add_argument("--queue-sheet", default="Signal_Review_Queue", help="Queue workbook sheet to read")
    parser.add_argument("--out-csv", required=True, help="Where to write packet CSV")
    parser.add_argument("--out-json", required=True, help="Where to write packet summary JSON")
    parser.add_argument("--out-xlsx", help="Optional XLSX packet output")
    args = parser.parse_args()

    queue_path = Path(args.queue_xlsx).expanduser().resolve()
    out_csv = Path(args.out_csv).expanduser().resolve()
    out_json = Path(args.out_json).expanduser().resolve()

    queue_rows = read_queue_rows(queue_path, args.queue_sheet)
    packet_rows = build_packet_rows(queue_rows)

    write_csv(out_csv, PACKET_FIELDS, packet_rows)
    if args.out_xlsx:
        write_xlsx(Path(args.out_xlsx).expanduser().resolve(), args.queue_sheet, packet_rows)

    bucket_counts = Counter(row["suggested_signal_bucket"] for row in packet_rows)
    confidence_counts = Counter(row["suggestion_confidence"] for row in packet_rows)
    summary = {
        "queue_workbook": str(queue_path),
        "queue_sheet": args.queue_sheet,
        "row_count": len(packet_rows),
        "suggested_signal_bucket_counts": dict(bucket_counts),
        "confidence_counts": dict(confidence_counts),
        "out_csv": str(out_csv),
        "out_xlsx": str(Path(args.out_xlsx).expanduser().resolve()) if args.out_xlsx else "",
    }

    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
