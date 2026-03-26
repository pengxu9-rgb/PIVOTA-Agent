#!/usr/bin/env python3
"""Apply new-canonical ingredient patch rows to a workbook copy."""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to apply new canonical ingredient patches. Install it in the local Python environment first."
    ) from exc


PREFERRED_TARGET_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]

PAREN_RE = re.compile(r"\([^)]*\)")


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    return "".join(ch for ch in raw if ch.isalnum())


def canonical_name_key(value: Any) -> str:
    return " ".join(normalize_text(value).casefold().split())


def semantic_match_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    raw = PAREN_RE.sub(" ", raw)
    raw = " ".join(raw.split())
    return "".join(ch for ch in raw if ch.isalnum())


def load_patch_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def build_header_index(sheet) -> dict[str, int]:
    header = [normalize_text(cell.value) for cell in sheet[1]]
    return {name: index + 1 for index, name in enumerate(header)}


def resolve_target_sheet_name(workbook, requested_sheet: str | None) -> str:
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise SystemExit(f"Workbook missing requested target sheet '{requested_sheet}'.")
        return requested_sheet
    for sheet_name in PREFERRED_TARGET_SHEETS:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    raise SystemExit(
        f"Workbook did not contain any supported target sheet names ({', '.join(PREFERRED_TARGET_SHEETS)})."
    )


def build_existing_indexes(sheet, header_index: dict[str, int]) -> dict[str, Any]:
    record_ids: dict[str, int] = {}
    normalized_keys: dict[str, dict[str, str]] = {}
    canonical_names: dict[str, dict[str, str]] = {}
    semantic_keys: dict[str, list[dict[str, str]]] = {}

    for row_number in range(2, sheet.max_row + 1):
        record_id = normalize_text(sheet.cell(row=row_number, column=header_index["record_id"]).value)
        canonical = normalize_text(sheet.cell(row=row_number, column=header_index["canonical_inci_name"]).value)
        normalized = normalize_text(sheet.cell(row=row_number, column=header_index["normalized_key"]).value)

        if record_id:
            record_ids[record_id] = row_number
        if normalized:
            normalized_keys[normalize_key(normalized)] = {
                "record_id": record_id,
                "canonical_inci_name": canonical,
            }
        if canonical:
            canonical_names[canonical_name_key(canonical)] = {
                "record_id": record_id,
                "canonical_inci_name": canonical,
            }
            semantic = semantic_match_key(canonical)
            semantic_keys.setdefault(semantic, []).append(
                {
                    "record_id": record_id,
                    "canonical_inci_name": canonical,
                    "normalized_key": normalize_key(normalized or canonical),
                }
            )

    return {
        "record_ids": record_ids,
        "normalized_keys": normalized_keys,
        "canonical_names": canonical_names,
        "semantic_keys": semantic_keys,
    }


def ensure_new_rows_sheet(workbook, header: list[str]):
    if "New_Rows_Only" in workbook.sheetnames:
        return workbook["New_Rows_Only"]
    sheet = workbook.create_sheet("New_Rows_Only")
    sheet.append(header)
    return sheet


def append_patch_row(sheet, header: list[str], patch: dict[str, str]) -> None:
    sheet.append([normalize_text(patch.get(column, "")) for column in header])


def update_readme_sheet(
    workbook,
    applied_count: int,
    patch_csv: Path,
    readme_metric_prefix: str | None,
    readme_note: str | None,
) -> None:
    if "README" not in workbook.sheetnames or applied_count <= 0:
        return

    sheet = workbook["README"]
    metric_rows: dict[str, int] = {}
    for row_number in range(1, sheet.max_row + 1):
        metric = normalize_text(sheet.cell(row=row_number, column=1).value)
        if metric:
            metric_rows[metric] = row_number

    for metric in ["new_canonical_patch_rows", "new_canonical_rows_appended", "merged_total_rows"]:
        row_number = metric_rows.get(metric)
        if not row_number:
            continue
        current_value = sheet.cell(row=row_number, column=2).value
        try:
            numeric = int(current_value or 0)
        except Exception:
            continue
        sheet.cell(row=row_number, column=2).value = numeric + applied_count

    metric_prefix = normalize_text(readme_metric_prefix) or "v2_2_full_inci_priority"
    note_value = normalize_text(readme_note) or "Applied clean Full_INCI priority canonical candidates into a new workbook copy."
    extra_metrics = {
        f"{metric_prefix}_apply_rows": applied_count,
        f"{metric_prefix}_patch_csv": str(patch_csv),
        f"{metric_prefix}_note": note_value,
    }
    for metric, value in extra_metrics.items():
        if metric in metric_rows:
            sheet.cell(row=metric_rows[metric], column=2).value = value
        else:
            sheet.append([metric, value])


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply new canonical ingredient patch rows to a workbook copy.")
    parser.add_argument("--ingredient-xlsx", required=True, help="Source ingredient workbook")
    parser.add_argument("--patch-csv", required=True, help="Apply-ready patch CSV")
    parser.add_argument("--out-xlsx", required=True, help="Path to write the patched workbook copy")
    parser.add_argument("--target-sheet", help="Optional target sheet name")
    parser.add_argument("--out-report-json", help="Optional path for apply report JSON")
    parser.add_argument("--readme-metric-prefix", help="Optional README metric prefix for appended patch metadata")
    parser.add_argument("--readme-note", help="Optional README note for appended patch metadata")
    args = parser.parse_args()

    workbook_path = Path(args.ingredient_xlsx).expanduser().resolve()
    patch_path = Path(args.patch_csv).expanduser().resolve()
    out_xlsx = Path(args.out_xlsx).expanduser().resolve()

    workbook = load_workbook(workbook_path)
    target_sheet_name = resolve_target_sheet_name(workbook, args.target_sheet)
    target_sheet = workbook[target_sheet_name]
    header = [normalize_text(cell.value) for cell in target_sheet[1]]
    header_index = build_header_index(target_sheet)

    required_columns = {"record_id", "canonical_inci_name", "normalized_key"}
    missing_columns = [column for column in required_columns if column not in header_index]
    if missing_columns:
        raise SystemExit(f"Workbook target sheet is missing required columns: {', '.join(missing_columns)}")

    patch_rows = load_patch_rows(patch_path)
    indexes = build_existing_indexes(target_sheet, header_index)
    record_ids = indexes["record_ids"]
    normalized_keys = indexes["normalized_keys"]
    canonical_names = indexes["canonical_names"]
    semantic_keys = indexes["semantic_keys"]

    new_rows_sheet = ensure_new_rows_sheet(workbook, header)

    applied: list[str] = []
    skipped_conflicts: list[dict[str, str]] = []
    skipped_invalid: list[dict[str, str]] = []

    for patch in patch_rows:
        record_id = normalize_text(patch.get("record_id"))
        canonical = normalize_text(patch.get("canonical_inci_name"))
        normalized = normalize_key(patch.get("normalized_key") or canonical)
        semantic = normalize_text(patch.get("semantic_match_key")) or semantic_match_key(canonical)

        if not record_id or not canonical or not normalized:
            skipped_invalid.append(
                {
                    "record_id": record_id,
                    "canonical_inci_name": canonical,
                    "normalized_key": normalized,
                }
            )
            continue

        conflict: dict[str, str] | None = None
        if record_id in record_ids:
            existing = {
                "record_id": record_id,
                "canonical_inci_name": canonical,
            }
            conflict = {
                "record_id": record_id,
                "conflict_type": "record_id_exists",
                "existing_record_id": existing["record_id"],
                "existing_canonical_inci_name": existing["canonical_inci_name"],
            }
        elif normalized in normalized_keys:
            existing = normalized_keys[normalized]
            conflict = {
                "record_id": record_id,
                "conflict_type": "normalized_key_exists",
                "existing_record_id": existing.get("record_id", ""),
                "existing_canonical_inci_name": existing.get("canonical_inci_name", ""),
            }
        elif canonical_name_key(canonical) in canonical_names:
            existing = canonical_names[canonical_name_key(canonical)]
            conflict = {
                "record_id": record_id,
                "conflict_type": "canonical_name_exists",
                "existing_record_id": existing.get("record_id", ""),
                "existing_canonical_inci_name": existing.get("canonical_inci_name", ""),
            }
        else:
            semantic_conflicts = [
                entry
                for entry in semantic_keys.get(semantic, [])
                if normalize_key(entry.get("canonical_inci_name")) != normalized
            ]
            if semantic_conflicts:
                existing = semantic_conflicts[0]
                conflict = {
                    "record_id": record_id,
                    "conflict_type": "semantic_key_exists",
                    "existing_record_id": existing.get("record_id", ""),
                    "existing_canonical_inci_name": existing.get("canonical_inci_name", ""),
                }

        if conflict:
            skipped_conflicts.append(conflict)
            continue

        append_patch_row(target_sheet, header, patch)
        append_patch_row(new_rows_sheet, header, patch)
        applied.append(record_id)

        record_ids[record_id] = target_sheet.max_row
        normalized_keys[normalized] = {
            "record_id": record_id,
            "canonical_inci_name": canonical,
        }
        canonical_names[canonical_name_key(canonical)] = {
            "record_id": record_id,
            "canonical_inci_name": canonical,
        }
        semantic_keys.setdefault(semantic, []).append(
            {
                "record_id": record_id,
                "canonical_inci_name": canonical,
                "normalized_key": normalized,
            }
        )

    update_readme_sheet(
        workbook,
        len(applied),
        patch_path,
        args.readme_metric_prefix,
        args.readme_note,
    )

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_xlsx)

    report = {
        "source_workbook": str(workbook_path),
        "patch_csv": str(patch_path),
        "out_workbook": str(out_xlsx),
        "target_sheet": target_sheet_name,
        "patch_row_count": len(patch_rows),
        "applied_count": len(applied),
        "skipped_conflict_count": len(skipped_conflicts),
        "skipped_invalid_count": len(skipped_invalid),
        "readme_metric_prefix": normalize_text(args.readme_metric_prefix),
        "readme_note": normalize_text(args.readme_note),
        "applied_record_ids": applied,
        "skipped_conflicts": skipped_conflicts,
        "skipped_invalid_rows": skipped_invalid,
    }

    rendered = json.dumps(report, ensure_ascii=True, indent=2)
    if args.out_report_json:
        report_path = Path(args.out_report_json).expanduser().resolve()
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(rendered + "\n", encoding="utf-8")
    print(rendered)


if __name__ == "__main__":
    main()
