#!/usr/bin/env python3
"""Build deterministic alias backfill proposals for ingredient workbook rows.

This script is read-only. It does not modify workbook files or databases.
It proposes conservative alias and alias_quality fills for rows where:

- aliases_common is missing
- alias_quality is missing
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to propose workbook alias backfills. Install it in the local Python environment first."
    ) from exc


PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def is_nonempty(value: Any) -> bool:
    return bool(normalize_text(value))


def parse_multi(value: Any) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split(";") if part and part.strip()]


def resolve_sheet_name(workbook_path: Path, requested_sheet: str | None = None) -> str:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    if requested_sheet:
        if requested_sheet not in sheet_names:
            raise SystemExit(f"Workbook sheet not found: {requested_sheet}")
        return requested_sheet
    for candidate in PREFERRED_INGREDIENT_SHEETS:
        if candidate in sheet_names:
            return candidate
    raise SystemExit(
        "Workbook is missing a supported ingredient sheet. Expected one of: "
        + ", ".join(PREFERRED_INGREDIENT_SHEETS)
    )


def read_records(workbook_path: Path, sheet_name: str) -> list[dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if not any(cell is not None and normalize_text(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        records.append(dict(zip(header, padded)))
    return records


def normalize_key_like(value: str) -> str:
    text = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    return "".join(ch for ch in text if ch.isalnum())


def same_normalized(left: str, right: str) -> bool:
    return bool(left and right and normalize_key_like(left) == normalize_key_like(right))


def dedupe_keep_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        text = normalize_text(value)
        key = text.casefold()
        if not text or key in seen:
            continue
        seen.add(key)
        out.append(text)
    return out


def is_displayable_alias(value: str) -> bool:
    text = normalize_text(value)
    if not text:
        return False
    if re.search(r"[ /,()\-]", text):
        return True
    return not text.islower()


def collect_exact_label_aliases(record: dict[str, Any], canonical: str, display_name: str) -> tuple[list[str], list[str]]:
    aliases: list[str] = []
    sources: list[str] = []

    def add_alias(candidate: str, source: str) -> None:
        text = normalize_text(candidate)
        if not text or text == canonical or text == display_name:
            return
        aliases.append(text)
        sources.append(source)

    for column in ["us_label_name", "eu_label_name"]:
        add_alias(record.get(column), column)

    for column in ["us_label_variants", "eu_label_variants"]:
        for variant in parse_multi(record.get(column)):
            add_alias(variant, column)

    for variant in parse_multi(record.get("parser_variants")):
        if not is_displayable_alias(variant):
            continue
        if variant == canonical or variant == display_name:
            continue
        # Only keep parser variants that are separator/casing variations of the same ingredient.
        if same_normalized(variant, canonical):
            aliases.append(variant)
            sources.append("parser_variants")

    return dedupe_keep_order(aliases), dedupe_keep_order(sources)


def collect_common_aliases(record: dict[str, Any], canonical: str) -> tuple[list[str], list[str]]:
    aliases: list[str] = []
    sources: list[str] = []
    display_name = normalize_text(record.get("canonical_display_name"))
    if display_name and display_name != canonical:
        aliases.append(display_name)
        sources.append("canonical_display_name")
    return dedupe_keep_order(aliases), dedupe_keep_order(sources)


def propose_alias_quality(existing_aliases: list[str], common_aliases: list[str], exact_aliases: list[str]) -> tuple[str, str]:
    if common_aliases and exact_aliases:
        return "mixed_review_required", "common+exact_label_sources"
    if common_aliases:
        return "common_alias", "canonical_display_name"
    if exact_aliases:
        return "exact_label_alias", "label_or_parser_variant"

    if existing_aliases:
        return "manual_review_required", "existing_aliases_without_deterministic_quality"
    return "", ""


def build_proposals(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    proposals: list[dict[str, Any]] = []

    for record in records:
        canonical = normalize_text(record.get("canonical_inci_name"))
        display_name = normalize_text(record.get("canonical_display_name"))
        existing_aliases = dedupe_keep_order(parse_multi(record.get("aliases_common")))
        existing_alias_quality = normalize_text(record.get("alias_quality"))

        if existing_aliases and existing_alias_quality:
            continue

        common_aliases, common_sources = collect_common_aliases(record, canonical)
        exact_aliases, exact_sources = collect_exact_label_aliases(record, canonical, display_name)

        suggested_aliases = existing_aliases or dedupe_keep_order(common_aliases + exact_aliases)
        suggested_alias_quality, quality_reason = propose_alias_quality(existing_aliases, common_aliases, exact_aliases)

        if existing_alias_quality:
            suggested_alias_quality = existing_alias_quality
            quality_reason = "already_present"

        proposal_confidence = "none"
        if suggested_alias_quality in {"common_alias", "exact_label_alias"} and suggested_aliases:
            proposal_confidence = "high"
        elif suggested_alias_quality == "mixed_review_required" or suggested_aliases:
            proposal_confidence = "medium"

        needs_manual_review = not suggested_aliases or suggested_alias_quality in {"", "manual_review_required", "mixed_review_required"}

        proposals.append(
            {
                "record_id": normalize_text(record.get("record_id")),
                "canonical_inci_name": canonical,
                "canonical_display_name": display_name,
                "existing_aliases_common": "; ".join(existing_aliases),
                "existing_alias_quality": existing_alias_quality,
                "suggested_aliases_common": "; ".join(suggested_aliases),
                "suggested_alias_quality": suggested_alias_quality,
                "proposal_confidence": proposal_confidence,
                "proposal_sources": "; ".join(dedupe_keep_order(common_sources + exact_sources)),
                "quality_reason": quality_reason,
                "needs_manual_review": "yes" if needs_manual_review else "no",
            }
        )

    return proposals


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "record_id",
        "canonical_inci_name",
        "canonical_display_name",
        "existing_aliases_common",
        "existing_alias_quality",
        "suggested_aliases_common",
        "suggested_alias_quality",
        "proposal_confidence",
        "proposal_sources",
        "quality_reason",
        "needs_manual_review",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Propose deterministic alias backfills for the ingredient workbook.")
    parser.add_argument("--ingredient-xlsx", required=True, help="Path to the ingredient reference workbook")
    parser.add_argument("--sheet-name", help="Optional workbook sheet name")
    parser.add_argument("--out-json", help="Optional path to write JSON summary")
    parser.add_argument("--out-csv", help="Optional path to write proposal CSV")
    args = parser.parse_args()

    workbook_path = Path(args.ingredient_xlsx).expanduser().resolve()
    sheet_name = resolve_sheet_name(workbook_path, args.sheet_name)
    records = read_records(workbook_path, sheet_name)
    proposals = build_proposals(records)

    summary = {
        "workbook": str(workbook_path),
        "sheet_name": sheet_name,
        "proposal_row_count": len(proposals),
        "high_confidence_count": sum(row["proposal_confidence"] == "high" for row in proposals),
        "medium_confidence_count": sum(row["proposal_confidence"] == "medium" for row in proposals),
        "manual_review_count": sum(row["needs_manual_review"] == "yes" for row in proposals),
        "proposals": proposals,
    }

    rendered = json.dumps(summary, ensure_ascii=True, indent=2)
    if args.out_json:
        out_json = Path(args.out_json).expanduser().resolve()
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(rendered + "\n", encoding="utf-8")
    if args.out_csv:
        write_csv(Path(args.out_csv).expanduser().resolve(), proposals)
    print(rendered)


if __name__ == "__main__":
    main()
