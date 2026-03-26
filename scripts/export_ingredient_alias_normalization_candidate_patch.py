#!/usr/bin/env python3
"""Export new-canonical candidate patch rows from alias normalization packet."""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to export alias normalization candidate patches. Install it in the local Python environment first."
    ) from exc


PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]

RECORD_ID_PATTERN = re.compile(r"^ing_patch_v13_(\d+)$", re.IGNORECASE)
PAREN_RE = re.compile(r"\([^)]*\)")


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    return "".join(ch for ch in raw if ch.isalnum())


def semantic_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    raw = PAREN_RE.sub(" ", raw)
    raw = " ".join(raw.split())
    return "".join(ch for ch in raw if ch.isalnum())


def split_semicolon_values(value: Any) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    return [part.strip() for part in raw.split(";") if part and part.strip()]


def dedupe_join(values: list[str]) -> str:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        text = normalize_text(value)
        key = text.casefold()
        if not text or key in seen:
            continue
        seen.add(key)
        out.append(text)
    return "; ".join(out)


def resolve_ingredient_sheet_name(path: Path, requested_sheet: str | None) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise SystemExit(f"Ingredient workbook missing requested sheet '{requested_sheet}'.")
        return requested_sheet
    for sheet_name in PREFERRED_INGREDIENT_SHEETS:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    raise SystemExit(
        f"Ingredient workbook did not contain any supported sheet names ({', '.join(PREFERRED_INGREDIENT_SHEETS)})."
    )


def load_packet_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def read_ingredient_header_and_max_patch(path: Path, sheet_name: str) -> tuple[list[str], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    max_patch = 0
    record_id_index = header.index("record_id")
    for row in rows[1:]:
        record_id = normalize_text(row[record_id_index] if record_id_index < len(row) else "")
        match = RECORD_ID_PATTERN.match(record_id)
        if match:
            max_patch = max(max_patch, int(match.group(1)))
    return header, max_patch


def set_if_present(row: dict[str, str], header: list[str], key: str, value: Any) -> None:
    if key in header:
        row[key] = normalize_text(value)


def build_patch_row(
    header: list[str],
    packet_row: dict[str, str],
    record_id: str,
    canonical_name: str,
) -> dict[str, str]:
    row = {column: "" for column in header}
    raw_token = normalize_text(packet_row.get("raw_token"))
    normalized = normalize_key(packet_row.get("normalized_token") or canonical_name)
    parser_variants = dedupe_join(
        [canonical_name]
        + split_semicolon_values(packet_row.get("suggested_parser_variants_addition"))
        + ([raw_token] if raw_token else [])
    )

    set_if_present(row, header, "record_id", record_id)
    set_if_present(row, header, "canonical_inci_name", canonical_name)
    set_if_present(row, header, "canonical_display_name", canonical_name)
    set_if_present(row, header, "ingredient_family", "other")
    set_if_present(row, header, "us_label_name", canonical_name)
    set_if_present(row, header, "eu_label_name", canonical_name)
    set_if_present(row, header, "us_label_variants", canonical_name)
    set_if_present(row, header, "eu_label_variants", canonical_name)
    set_if_present(row, header, "cross_market_notes", "Alias normalization candidate patch from v2.2 queue")
    set_if_present(row, header, "normalized_key", normalized)
    set_if_present(row, header, "aliases_common", "")
    set_if_present(row, header, "parser_variants", parser_variants)
    set_if_present(row, header, "deprecated_aliases", "")
    set_if_present(row, header, "alias_quality", "")
    set_if_present(
        row,
        header,
        "notes_for_parser",
        "Generated from Alias_Normalization_Queue; preserve case variants, label forms, and slash-separated forms in parser_variants.",
    )
    set_if_present(row, header, "primary_bucket", "")
    set_if_present(row, header, "all_buckets", "")
    set_if_present(row, header, "function_tags", "")
    set_if_present(row, header, "benefit_tags", "")
    set_if_present(row, header, "risk_flags", "")
    for boolean_column in [
        "is_humectant",
        "is_barrier_support",
        "is_retinoid",
        "is_exfoliant",
        "is_uv_filter",
        "is_preservative",
        "is_surfactant",
        "is_fragrance_or_eo",
    ]:
        set_if_present(row, header, boolean_column, "no")
    set_if_present(row, header, "regulatory_bucket", "patch_candidate_review")
    set_if_present(row, header, "source_urls", packet_row.get("example_urls", ""))
    set_if_present(row, header, "source_authorities", "brand_official_pdp")
    set_if_present(row, header, "source_types", "official_brand_site")
    set_if_present(row, header, "review_status", "draft")
    confidence = normalize_text(packet_row.get("suggestion_confidence")) or "low"
    set_if_present(row, header, "confidence", confidence)
    set_if_present(row, header, "last_reviewed_at", "")
    set_if_present(
        row,
        header,
        "review_notes",
        "Generated from Alias_Normalization_Queue packet; verify canonical naming and parser variants before promotion.",
    )
    set_if_present(
        row,
        header,
        "notes",
        dedupe_join(
            [
                f"priority_score={packet_row.get('priority_score', '')}",
                f"unmatched_count={packet_row.get('unmatched_count', '')}",
                f"sku_row_count={packet_row.get('sku_row_count', '')}",
                f"full_inci_count={packet_row.get('full_inci_count', '')}",
                f"brands={packet_row.get('example_brands', '')}",
            ]
        ),
    )
    set_if_present(row, header, "kb_version", "spec_v1_patch_v13_alias_normalization")

    row["queue_priority_score"] = normalize_text(packet_row.get("priority_score"))
    row["queue_example_brands"] = normalize_text(packet_row.get("example_brands"))
    row["queue_example_products"] = normalize_text(packet_row.get("example_products"))
    row["queue_example_urls"] = normalize_text(packet_row.get("example_urls"))
    row["queue_top_categories"] = normalize_text(packet_row.get("top_categories"))
    row["source_packet_resolution"] = normalize_text(packet_row.get("suggested_resolution"))
    row["source_packet_confidence"] = normalize_text(packet_row.get("suggestion_confidence"))
    row["source_packet_raw_token"] = raw_token
    row["semantic_match_key"] = semantic_key(canonical_name)
    return row


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export new-canonical candidate patch rows from alias normalization packet.")
    parser.add_argument("--packet-csv", required=True, help="Alias normalization packet CSV")
    parser.add_argument("--ingredient-xlsx", required=True, help="Current ingredient workbook")
    parser.add_argument("--ingredient-sheet", help="Optional ingredient workbook sheet")
    parser.add_argument("--out-apply-csv", required=True, help="Where to write apply-ready candidate patch CSV")
    parser.add_argument("--out-remainder-csv", required=True, help="Where to write non-exported packet rows")
    parser.add_argument("--out-summary-json", required=True, help="Where to write summary JSON")
    args = parser.parse_args()

    packet_csv = Path(args.packet_csv).expanduser().resolve()
    ingredient_xlsx = Path(args.ingredient_xlsx).expanduser().resolve()
    out_apply_csv = Path(args.out_apply_csv).expanduser().resolve()
    out_remainder_csv = Path(args.out_remainder_csv).expanduser().resolve()
    out_summary_json = Path(args.out_summary_json).expanduser().resolve()

    ingredient_sheet = resolve_ingredient_sheet_name(ingredient_xlsx, args.ingredient_sheet)
    header, max_patch = read_ingredient_header_and_max_patch(ingredient_xlsx, ingredient_sheet)
    packet_rows = load_packet_rows(packet_csv)

    apply_rows: list[dict[str, str]] = []
    remainder_rows: list[dict[str, str]] = []
    next_patch_number = max_patch + 1

    for packet_row in packet_rows:
        resolution = normalize_text(packet_row.get("suggested_resolution"))
        confidence = normalize_text(packet_row.get("suggestion_confidence"))
        existing_target = normalize_text(packet_row.get("existing_target_record_id"))
        canonical_name = normalize_text(packet_row.get("approved_new_canonical_inci_name") or packet_row.get("suggested_new_canonical_inci_name"))

        if (
            resolution == "new_canonical_candidate_with_parser_variants"
            and confidence in {"high", "medium"}
            and not existing_target
            and canonical_name
        ):
            record_id = f"ing_patch_v13_{next_patch_number:03d}"
            next_patch_number += 1
            apply_rows.append(build_patch_row(header, packet_row, record_id, canonical_name))
            continue

        remainder_rows.append(packet_row)

    apply_fieldnames = header + [
        "queue_priority_score",
        "queue_example_brands",
        "queue_example_products",
        "queue_example_urls",
        "queue_top_categories",
        "source_packet_resolution",
        "source_packet_confidence",
        "source_packet_raw_token",
        "semantic_match_key",
    ]
    remainder_fieldnames = list(packet_rows[0].keys()) if packet_rows else []

    write_csv(out_apply_csv, apply_fieldnames, apply_rows)
    write_csv(out_remainder_csv, remainder_fieldnames, remainder_rows)

    summary = {
        "packet_csv": str(packet_csv),
        "ingredient_workbook": str(ingredient_xlsx),
        "ingredient_sheet": ingredient_sheet,
        "input_row_count": len(packet_rows),
        "apply_ready_count": len(apply_rows),
        "remainder_count": len(remainder_rows),
        "record_id_start": apply_rows[0]["record_id"] if apply_rows else None,
        "record_id_end": apply_rows[-1]["record_id"] if apply_rows else None,
        "out_apply_csv": str(out_apply_csv),
        "out_remainder_csv": str(out_remainder_csv),
    }

    out_summary_json.parent.mkdir(parents=True, exist_ok=True)
    out_summary_json.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
