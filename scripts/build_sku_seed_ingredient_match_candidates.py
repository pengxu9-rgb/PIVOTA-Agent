#!/usr/bin/env python3
"""Build deterministic SKU-seed x ingredient-reference candidate matches.

This script is intentionally conservative:
- it never writes to a database
- it only emits exact normalized matches against reviewed ingredient reference terms
- it is designed to create a reviewable candidate layer before harvester / KB ingest
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build SKU seed ingredient match candidates. Install it in the local Python environment first."
    ) from exc


SKU_REQUIRED_COLUMNS = [
    "brand_name",
    "product_name",
    "official_product_url",
    "market",
    "category",
    "ingredient_granularity",
    "ingredients_or_key_ingredients",
    "extraction_status",
]

INGREDIENT_REQUIRED_COLUMNS = [
    "record_id",
    "canonical_inci_name",
    "canonical_display_name",
    "normalized_key",
    "parser_variants",
]

PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]

PAREN_RE = re.compile(r"\s*\([^)]*\)\s*")


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    return "".join(ch for ch in raw if ch.isalnum())


def split_semicolon_values(value: Any) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    return [part.strip() for part in raw.split(";") if part and part.strip()]


def sha1_text(value: str, length: int = 20) -> str:
    return hashlib.sha1(value.encode("utf-8")).hexdigest()[:length]


def resolve_sheet_name(path: Path, requested_sheet: str | None, preferred_sheets: list[str]) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise SystemExit(f"Workbook missing required '{requested_sheet}' sheet: {path}")
        return requested_sheet
    for sheet_name in preferred_sheets:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    raise SystemExit(f"Workbook missing any supported sheet ({', '.join(preferred_sheets)}): {path}")


def read_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Workbook missing required '{sheet_name}' sheet: {path}")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"Workbook sheet '{sheet_name}' is empty: {path}")

    header = [normalize_text(cell) for cell in rows[0]]
    records: list[tuple[int, dict[str, str]]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(normalize_text(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        record = {
            key: normalize_text(padded[index] if index < len(padded) else "")
            for index, key in enumerate(header)
        }
        records.append((row_number, record))
    return header, records


def assert_required_columns(header: list[str], required: list[str], label: str) -> None:
    missing = [column for column in required if column not in header]
    if missing:
        raise SystemExit(f"{label} missing required columns: {', '.join(missing)}")


def build_reference_lookup(records: list[tuple[int, dict[str, str]]]) -> dict[str, list[dict[str, Any]]]:
    lookup: dict[str, list[dict[str, Any]]] = defaultdict(list)

    term_specs = [
        ("canonical_inci_name", "canonical_inci_name", 100),
        ("canonical_display_name", "canonical_display_name", 98),
        ("us_label_name", "us_label_name", 96),
        ("eu_label_name", "eu_label_name", 96),
        ("parser_variants", "parser_variant", 94),
        ("aliases_common", "common_alias", 90),
        ("us_label_variants", "us_label_variant", 88),
        ("eu_label_variants", "eu_label_variant", 88),
        ("deprecated_aliases", "deprecated_alias", 70),
    ]

    for source_row_number, record in records:
        base = {
            "record_id": record.get("record_id", ""),
            "canonical_inci_name": record.get("canonical_inci_name", ""),
            "canonical_display_name": record.get("canonical_display_name", ""),
            "ingredient_family": record.get("ingredient_family", ""),
            "primary_bucket": record.get("primary_bucket", ""),
            "normalized_key": record.get("normalized_key", ""),
            "source_row_number": source_row_number,
        }
        normalized_record_key = normalize_key(record.get("normalized_key") or record.get("canonical_inci_name"))
        if normalized_record_key:
            lookup[normalized_record_key].append(
                {
                    **base,
                    "term_type": "normalized_key",
                    "matched_term": record.get("normalized_key") or record.get("canonical_inci_name", ""),
                    "score": 110,
                }
            )

        for field_name, term_type, score in term_specs:
            values = split_semicolon_values(record.get(field_name)) if field_name.endswith("variants") or field_name in {"parser_variants", "aliases_common", "deprecated_aliases"} else [record.get(field_name, "")]
            for value in values:
                normalized = normalize_key(value)
                if not normalized:
                    continue
                lookup[normalized].append(
                    {
                        **base,
                        "term_type": term_type,
                        "matched_term": value,
                        "score": score,
                    }
                )

    for normalized_key, options in lookup.items():
        deduped: dict[tuple[str, str], dict[str, Any]] = {}
        for option in sorted(options, key=lambda row: (-int(row["score"]), row["record_id"], row["term_type"])):
            deduped.setdefault((option["record_id"], option["term_type"]), option)
        lookup[normalized_key] = list(deduped.values())

    return lookup


def build_token_variants(raw_token: str) -> list[dict[str, str]]:
    variants: list[dict[str, str]] = []

    def push(value: str, variant_type: str) -> None:
        text = normalize_text(value)
        normalized = normalize_key(text)
        if not text or not normalized:
            return
        if any(row["candidate_normalized"] == normalized and row["variant_type"] == variant_type for row in variants):
            return
        variants.append(
            {
                "candidate_text": text,
                "candidate_normalized": normalized,
                "variant_type": variant_type,
            }
        )

    push(raw_token, "raw_token")

    paren_stripped = PAREN_RE.sub(" ", raw_token).strip()
    if paren_stripped and paren_stripped != raw_token:
        push(paren_stripped, "paren_stripped")

    if "/" in raw_token:
        for part in re.split(r"\s*/\s*", raw_token):
            push(part, "slash_variant")

    return variants


def split_seed_tokens(value: Any) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    tokens: list[str] = []
    current: list[str] = []
    depth = 0
    for ch in raw:
        if ch == "(":
            depth += 1
        elif ch == ")" and depth > 0:
            depth -= 1
        if ch in {",", ";"} and depth == 0:
            token = "".join(current).strip()
            if token and not normalize_key(token).isdigit():
                tokens.append(token)
            current = []
            continue
        current.append(ch)

    tail = "".join(current).strip()
    if tail and not normalize_key(tail).isdigit():
        tokens.append(tail)
    return tokens


def choose_best_option(candidates: list[dict[str, Any]]) -> tuple[dict[str, Any] | None, bool]:
    if not candidates:
        return None, False
    ranked = sorted(
        candidates,
        key=lambda row: (
            -int(row["score"]),
            row["record_id"],
            row["term_type"],
            row["variant_type"],
        ),
    )
    top = ranked[0]
    top_score = int(top["score"])
    top_record_ids = {row["record_id"] for row in ranked if int(row["score"]) == top_score}
    is_ambiguous = len(top_record_ids) > 1
    return top, is_ambiguous


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def main() -> None:
    parser = argparse.ArgumentParser(description="Build deterministic SKU seed x ingredient reference candidate matches.")
    parser.add_argument("--sku-xlsx", required=True, help="Brand SKU inventory workbook path")
    parser.add_argument("--ingredient-xlsx", required=True, help="Ingredient reference workbook path")
    parser.add_argument("--sku-sheet", default="SKU_Seed_Inventory", help="SKU workbook sheet to read")
    parser.add_argument("--ingredient-sheet", help="Ingredient workbook sheet to read")
    parser.add_argument("--out-json", required=True, help="Summary/report JSON output path")
    parser.add_argument("--out-csv", required=True, help="Row-level candidate CSV output path")
    args = parser.parse_args()

    sku_path = Path(args.sku_xlsx).expanduser().resolve()
    ingredient_path = Path(args.ingredient_xlsx).expanduser().resolve()
    out_json = Path(args.out_json).expanduser().resolve()
    out_csv = Path(args.out_csv).expanduser().resolve()

    ingredient_sheet = resolve_sheet_name(ingredient_path, args.ingredient_sheet, PREFERRED_INGREDIENT_SHEETS)

    sku_header, sku_records = read_sheet(sku_path, args.sku_sheet)
    ingredient_header, ingredient_records = read_sheet(ingredient_path, ingredient_sheet)
    assert_required_columns(sku_header, SKU_REQUIRED_COLUMNS, "SKU workbook")
    assert_required_columns(ingredient_header, INGREDIENT_REQUIRED_COLUMNS, "Ingredient workbook")

    reference_lookup = build_reference_lookup(ingredient_records)
    exported_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")

    candidate_rows: list[dict[str, Any]] = []
    matched_counter: Counter[str] = Counter()
    match_status_counter: Counter[str] = Counter()
    granularity_status_counter: Counter[str] = Counter()
    rows_with_any_match = 0

    for source_row_number, sku_row in sku_records:
        raw_ingredient_text = sku_row.get("ingredients_or_key_ingredients", "")
        raw_tokens = split_seed_tokens(raw_ingredient_text)
        if not raw_tokens:
            continue

        row_key = sha1_text(
            "|".join(
                [
                    sku_path.name,
                    args.sku_sheet,
                    str(source_row_number),
                    sku_row.get("brand_name", ""),
                    sku_row.get("product_name", ""),
                    sku_row.get("official_product_url", ""),
                ]
            ),
            length=16,
        )
        row_has_match = False

        for token_index, raw_token in enumerate(raw_tokens, start=1):
            variants = build_token_variants(raw_token)
            options: list[dict[str, Any]] = []
            for variant in variants:
                for option in reference_lookup.get(variant["candidate_normalized"], []):
                    options.append(
                        {
                            **option,
                            "candidate_text": variant["candidate_text"],
                            "candidate_normalized": variant["candidate_normalized"],
                            "variant_type": variant["variant_type"],
                        }
                    )

            best, is_ambiguous = choose_best_option(options)
            if best and not is_ambiguous:
                match_status = "matched"
                row_has_match = True
                matched_counter[best["canonical_inci_name"]] += 1
            elif is_ambiguous:
                match_status = "ambiguous"
            else:
                match_status = "unmatched"

            match_status_counter[match_status] += 1
            granularity = sku_row.get("ingredient_granularity", "")
            granularity_status_counter[f"{granularity}::{match_status}"] += 1

            candidate_rows.append(
                {
                    "candidate_match_key": sha1_text(
                        f"{row_key}|{token_index}|{raw_token}|{best['record_id'] if best else normalize_key(raw_token)}"
                    ),
                    "sku_row_key": row_key,
                    "source_file": sku_path.name,
                    "source_sheet": args.sku_sheet,
                    "source_row_number": source_row_number,
                    "brand_name": sku_row.get("brand_name", ""),
                    "product_name": sku_row.get("product_name", ""),
                    "official_product_url": sku_row.get("official_product_url", ""),
                    "market": sku_row.get("market", ""),
                    "category": sku_row.get("category", ""),
                    "ingredient_granularity": granularity,
                    "extraction_status": sku_row.get("extraction_status", ""),
                    "raw_ingredient_text": raw_ingredient_text,
                    "raw_token": raw_token,
                    "token_index": token_index,
                    "token_normalized": normalize_key(raw_token),
                    "matched_input": best["candidate_text"] if best else "",
                    "matched_input_normalized": best["candidate_normalized"] if best else "",
                    "matched_input_variant_type": best["variant_type"] if best else "",
                    "match_status": match_status,
                    "match_method": best["term_type"] if best and not is_ambiguous else "",
                    "match_confidence": "high" if best and not is_ambiguous else "review" if is_ambiguous else "none",
                    "ingredient_record_id": best["record_id"] if best and not is_ambiguous else "",
                    "canonical_inci_name": best["canonical_inci_name"] if best and not is_ambiguous else "",
                    "canonical_display_name": best["canonical_display_name"] if best and not is_ambiguous else "",
                    "ingredient_normalized_key": best["normalized_key"] if best and not is_ambiguous else "",
                    "ingredient_family": best["ingredient_family"] if best and not is_ambiguous else "",
                    "primary_bucket": best["primary_bucket"] if best and not is_ambiguous else "",
                    "matched_reference_term": best["matched_term"] if best and not is_ambiguous else "",
                    "ambiguity_record_ids": ";".join(
                        sorted({row["record_id"] for row in options if int(row["score"]) == int(best["score"])})
                    )
                    if best and is_ambiguous
                    else "",
                    "exported_at": exported_at,
                }
            )

        if row_has_match:
            rows_with_any_match += 1

    csv_fieldnames = [
        "candidate_match_key",
        "sku_row_key",
        "source_file",
        "source_sheet",
        "source_row_number",
        "brand_name",
        "product_name",
        "official_product_url",
        "market",
        "category",
        "ingredient_granularity",
        "extraction_status",
        "raw_ingredient_text",
        "raw_token",
        "token_index",
        "token_normalized",
        "matched_input",
        "matched_input_normalized",
        "matched_input_variant_type",
        "match_status",
        "match_method",
        "match_confidence",
        "ingredient_record_id",
        "canonical_inci_name",
        "canonical_display_name",
        "ingredient_normalized_key",
        "ingredient_family",
        "primary_bucket",
        "matched_reference_term",
        "ambiguity_record_ids",
        "exported_at",
    ]
    write_csv(out_csv, csv_fieldnames, candidate_rows)

    summary = {
        "sku_workbook": str(sku_path),
        "ingredient_workbook": str(ingredient_path),
        "sku_sheet": args.sku_sheet,
        "ingredient_sheet": ingredient_sheet,
        "row_count": len(sku_records),
        "token_count": len(candidate_rows),
        "matched_token_count": match_status_counter["matched"],
        "ambiguous_token_count": match_status_counter["ambiguous"],
        "unmatched_token_count": match_status_counter["unmatched"],
        "sku_rows_with_any_match": rows_with_any_match,
        "match_status_counts": dict(match_status_counter),
        "granularity_status_counts": dict(granularity_status_counter),
        "top_canonical_matches": dict(matched_counter.most_common(25)),
        "recommended_target": {
            "table": "seed_preview.sku_ingredient_reference_match_candidates",
            "role": "deterministic reviewable join between SKU seed rows and ingredient reference rows",
            "ready_for_direct_runtime_use": False,
        },
        "promotion_path": [
            "keep workbook rows in seed_ingest.sku_seed_inventory_seed and seed_ingest.ingredient_reference_seed",
            "generate deterministic match candidates into a reviewable candidate layer",
            "review ambiguous and unmatched key-ingredient rows before using the candidate layer for harvest planning",
            "prefer harvested/reviewed full INCI evidence for pci_kb.sku_ingredients ingest",
            "publish a runtime-facing projection only after reviewed SKU ingredient evidence exists",
        ],
        "out_csv": str(out_csv),
        "exported_at": exported_at,
    }

    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
