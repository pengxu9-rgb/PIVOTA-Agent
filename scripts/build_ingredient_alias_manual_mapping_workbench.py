#!/usr/bin/env python3
"""Enrich alias manual-mapping packet rows with deterministic target suggestions."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build ingredient alias manual-mapping workbenches. Install it in the local Python environment first."
    ) from exc


PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]

PAREN_RE = re.compile(r"\(([^)]*)\)")

WORKBENCH_FIELDS = [
    "priority_score",
    "raw_token",
    "normalized_token",
    "manual_mapping_subtype",
    "suggested_new_canonical_inci_name",
    "suggested_parser_variants_addition",
    "example_brands",
    "example_products",
    "example_urls",
    "resolution_rationale",
    "suggested_existing_target_record_id",
    "suggested_existing_target_canonical_inci_name",
    "suggested_existing_aliases_common",
    "suggested_existing_parser_variants",
    "suggested_existing_alias_quality",
    "suggested_decision",
    "suggested_alias_quality",
    "suggestion_confidence",
    "suggestion_rationale",
    "decision",
    "approved_existing_target_record_id",
    "approved_existing_target_canonical_inci_name",
    "approved_new_canonical_inci_name",
    "approved_parser_variants_addition",
    "approved_alias_quality",
    "reviewer_notes",
]


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: Any) -> str:
    return "".join(ch for ch in normalize_text(value).lower() if ch.isalnum())


def load_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def resolve_ingredient_sheet_name(path: Path, requested_sheet: str | None) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise SystemExit(f"Ingredient workbook missing requested sheet '{requested_sheet}'.")
        return requested_sheet
    for sheet_name in PREFERRED_INGREDIENT_SHEETS:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    raise SystemExit(
        f"Ingredient workbook did not contain any supported sheet names ({', '.join(PREFERRED_INGREDIENT_SHEETS)})."
    )


def build_canonical_index(path: Path, sheet_name: str) -> dict[str, list[dict[str, str]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    idx = {name: i for i, name in enumerate(header)}

    out: dict[str, list[dict[str, str]]] = {}
    for row in rows[1:]:
        if not row:
            continue
        canon = normalize_text(row[idx["canonical_inci_name"]]) if "canonical_inci_name" in idx else ""
        if not canon:
            continue
        key = normalize_key(canon)
        out.setdefault(key, []).append(
            {
                "record_id": normalize_text(row[idx["record_id"]]) if "record_id" in idx else "",
                "canonical_inci_name": canon,
                "aliases_common": normalize_text(row[idx["aliases_common"]]) if "aliases_common" in idx else "",
                "parser_variants": normalize_text(row[idx["parser_variants"]]) if "parser_variants" in idx else "",
                "alias_quality": normalize_text(row[idx["alias_quality"]]) if "alias_quality" in idx else "",
            }
        )
    return out


def merge_alias_value(existing_aliases: str, raw_token: str) -> str:
    values: list[str] = []
    seen: set[str] = set()
    for part in [p.strip() for p in normalize_text(existing_aliases).split(";") if p.strip()] + [normalize_text(raw_token)]:
        key = part.casefold()
        if not part or key in seen:
            continue
        seen.add(key)
        values.append(part)
    return "; ".join(values)


def suggest_existing_target(row: dict[str, str], canonical_index: dict[str, list[dict[str, str]]]) -> tuple[dict[str, str], str, str, str, str]:
    raw_token = normalize_text(row.get("raw_token"))
    subtype = normalize_text(row.get("manual_mapping_subtype"))
    suggested_new = normalize_text(row.get("suggested_new_canonical_inci_name"))

    # Safe rule: CI tokens that include a unique existing canonical inside parentheses.
    if subtype == "ci_color_index_token":
        match = PAREN_RE.search(raw_token)
        if match:
            inner = normalize_text(match.group(1))
            targets = canonical_index.get(normalize_key(inner), [])
            if len(targets) == 1:
                target = targets[0]
                return (
                    target,
                    "map_to_existing_canonical",
                    "exact_label_alias",
                    "high",
                    "Parenthetical CI token resolves uniquely to an existing canonical ingredient; safe to map as exact label alias.",
                )

    # Safe rule: slash variants can map when the left side is already an existing unique canonical.
    if subtype == "bilingual_or_slash_label_variant":
        left = normalize_text(raw_token.split("/")[0])
        targets = canonical_index.get(normalize_key(left), [])
        if len(targets) == 1:
            target = targets[0]
            return (
                target,
                "map_to_existing_canonical",
                "exact_label_alias",
                "high",
                "Slash-separated label variant resolves uniquely to existing canonical via left-side label form.",
            )

    # Default: keep as create-new-canonical candidate for reviewer judgment.
    return (
        {
            "record_id": "",
                "canonical_inci_name": "",
                "aliases_common": "",
                "parser_variants": "",
                "alias_quality": "",
            },
        "create_new_canonical",
        "",
        "low",
        "No deterministic existing canonical target found; reviewer must decide whether to create a new canonical or keep as alias/signal only.",
    )


def build_rows(packet_rows: list[dict[str, str]], canonical_index: dict[str, list[dict[str, str]]]) -> list[dict[str, str]]:
    out: list[dict[str, str]] = []
    for row in packet_rows:
        target, suggested_decision, suggested_alias_quality, confidence, suggestion_rationale = suggest_existing_target(row, canonical_index)
        out.append(
            {
                "priority_score": normalize_text(row.get("priority_score")),
                "raw_token": normalize_text(row.get("raw_token")),
                "normalized_token": normalize_text(row.get("normalized_token")),
                "manual_mapping_subtype": normalize_text(row.get("manual_mapping_subtype")),
                "suggested_new_canonical_inci_name": normalize_text(row.get("suggested_new_canonical_inci_name")),
                "suggested_parser_variants_addition": normalize_text(row.get("suggested_parser_variants_addition")),
                "example_brands": normalize_text(row.get("example_brands")),
                "example_products": normalize_text(row.get("example_products")),
                "example_urls": normalize_text(row.get("example_urls")),
                "resolution_rationale": normalize_text(row.get("resolution_rationale")),
                "suggested_existing_target_record_id": normalize_text(target.get("record_id")),
                "suggested_existing_target_canonical_inci_name": normalize_text(target.get("canonical_inci_name")),
                "suggested_existing_aliases_common": normalize_text(target.get("aliases_common")),
                "suggested_existing_parser_variants": normalize_text(target.get("parser_variants")),
                "suggested_existing_alias_quality": normalize_text(target.get("alias_quality")),
                "suggested_decision": suggested_decision,
                "suggested_alias_quality": suggested_alias_quality,
                "suggestion_confidence": confidence,
                "suggestion_rationale": suggestion_rationale,
                "decision": "",
                "approved_existing_target_record_id": normalize_text(row.get("approved_existing_target_record_id")),
                "approved_existing_target_canonical_inci_name": normalize_text(row.get("approved_existing_target_canonical_inci_name")),
                "approved_new_canonical_inci_name": normalize_text(row.get("approved_new_canonical_inci_name")),
                "approved_parser_variants_addition": normalize_text(row.get("approved_parser_variants_addition")),
                "approved_alias_quality": normalize_text(row.get("approved_alias_quality")),
                "reviewer_notes": normalize_text(row.get("reviewer_notes")),
            }
        )
    return out


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=WORKBENCH_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, sheet_name: str, rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = sheet_name
    sheet.append(WORKBENCH_FIELDS)
    for row in rows:
        sheet.append([row.get(field, "") for field in WORKBENCH_FIELDS])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Enrich alias manual-mapping packet rows with deterministic target suggestions.")
    parser.add_argument("--packet-csv", required=True, help="Alias manual-mapping packet CSV")
    parser.add_argument("--ingredient-xlsx", required=True, help="Current ingredient workbook")
    parser.add_argument("--ingredient-sheet", help="Optional ingredient workbook sheet")
    parser.add_argument("--out-csv", required=True, help="Where to write the enriched workbench CSV")
    parser.add_argument("--out-json", required=True, help="Where to write summary JSON")
    parser.add_argument("--out-xlsx", help="Optional XLSX output")
    parser.add_argument("--sheet-name", default="Alias_Manual_Mapping_Workbench", help="Optional XLSX sheet name")
    args = parser.parse_args()

    packet_csv = Path(args.packet_csv).expanduser().resolve()
    ingredient_xlsx = Path(args.ingredient_xlsx).expanduser().resolve()
    ingredient_sheet = resolve_ingredient_sheet_name(ingredient_xlsx, args.ingredient_sheet)

    packet_rows = load_rows(packet_csv)
    canonical_index = build_canonical_index(ingredient_xlsx, ingredient_sheet)
    workbench_rows = build_rows(packet_rows, canonical_index)

    out_csv = Path(args.out_csv).expanduser().resolve()
    out_json = Path(args.out_json).expanduser().resolve()
    write_csv(out_csv, workbench_rows)
    if args.out_xlsx:
        write_xlsx(Path(args.out_xlsx).expanduser().resolve(), normalize_text(args.sheet_name) or "Alias_Manual_Mapping_Workbench", workbench_rows)

    summary = {
        "packet_csv": str(packet_csv),
        "ingredient_workbook": str(ingredient_xlsx),
        "ingredient_sheet": ingredient_sheet,
        "row_count": len(workbench_rows),
        "suggested_decision_counts": dict(Counter(row["suggested_decision"] for row in workbench_rows)),
        "suggestion_confidence_counts": dict(Counter(row["suggestion_confidence"] for row in workbench_rows)),
        "out_csv": str(out_csv),
        "out_xlsx": str(Path(args.out_xlsx).expanduser().resolve()) if args.out_xlsx else "",
    }
    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
