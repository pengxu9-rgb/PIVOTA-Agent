#!/usr/bin/env python3
"""Export a conservative canonical-candidate patch package from Full_INCI priority queue.

This helper is read-only:
- it never edits the source workbook
- it exports clean candidate rows as an apply-ready CSV
- it moves parser-contaminated or duplicate-risk rows into a review remainder CSV
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to export Full_INCI priority patches. Install it in the local Python environment first."
    ) from exc


QUEUE_REQUIRED_COLUMNS = [
    "priority_score",
    "recommended_bucket",
    "recommended_action",
    "triage_reason",
    "raw_token",
    "normalized_token",
    "unmatched_count",
    "sku_row_count",
    "full_inci_count",
    "key_count",
    "product_only_count",
    "top_categories",
    "example_brands",
    "example_products",
    "example_urls",
    "in_current_master_like",
]

INGREDIENT_REQUIRED_COLUMNS = [
    "record_id",
    "canonical_inci_name",
    "normalized_key",
]

PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]

RECORD_ID_PATTERN = re.compile(r"^ing_patch_v13_(\d+)$", re.IGNORECASE)
PAREN_RE = re.compile(r"\([^)]*\)")


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    return "".join(ch for ch in raw if ch.isalnum())


def canonical_name_key(value: Any) -> str:
    return " ".join(normalize_text(value).casefold().split())


def semantic_match_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    raw = PAREN_RE.sub(" ", raw)
    raw = " ".join(raw.split())
    return "".join(ch for ch in raw if ch.isalnum())


def split_semicolon_values(value: Any) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    return [part.strip() for part in raw.split(";") if part and part.strip()]


def dedupe_join(values: list[str]) -> str:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        text = normalize_text(value)
        key = canonical_name_key(text)
        if not text or key in seen:
            continue
        seen.add(key)
        ordered.append(text)
    return "; ".join(ordered)


def read_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Workbook missing required '{sheet_name}' sheet: {path}")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"Workbook sheet '{sheet_name}' is empty: {path}")

    header = [normalize_text(cell) for cell in rows[0]]
    records: list[tuple[int, dict[str, str]]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(normalize_text(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        record = {
            key: normalize_text(padded[index] if index < len(padded) else "")
            for index, key in enumerate(header)
        }
        records.append((row_number, record))
    return header, records


def assert_required_columns(header: list[str], required: list[str], label: str) -> None:
    missing = [column for column in required if column not in header]
    if missing:
        raise SystemExit(f"{label} missing required columns: {', '.join(missing)}")


def resolve_ingredient_sheet_name(path: Path, requested_sheet: str | None) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheetnames = workbook.sheetnames
    if requested_sheet:
        if requested_sheet not in sheetnames:
            raise SystemExit(f"Ingredient workbook missing requested sheet '{requested_sheet}': {path}")
        return requested_sheet

    for sheet_name in PREFERRED_INGREDIENT_SHEETS:
        if sheet_name in sheetnames:
            return sheet_name
    raise SystemExit(
        f"Ingredient workbook did not contain any supported sheet names ({', '.join(PREFERRED_INGREDIENT_SHEETS)}): {path}"
    )


def detect_review_reasons(raw_token: str) -> list[str]:
    lowered = normalize_text(raw_token).lower()
    reasons: list[str] = []

    if "%" in raw_token:
        reasons.append("parser_contaminated_percent_token")
    if "active ingredient" in lowered or "inactive ingredient" in lowered:
        reasons.append("parser_contaminated_active_inactive_segment")
    if "ingredients:" in lowered or "ingredient:" in lowered:
        reasons.append("parser_contaminated_ingredient_prefix")
    if ";" in raw_token:
        reasons.append("parser_contaminated_multi_segment_token")

    return reasons


def build_existing_indexes(records: list[tuple[int, dict[str, str]]]) -> dict[str, Any]:
    normalized_index: dict[str, list[dict[str, str]]] = defaultdict(list)
    canonical_index: dict[str, list[dict[str, str]]] = defaultdict(list)
    semantic_index: dict[str, list[dict[str, str]]] = defaultdict(list)
    max_patch_number = 0

    for source_row_number, record in records:
        entry = {
            "record_id": record.get("record_id", ""),
            "canonical_inci_name": record.get("canonical_inci_name", ""),
            "normalized_key": record.get("normalized_key", ""),
            "source_row_number": str(source_row_number),
        }

        normalized = normalize_key(record.get("normalized_key") or record.get("canonical_inci_name"))
        if normalized:
            normalized_index[normalized].append(entry)

        canonical = canonical_name_key(record.get("canonical_inci_name"))
        if canonical:
            canonical_index[canonical].append(entry)

        semantic = semantic_match_key(record.get("canonical_inci_name"))
        if semantic:
            semantic_index[semantic].append(entry)

        record_id = normalize_text(record.get("record_id"))
        match = RECORD_ID_PATTERN.match(record_id)
        if match:
            max_patch_number = max(max_patch_number, int(match.group(1)))

    return {
        "normalized_index": normalized_index,
        "canonical_index": canonical_index,
        "semantic_index": semantic_index,
        "max_patch_number": max_patch_number,
    }


def set_if_present(row: dict[str, str], header: list[str], key: str, value: Any) -> None:
    if key in header:
        row[key] = normalize_text(value)


def build_candidate_row(
    header: list[str],
    queue_row: dict[str, str],
    record_id: str,
    candidate_name: str,
    normalized: str,
    semantic: str,
) -> dict[str, str]:
    row = {column: "" for column in header}

    parser_variants = dedupe_join([candidate_name, candidate_name.lower()])
    notes = dedupe_join(
        [
            f"priority_score={queue_row.get('priority_score', '')}",
            f"unmatched_count={queue_row.get('unmatched_count', '')}",
            f"sku_row_count={queue_row.get('sku_row_count', '')}",
            f"full_inci_count={queue_row.get('full_inci_count', '')}",
            f"key_count={queue_row.get('key_count', '')}",
            f"brands={queue_row.get('example_brands', '')}",
        ]
    )

    set_if_present(row, header, "record_id", record_id)
    set_if_present(row, header, "canonical_inci_name", candidate_name)
    set_if_present(row, header, "canonical_display_name", candidate_name)
    set_if_present(row, header, "ingredient_family", "other")
    set_if_present(row, header, "us_label_name", candidate_name)
    set_if_present(row, header, "eu_label_name", candidate_name)
    set_if_present(row, header, "us_label_variants", candidate_name)
    set_if_present(row, header, "eu_label_variants", candidate_name)
    set_if_present(row, header, "cross_market_notes", "Full_INCI priority patch candidate from v13 unmatched queue")
    set_if_present(row, header, "normalized_key", normalized)
    set_if_present(row, header, "aliases_common", "")
    set_if_present(row, header, "parser_variants", parser_variants)
    set_if_present(row, header, "deprecated_aliases", "")
    set_if_present(row, header, "alias_quality", "")
    set_if_present(
        row,
        header,
        "notes_for_parser",
        "Generated from Full_INCI_Priority_Queue; verify label punctuation, separators, and casing before promotion.",
    )
    set_if_present(row, header, "primary_bucket", "")
    set_if_present(row, header, "all_buckets", "")
    set_if_present(row, header, "function_tags", "")
    set_if_present(row, header, "benefit_tags", "")
    set_if_present(row, header, "risk_flags", "")
    for boolean_column in [
        "is_humectant",
        "is_barrier_support",
        "is_retinoid",
        "is_exfoliant",
        "is_uv_filter",
        "is_preservative",
        "is_surfactant",
        "is_fragrance_or_eo",
    ]:
        set_if_present(row, header, boolean_column, "no")
    set_if_present(row, header, "regulatory_bucket", "patch_candidate_review")
    set_if_present(row, header, "source_urls", queue_row.get("example_urls", ""))
    set_if_present(row, header, "source_authorities", "brand_official_pdp")
    set_if_present(row, header, "source_types", "official_brand_site")
    set_if_present(row, header, "review_status", "draft")
    set_if_present(row, header, "confidence", "low")
    set_if_present(row, header, "last_reviewed_at", "")
    set_if_present(
        row,
        header,
        "review_notes",
        "Generated from Full_INCI_Priority_Queue; verify canonical naming, taxonomy, and regulatory metadata before promotion.",
    )
    set_if_present(row, header, "notes", notes)
    set_if_present(row, header, "kb_version", "spec_v1_patch_v13_full_inci_priority")

    row["queue_priority_score"] = normalize_text(queue_row.get("priority_score"))
    row["queue_triage_reason"] = normalize_text(queue_row.get("triage_reason"))
    row["queue_top_categories"] = normalize_text(queue_row.get("top_categories"))
    row["queue_example_brands"] = normalize_text(queue_row.get("example_brands"))
    row["queue_example_products"] = normalize_text(queue_row.get("example_products"))
    row["queue_example_urls"] = normalize_text(queue_row.get("example_urls"))
    row["semantic_match_key"] = semantic

    return row


def summarize_entries(entries: list[dict[str, str]]) -> tuple[str, str]:
    record_ids = dedupe_join([entry.get("record_id", "") for entry in entries])
    names = dedupe_join([entry.get("canonical_inci_name", "") for entry in entries])
    return record_ids, names


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            writer.writerow({key: row.get(key, "") for key in fieldnames})


def main() -> None:
    parser = argparse.ArgumentParser(description="Export apply-ready canonical candidates from Full_INCI priority queue.")
    parser.add_argument("--queue-xlsx", required=True, help="Path to ingredient unmatched priority queue workbook")
    parser.add_argument("--ingredient-xlsx", required=True, help="Path to current ingredient workbook")
    parser.add_argument("--queue-sheet", default="Full_INCI_Priority_Queue", help="Queue workbook sheet to read")
    parser.add_argument("--ingredient-sheet", help="Optional ingredient workbook sheet to read")
    parser.add_argument("--out-apply-csv", required=True, help="Path to write clean apply-ready candidate rows")
    parser.add_argument("--out-review-csv", required=True, help="Path to write review remainder rows")
    parser.add_argument("--out-summary-json", required=True, help="Path to write summary JSON")
    args = parser.parse_args()

    queue_path = Path(args.queue_xlsx).expanduser().resolve()
    ingredient_path = Path(args.ingredient_xlsx).expanduser().resolve()
    out_apply_csv = Path(args.out_apply_csv).expanduser().resolve()
    out_review_csv = Path(args.out_review_csv).expanduser().resolve()
    out_summary_json = Path(args.out_summary_json).expanduser().resolve()

    ingredient_sheet = resolve_ingredient_sheet_name(ingredient_path, args.ingredient_sheet)
    queue_header, queue_records = read_sheet(queue_path, args.queue_sheet)
    ingredient_header, ingredient_records = read_sheet(ingredient_path, ingredient_sheet)

    assert_required_columns(queue_header, QUEUE_REQUIRED_COLUMNS, "Priority queue sheet")
    assert_required_columns(ingredient_header, INGREDIENT_REQUIRED_COLUMNS, "Ingredient workbook sheet")

    indexes = build_existing_indexes(ingredient_records)
    normalized_index = indexes["normalized_index"]
    canonical_index = indexes["canonical_index"]
    semantic_index = indexes["semantic_index"]
    next_patch_number = int(indexes["max_patch_number"]) + 1

    planned_normalized: dict[str, str] = {}
    planned_canonical: dict[str, str] = {}
    planned_semantic: dict[str, str] = {}

    apply_rows: list[dict[str, str]] = []
    review_rows: list[dict[str, str]] = []
    reason_counter: Counter[str] = Counter()

    for _source_row_number, queue_row in queue_records:
        candidate_name = normalize_text(queue_row.get("raw_token"))
        normalized = normalize_key(candidate_name)
        semantic = semantic_match_key(candidate_name)
        reason_codes = detect_review_reasons(candidate_name)

        normalized_conflicts = normalized_index.get(normalized, [])
        canonical_conflicts = canonical_index.get(canonical_name_key(candidate_name), [])
        semantic_conflicts = [
            entry
            for entry in semantic_index.get(semantic, [])
            if normalize_key(entry.get("canonical_inci_name")) != normalized
        ]

        if normalized_conflicts:
            reason_codes.append("existing_normalized_key_conflict")
        if canonical_conflicts:
            reason_codes.append("existing_canonical_name_conflict")
        if semantic and semantic in planned_semantic and planned_semantic[semantic] != normalized:
            reason_codes.append("batch_semantic_key_conflict")
        if normalized and normalized in planned_normalized:
            reason_codes.append("batch_normalized_key_conflict")
        if candidate_name and canonical_name_key(candidate_name) in planned_canonical:
            reason_codes.append("batch_canonical_name_conflict")
        if semantic_conflicts:
            reason_codes.append("existing_semantic_key_conflict")

        if reason_codes:
            existing_record_ids, existing_names = summarize_entries(
                normalized_conflicts + canonical_conflicts + semantic_conflicts
            )
            review_rows.append(
                {
                    **queue_row,
                    "suggested_canonical_inci_name": candidate_name,
                    "suggested_normalized_key": normalized,
                    "semantic_match_key": semantic,
                    "review_reason_codes": dedupe_join(reason_codes),
                    "existing_record_ids": existing_record_ids,
                    "existing_canonical_inci_names": existing_names,
                }
            )
            reason_counter.update(reason_codes)
            continue

        record_id = f"ing_patch_v13_{next_patch_number:03d}"
        next_patch_number += 1

        apply_rows.append(
            build_candidate_row(
                header=ingredient_header,
                queue_row=queue_row,
                record_id=record_id,
                candidate_name=candidate_name,
                normalized=normalized,
                semantic=semantic,
            )
        )

        planned_normalized[normalized] = record_id
        planned_canonical[canonical_name_key(candidate_name)] = record_id
        planned_semantic[semantic] = normalized

    apply_fieldnames = ingredient_header + [
        "queue_priority_score",
        "queue_triage_reason",
        "queue_top_categories",
        "queue_example_brands",
        "queue_example_products",
        "queue_example_urls",
        "semantic_match_key",
    ]
    review_fieldnames = queue_header + [
        "suggested_canonical_inci_name",
        "suggested_normalized_key",
        "semantic_match_key",
        "review_reason_codes",
        "existing_record_ids",
        "existing_canonical_inci_names",
    ]

    write_csv(out_apply_csv, apply_fieldnames, apply_rows)
    write_csv(out_review_csv, review_fieldnames, review_rows)

    summary = {
        "queue_workbook": str(queue_path),
        "queue_sheet": args.queue_sheet,
        "ingredient_workbook": str(ingredient_path),
        "ingredient_sheet": ingredient_sheet,
        "input_row_count": len(queue_records),
        "apply_ready_count": len(apply_rows),
        "review_count": len(review_rows),
        "reason_counts": dict(sorted(reason_counter.items())),
        "record_id_start": apply_rows[0]["record_id"] if apply_rows else None,
        "record_id_end": apply_rows[-1]["record_id"] if apply_rows else None,
        "out_apply_csv": str(out_apply_csv),
        "out_review_csv": str(out_review_csv),
    }

    out_summary_json.parent.mkdir(parents=True, exist_ok=True)
    out_summary_json.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
