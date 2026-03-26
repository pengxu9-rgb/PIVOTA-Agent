#!/usr/bin/env python3
"""Export reviewed SKU ingredient/signal workbook rows into downstream lanes."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to export reviewed SKU ingredient/signal layers. Install it in the local Python environment first."
    ) from exc


MANUAL_REVIEWED_SHEET = "Manual_Queue_Reviewed"

RUNTIME_EVIDENCE_FIELDS = [
    "sku_row_key",
    "brand_name",
    "product_name",
    "official_product_url",
    "market",
    "category",
    "evidence_lane",
    "runtime_evidence_eligible",
    "downstream_handoff_path",
    "canonical_ingredients",
    "signal_display_names",
    "signal_keys",
    "signal_handling_lane",
    "approved_decision",
    "review_status",
    "trust_tier",
    "decision_rationale",
    "source_packet",
]

PARSER_CONTROL_FIELDS = [
    "sku_row_key",
    "brand_name",
    "product_name",
    "official_product_url",
    "market",
    "category",
    "control_lane",
    "downstream_handoff_path",
    "canonical_ingredients",
    "signal_display_names",
    "signal_keys",
    "parser_excluded_fragments",
    "approved_decision",
    "review_status",
    "trust_tier",
    "decision_rationale",
    "source_packet",
]

SIGNAL_PREVIEW_FIELDS = [
    "sku_row_key",
    "brand_name",
    "product_name",
    "official_product_url",
    "market",
    "category",
    "preview_lane",
    "runtime_evidence_eligible",
    "signal_display_names",
    "signal_keys",
    "canonical_ingredients",
    "approved_decision",
    "review_status",
    "trust_tier",
    "decision_rationale",
    "source_packet",
]

FOLLOWUP_FIELDS = [
    "sku_row_key",
    "brand_name",
    "product_name",
    "official_product_url",
    "market",
    "category",
    "followup_lane",
    "signal_display_names",
    "signal_keys",
    "canonical_ingredients",
    "approved_decision",
    "review_status",
    "trust_tier",
    "decision_rationale",
    "source_packet",
]


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def read_reviewed_rows(path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Workbook missing required sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    out: list[dict[str, str]] = []
    for row in rows[1:]:
        if not any(normalize_text(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        out.append({header[index]: normalize_text(padded[index]) for index in range(len(header))})
    return out


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return [{key: normalize_text(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def append_sheet(workbook: Workbook, title: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])


def hybrid_signal_lane(approved_decision: str) -> tuple[str, str]:
    mapping = {
        "approve_hybrid_review_candidate": (
            "secondary_review_context",
            "eligible_for_hybrid_review_with_signal_context",
        ),
        "approve_hybrid_keep_ingredients__signal_preview_only": (
            "preview_only",
            "eligible_for_hybrid_review__signal_preview_only",
        ),
        "approve_hybrid_keep_ingredients__signal_family_only": (
            "family_only",
            "eligible_for_hybrid_review__signal_family_only",
        ),
        "approve_hybrid_keep_ingredients__route_signal_to_claim_or_flag_layer": (
            "claim_or_flag_only",
            "eligible_for_hybrid_review__signal_claim_or_flag_only",
        ),
    }
    return mapping.get(approved_decision, ("review_required", "needs_review"))


def signal_preview_lane(approved_decision: str) -> str:
    mapping = {
        "approve_signal_preview_only__family_or_group_term": "family_or_group_term",
        "approve_signal_preview_only__candidate_for_alias_or_canonical_followup": "candidate_for_alias_or_canonical_followup",
        "approve_signal_preview_only__proprietary_or_marketing": "proprietary_or_marketing",
        "approve_signal_preview_only": "generic_signal_preview_only",
    }
    return mapping.get(approved_decision, "needs_review")


def main() -> None:
    parser = argparse.ArgumentParser(description="Export reviewed SKU ingredient/signal workbook rows into downstream lanes.")
    parser.add_argument("--reviewed-xlsx", required=True, help="Reviewed manual handoff queue workbook")
    parser.add_argument("--existing-handoff-csv", required=True, help="Existing approved downstream handoff CSV")
    parser.add_argument("--out-runtime-evidence-csv", required=True, help="Where to write runtime-evidence-eligible rows")
    parser.add_argument("--out-parser-controls-csv", required=True, help="Where to write parser control rows")
    parser.add_argument("--out-signal-preview-csv", required=True, help="Where to write signal preview rows")
    parser.add_argument("--out-followup-csv", required=True, help="Where to write alias/canonical followup rows")
    parser.add_argument("--out-summary-json", required=True, help="Where to write summary JSON")
    parser.add_argument("--out-xlsx", help="Optional XLSX output with all lanes")
    args = parser.parse_args()

    reviewed_path = Path(args.reviewed_xlsx).expanduser().resolve()
    reviewed_rows = read_reviewed_rows(reviewed_path, MANUAL_REVIEWED_SHEET)
    existing_handoff_rows = read_csv_rows(Path(args.existing_handoff_csv).expanduser().resolve())

    runtime_rows: list[dict[str, str]] = []
    parser_control_rows: list[dict[str, str]] = []
    signal_preview_rows: list[dict[str, str]] = []
    followup_rows: list[dict[str, str]] = []

    for row in existing_handoff_rows:
        action = normalize_text(row.get("recommended_review_action"))
        if action == "ready_reference_only_sku":
            runtime_rows.append(
                {
                    "sku_row_key": normalize_text(row.get("sku_row_key")),
                    "brand_name": normalize_text(row.get("brand_name")),
                    "product_name": normalize_text(row.get("product_name")),
                    "official_product_url": normalize_text(row.get("official_product_url")),
                    "market": normalize_text(row.get("market")),
                    "category": normalize_text(row.get("category")),
                    "evidence_lane": "reference_only",
                    "runtime_evidence_eligible": "yes",
                    "downstream_handoff_path": normalize_text(row.get("downstream_handoff_path")),
                    "canonical_ingredients": normalize_text(row.get("canonical_ingredients")),
                    "signal_display_names": normalize_text(row.get("signal_display_names")),
                    "signal_keys": normalize_text(row.get("signal_keys")),
                    "signal_handling_lane": "none",
                    "approved_decision": "confirm_reference_only_sku",
                    "review_status": "approved",
                    "trust_tier": "high",
                    "decision_rationale": normalize_text(row.get("decision_rationale")),
                    "source_packet": normalize_text(row.get("source_packet")),
                }
            )
        elif action == "review_parser_fragment_series":
            parser_control_rows.append(
                {
                    "sku_row_key": normalize_text(row.get("sku_row_key")),
                    "brand_name": normalize_text(row.get("brand_name")),
                    "product_name": normalize_text(row.get("product_name")),
                    "official_product_url": normalize_text(row.get("official_product_url")),
                    "market": normalize_text(row.get("market")),
                    "category": normalize_text(row.get("category")),
                    "control_lane": "parser_fragment_exclusion_confirmed",
                    "downstream_handoff_path": normalize_text(row.get("downstream_handoff_path")),
                    "canonical_ingredients": normalize_text(row.get("canonical_ingredients")),
                    "signal_display_names": normalize_text(row.get("signal_display_names")),
                    "signal_keys": normalize_text(row.get("signal_keys")),
                    "parser_excluded_fragments": normalize_text(row.get("parser_excluded_fragments")),
                    "approved_decision": "confirm_parser_fragment_exclusion",
                    "review_status": "approved",
                    "trust_tier": "high",
                    "decision_rationale": normalize_text(row.get("decision_rationale")),
                    "source_packet": normalize_text(row.get("source_packet")),
                }
            )

    for row in reviewed_rows:
        action = normalize_text(row.get("recommended_review_action"))
        approved_decision = normalize_text(row.get("approved_decision")) or normalize_text(row.get("decision"))
        reviewer_notes = normalize_text(row.get("reviewer_notes")) or normalize_text(row.get("reviewer_notes_auto")) or normalize_text(row.get("decision_rationale"))

        if action == "review_hybrid_ingredient_signal_sku":
            signal_lane, handoff_path = hybrid_signal_lane(approved_decision)
            runtime_rows.append(
                {
                    "sku_row_key": normalize_text(row.get("sku_row_key")),
                    "brand_name": normalize_text(row.get("brand_name")),
                    "product_name": normalize_text(row.get("product_name")),
                    "official_product_url": normalize_text(row.get("official_product_url")),
                    "market": normalize_text(row.get("market")),
                    "category": normalize_text(row.get("category")),
                    "evidence_lane": "hybrid_canonical_primary",
                    "runtime_evidence_eligible": "yes",
                    "downstream_handoff_path": handoff_path,
                    "canonical_ingredients": normalize_text(row.get("canonical_ingredients")),
                    "signal_display_names": normalize_text(row.get("signal_display_names")),
                    "signal_keys": normalize_text(row.get("signal_keys")),
                    "signal_handling_lane": signal_lane,
                    "approved_decision": approved_decision,
                    "review_status": normalize_text(row.get("review_status")),
                    "trust_tier": normalize_text(row.get("trust_tier")),
                    "decision_rationale": reviewer_notes,
                    "source_packet": normalize_text(row.get("source_packet")) or MANUAL_REVIEWED_SHEET,
                }
            )
            continue

        if action == "review_signal_led_sku":
            preview_lane = signal_preview_lane(approved_decision)
            preview_row = {
                "sku_row_key": normalize_text(row.get("sku_row_key")),
                "brand_name": normalize_text(row.get("brand_name")),
                "product_name": normalize_text(row.get("product_name")),
                "official_product_url": normalize_text(row.get("official_product_url")),
                "market": normalize_text(row.get("market")),
                "category": normalize_text(row.get("category")),
                "preview_lane": preview_lane,
                "runtime_evidence_eligible": "no",
                "signal_display_names": normalize_text(row.get("signal_display_names")),
                "signal_keys": normalize_text(row.get("signal_keys")),
                "canonical_ingredients": normalize_text(row.get("canonical_ingredients")),
                "approved_decision": approved_decision,
                "review_status": normalize_text(row.get("review_status")),
                "trust_tier": normalize_text(row.get("trust_tier")),
                "decision_rationale": reviewer_notes,
                "source_packet": normalize_text(row.get("source_packet")) or MANUAL_REVIEWED_SHEET,
            }
            signal_preview_rows.append(preview_row)
            if preview_lane == "candidate_for_alias_or_canonical_followup":
                followup_rows.append(
                    {
                        "sku_row_key": preview_row["sku_row_key"],
                        "brand_name": preview_row["brand_name"],
                        "product_name": preview_row["product_name"],
                        "official_product_url": preview_row["official_product_url"],
                        "market": preview_row["market"],
                        "category": preview_row["category"],
                        "followup_lane": preview_lane,
                        "signal_display_names": preview_row["signal_display_names"],
                        "signal_keys": preview_row["signal_keys"],
                        "canonical_ingredients": preview_row["canonical_ingredients"],
                        "approved_decision": approved_decision,
                        "review_status": preview_row["review_status"],
                        "trust_tier": preview_row["trust_tier"],
                        "decision_rationale": reviewer_notes,
                        "source_packet": preview_row["source_packet"],
                    }
                )

    runtime_rows.sort(key=lambda row: (row["brand_name"].casefold(), row["product_name"].casefold()))
    parser_control_rows.sort(key=lambda row: (row["brand_name"].casefold(), row["product_name"].casefold()))
    signal_preview_rows.sort(key=lambda row: (row["brand_name"].casefold(), row["product_name"].casefold()))
    followup_rows.sort(key=lambda row: (row["brand_name"].casefold(), row["product_name"].casefold()))

    out_runtime = Path(args.out_runtime_evidence_csv).expanduser().resolve()
    out_parser = Path(args.out_parser_controls_csv).expanduser().resolve()
    out_preview = Path(args.out_signal_preview_csv).expanduser().resolve()
    out_followup = Path(args.out_followup_csv).expanduser().resolve()
    write_csv(out_runtime, RUNTIME_EVIDENCE_FIELDS, runtime_rows)
    write_csv(out_parser, PARSER_CONTROL_FIELDS, parser_control_rows)
    write_csv(out_preview, SIGNAL_PREVIEW_FIELDS, signal_preview_rows)
    write_csv(out_followup, FOLLOWUP_FIELDS, followup_rows)

    summary = {
        "reviewed_xlsx": str(reviewed_path),
        "existing_handoff_csv": str(Path(args.existing_handoff_csv).expanduser().resolve()),
        "runtime_evidence_count": len(runtime_rows),
        "parser_control_count": len(parser_control_rows),
        "signal_preview_count": len(signal_preview_rows),
        "followup_count": len(followup_rows),
        "runtime_evidence_lane_counts": dict(Counter(row["evidence_lane"] for row in runtime_rows)),
        "signal_handling_lane_counts": dict(Counter(row["signal_handling_lane"] for row in runtime_rows if row["signal_handling_lane"])),
        "signal_preview_lane_counts": dict(Counter(row["preview_lane"] for row in signal_preview_rows)),
        "out_runtime_evidence_csv": str(out_runtime),
        "out_parser_controls_csv": str(out_parser),
        "out_signal_preview_csv": str(out_preview),
        "out_followup_csv": str(out_followup),
    }

    if args.out_xlsx:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        append_sheet(workbook, "Runtime_Evidence", RUNTIME_EVIDENCE_FIELDS, runtime_rows)
        append_sheet(workbook, "Parser_Controls", PARSER_CONTROL_FIELDS, parser_control_rows)
        append_sheet(workbook, "Signal_Preview", SIGNAL_PREVIEW_FIELDS, signal_preview_rows)
        append_sheet(workbook, "Followup_14", FOLLOWUP_FIELDS, followup_rows)
        out_xlsx = Path(args.out_xlsx).expanduser().resolve()
        out_xlsx.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(out_xlsx)
        summary["out_xlsx"] = str(out_xlsx)

    out_summary_json = Path(args.out_summary_json).expanduser().resolve()
    out_summary_json.parent.mkdir(parents=True, exist_ok=True)
    out_summary_json.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
