#!/usr/bin/env python3
"""Build an ingredient-master followup packet from reviewed signal-led followup rows."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build the ingredient-master followup packet. Install it in the local Python environment first."
    ) from exc


ROW_FIELDS = [
    "sku_row_key",
    "brand_name",
    "product_name",
    "official_product_url",
    "market",
    "category",
    "term_count",
    "existing_bridge_term_count",
    "new_canonical_term_count",
    "hold_signal_term_count",
    "needs_species_confirmation_term_count",
    "recommended_row_action",
    "term_keys",
    "term_display_names",
    "term_lanes",
    "source_followup_lane",
    "source_decision",
    "source_review_status",
    "source_trust_tier",
    "decision_rationale",
]

TERM_FIELDS = [
    "signal_key",
    "display_signal_name",
    "signal_bucket",
    "term_lane",
    "recommended_master_action",
    "recommended_target_display_name",
    "recommended_target_inci_name",
    "reference_match_status",
    "reference_match_display_name",
    "reference_match_inci_name",
    "reference_match_aliases_common",
    "followup_priority",
    "sku_row_count",
    "example_brands",
    "example_products",
    "decision_rationale",
]


TERM_OVERRIDES: dict[str, dict[str, str]] = {
    "viniferine": {
        "term_lane": "new_canonical_candidate",
        "recommended_master_action": "open_new_named_active_candidate__needs_inci_confirmation",
        "recommended_target_display_name": "Viniferine",
        "recommended_target_inci_name": "",
        "reference_match_status": "no_current_reference_match",
        "followup_priority": "high",
        "decision_rationale": "High-value named active signal, but current ingredient reference does not contain a safe canonical row yet.",
    },
    "glycerin": {
        "term_lane": "existing_canonical_bridge",
        "recommended_master_action": "bridge_signal_to_existing_canonical__glycerin",
        "recommended_target_display_name": "Glycerin",
        "recommended_target_inci_name": "Glycerin",
        "reference_match_status": "existing_canonical_target_confirmed",
        "followup_priority": "high",
        "decision_rationale": "Canonical ingredient already exists; this is a routing/alias bridge problem, not a new ingredient-master row.",
    },
    "urea": {
        "term_lane": "existing_canonical_bridge",
        "recommended_master_action": "bridge_signal_to_existing_canonical__urea",
        "recommended_target_display_name": "Urea",
        "recommended_target_inci_name": "Urea",
        "reference_match_status": "existing_canonical_target_confirmed",
        "followup_priority": "high",
        "decision_rationale": "Canonical ingredient already exists; this should bridge to the existing Urea row instead of staying signal-only.",
    },
    "japanese_charcoal": {
        "term_lane": "hold_signal_only_pending_confirmation",
        "recommended_master_action": "keep_signal_only__needs_material_confirmation_before_charcoal_alias",
        "recommended_target_display_name": "Charcoal Powder",
        "recommended_target_inci_name": "Charcoal Powder",
        "reference_match_status": "candidate_existing_material_match_only",
        "followup_priority": "medium",
        "decision_rationale": "Material looks related to charcoal, but the country/material descriptor makes it unsafe to auto-promote into ingredient master without label confirmation.",
    },
    "chlorella": {
        "term_lane": "new_canonical_candidate",
        "recommended_master_action": "open_new_botanical_candidate__needs_inci_confirmation",
        "recommended_target_display_name": "Chlorella",
        "recommended_target_inci_name": "",
        "reference_match_status": "no_current_reference_match",
        "followup_priority": "high",
        "decision_rationale": "Botanical term looks ingredient-like and recurring, but the exact INCI form should be confirmed before a canonical row is added.",
    },
    "ginkgo_biloba": {
        "term_lane": "existing_canonical_bridge",
        "recommended_master_action": "add_common_alias_to_existing_canonical__ginkgo_biloba_leaf_extract",
        "recommended_target_display_name": "Ginkgo Biloba Leaf Extract",
        "recommended_target_inci_name": "Ginkgo Biloba Leaf Extract",
        "reference_match_status": "existing_canonical_candidate_needs_alias_review",
        "followup_priority": "high",
        "decision_rationale": "Ingredient master already contains the leaf-extract canonical; common-name routing should bridge to that existing row.",
    },
    "padina_pavonica": {
        "term_lane": "new_canonical_candidate",
        "recommended_master_action": "open_new_marine_extract_candidate__needs_inci_confirmation",
        "recommended_target_display_name": "Padina Pavonica",
        "recommended_target_inci_name": "",
        "reference_match_status": "no_current_reference_match",
        "followup_priority": "high",
        "decision_rationale": "Marine botanical term is ingredient-like, but the exact extract/thallus INCI should be confirmed before master writeback.",
    },
    "sakura_ferment": {
        "term_lane": "hold_signal_only_pending_confirmation",
        "recommended_master_action": "keep_signal_only__proprietary_or_ferment_blend_needs_source_confirmation",
        "recommended_target_display_name": "",
        "recommended_target_inci_name": "",
        "reference_match_status": "no_current_reference_match",
        "followup_priority": "low",
        "decision_rationale": "Looks like a branded ferment descriptor rather than a safe canonical ingredient row.",
    },
    "licochalcone_a": {
        "term_lane": "new_canonical_candidate",
        "recommended_master_action": "open_new_named_active_candidate__needs_inci_confirmation",
        "recommended_target_display_name": "Licochalcone A",
        "recommended_target_inci_name": "",
        "reference_match_status": "no_current_reference_match",
        "followup_priority": "high",
        "decision_rationale": "Named active is high-value, but the ingredient-master row should only be added once exact INCI labeling is confirmed.",
    },
    "thiamidol": {
        "term_lane": "new_canonical_candidate",
        "recommended_master_action": "open_new_named_active_candidate__needs_inci_confirmation",
        "recommended_target_display_name": "Thiamidol",
        "recommended_target_inci_name": "",
        "reference_match_status": "no_current_reference_match",
        "followup_priority": "high",
        "decision_rationale": "High-value branded active with repeated SKU presence; merits ingredient-master followup, but not blind canonical promotion.",
    },
    "dexpanthenol": {
        "term_lane": "existing_canonical_bridge",
        "recommended_master_action": "add_alias_to_existing_canonical__panthenol",
        "recommended_target_display_name": "Panthenol",
        "recommended_target_inci_name": "Panthenol",
        "reference_match_status": "existing_canonical_candidate_needs_alias_review",
        "followup_priority": "high",
        "decision_rationale": "Looks like an alias/variant of Panthenol rather than a net-new canonical ingredient row.",
    },
    "barbados_cherry": {
        "term_lane": "new_canonical_candidate",
        "recommended_master_action": "open_new_botanical_candidate__needs_inci_confirmation",
        "recommended_target_display_name": "Barbados Cherry",
        "recommended_target_inci_name": "",
        "reference_match_status": "no_current_reference_match",
        "followup_priority": "medium",
        "decision_rationale": "Common botanical name looks ingredient-like, but exact INCI species/extract form should be confirmed first.",
    },
    "kombucha": {
        "term_lane": "hold_signal_only_pending_confirmation",
        "recommended_master_action": "keep_signal_only__ferment_term_too_broad_for_master",
        "recommended_target_display_name": "",
        "recommended_target_inci_name": "",
        "reference_match_status": "no_current_reference_match",
        "followup_priority": "low",
        "decision_rationale": "Too broad as a ferment term to promote directly into ingredient master without exact underlying INCI mapping.",
    },
    "royal_jelly": {
        "term_lane": "existing_canonical_bridge",
        "recommended_master_action": "add_common_alias_to_existing_canonical__royal_jelly_extract",
        "recommended_target_display_name": "Royal Jelly Extract",
        "recommended_target_inci_name": "Royal Jelly Extract",
        "reference_match_status": "existing_canonical_candidate_needs_alias_review",
        "followup_priority": "high",
        "decision_rationale": "Existing ingredient-master row already covers the extract; common-name routing should attach to that row.",
    },
    "three_black_bee_honeys": {
        "term_lane": "hold_signal_only_pending_confirmation",
        "recommended_master_action": "keep_signal_only__marketing_blend_not_safe_for_master",
        "recommended_target_display_name": "",
        "recommended_target_inci_name": "",
        "reference_match_status": "no_current_reference_match",
        "followup_priority": "low",
        "decision_rationale": "Marketing/blend phrase is not safe to promote into ingredient master as a canonical ingredient.",
    },
    "calendula": {
        "term_lane": "existing_canonical_bridge",
        "recommended_master_action": "add_common_alias_to_existing_canonical__calendula_officinalis_flower_extract",
        "recommended_target_display_name": "Calendula Officinalis Flower Extract",
        "recommended_target_inci_name": "Calendula Officinalis Flower Extract",
        "reference_match_status": "existing_canonical_candidate_needs_alias_review",
        "followup_priority": "high",
        "decision_rationale": "Common-name botanical should bridge to the existing calendula extract row rather than stay signal-only.",
    },
    "chamomile": {
        "term_lane": "needs_species_or_form_confirmation",
        "recommended_master_action": "needs_species_or_form_confirmation_before_alias_candidate",
        "recommended_target_display_name": "Chamomile",
        "recommended_target_inci_name": "",
        "reference_match_status": "ambiguous_existing_botanical_family_match",
        "followup_priority": "medium",
        "decision_rationale": "Common name is ingredient-like, but species and oil/extract form are ambiguous in current master coverage.",
    },
    "pansy": {
        "term_lane": "new_canonical_candidate",
        "recommended_master_action": "open_new_botanical_candidate__needs_inci_confirmation",
        "recommended_target_display_name": "Pansy",
        "recommended_target_inci_name": "",
        "reference_match_status": "no_current_reference_match",
        "followup_priority": "medium",
        "decision_rationale": "Common botanical name looks ingredient-like, but exact INCI species/form should be confirmed before canonicalization.",
    },
}


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def norm_token(value: str) -> str:
    return "".join(ch.lower() for ch in normalize_text(value) if ch.isalnum())


def split_multi(value: str) -> list[str]:
    return [part.strip() for part in normalize_text(value).split(";") if part.strip()]


def pair_terms(displays: list[str], keys: list[str]) -> list[tuple[str, str]]:
    if not displays and not keys:
        return []
    if not displays:
        return [(key.replace("_", " "), key) for key in keys]
    if not keys:
        return [(display, norm_token(display)) for display in displays]

    unused_keys = keys[:]
    pairs: list[tuple[str, str]] = []
    for display in displays:
        display_norm = norm_token(display)
        matched_index = next((index for index, key in enumerate(unused_keys) if norm_token(key) == display_norm), None)
        if matched_index is None:
            matched_index = 0 if unused_keys else None
        if matched_index is None:
            pairs.append((display, norm_token(display)))
            continue
        key = unused_keys.pop(matched_index)
        pairs.append((display, key))
    for key in unused_keys:
        pairs.append((key.replace("_", " "), key))
    return pairs


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return [{key: normalize_text(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def append_sheet(workbook: Workbook, title: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])


def append_summary_sheet(workbook: Workbook, summary: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Summary")
    sheet.append(["metric", "value"])
    for key, value in summary.items():
        if isinstance(value, dict):
            sheet.append([key, json.dumps(value, ensure_ascii=True, sort_keys=True)])
        else:
            sheet.append([key, value])


def pick_reference_match(
    signal_key: str,
    term_display_name: str,
    reference_rows: list[dict[str, str]],
    recommended_target_display_name: str,
    recommended_target_inci_name: str,
) -> tuple[str, str, str]:
    preferred = norm_token(recommended_target_display_name or recommended_target_inci_name)
    search_targets = [preferred, norm_token(signal_key), norm_token(term_display_name)]
    search_targets = [target for target in search_targets if target]
    candidates: list[tuple[int, dict[str, str]]] = []
    for row in reference_rows:
        candidate_fields = [
            normalize_text(row.get("canonical_display_name")),
            normalize_text(row.get("canonical_inci_name")),
            normalize_text(row.get("aliases_common")),
            normalize_text(row.get("us_label_variants")),
            normalize_text(row.get("eu_label_variants")),
        ]
        field_norms = [norm_token(field) for field in candidate_fields if field]
        if not field_norms:
            continue
        score = 0
        for target in search_targets:
            if any(field == target for field in field_norms):
                score = max(score, 3)
            elif any(target and target in field for field in field_norms):
                score = max(score, 1)
        if score:
            canonical_display_norm = norm_token(row.get("canonical_display_name"))
            canonical_inci_norm = norm_token(row.get("canonical_inci_name"))
            if preferred and preferred in {canonical_display_norm, canonical_inci_norm}:
                score = 4
            candidates.append((score, row))
    if not candidates:
        return ("", "", "")
    candidates.sort(
        key=lambda item: (
            -item[0],
            normalize_text(item[1].get("canonical_display_name")).casefold(),
            normalize_text(item[1].get("canonical_inci_name")).casefold(),
        )
    )
    top = candidates[0][1]
    return (
        normalize_text(top.get("canonical_display_name")),
        normalize_text(top.get("canonical_inci_name")),
        normalize_text(top.get("aliases_common")),
    )


def recommended_row_action(lanes: list[str]) -> str:
    lane_set = set(lanes)
    if lane_set == {"existing_canonical_bridge"}:
        return "bridge_terms_to_existing_canonical_rows"
    if lane_set == {"new_canonical_candidate"}:
        return "open_new_canonical_candidate_terms"
    if lane_set == {"hold_signal_only_pending_confirmation"}:
        return "keep_signal_only_pending_source_confirmation"
    if lane_set == {"needs_species_or_form_confirmation"}:
        return "needs_species_or_form_confirmation_before_master_writeback"
    return "split_row_terms_by_master_action"


def main() -> None:
    parser = argparse.ArgumentParser(description="Build an ingredient-master followup packet from reviewed signal-led followup rows.")
    parser.add_argument("--followup-csv", required=True, help="Reviewed alias/canonical followup CSV")
    parser.add_argument("--ingredient-reference-csv", required=True, help="Current ingredient reference ingest CSV")
    parser.add_argument("--signal-dict-csv", required=True, help="Current signal dictionary CSV")
    parser.add_argument("--out-rows-csv", required=True, help="Where to write row-level followup packet CSV")
    parser.add_argument("--out-terms-csv", required=True, help="Where to write unique-term followup packet CSV")
    parser.add_argument("--out-summary-json", required=True, help="Where to write summary JSON")
    parser.add_argument("--out-xlsx", help="Optional workbook path")
    args = parser.parse_args()

    followup_rows = read_csv_rows(Path(args.followup_csv).expanduser().resolve())
    reference_rows = read_csv_rows(Path(args.ingredient_reference_csv).expanduser().resolve())
    signal_rows = read_csv_rows(Path(args.signal_dict_csv).expanduser().resolve())
    signal_by_key = {row["signal_key"]: row for row in signal_rows}

    row_outputs: list[dict[str, str]] = []
    term_context: dict[str, dict[str, Any]] = {}

    for row in followup_rows:
        term_pairs = pair_terms(split_multi(row.get("signal_display_names", "")), split_multi(row.get("signal_keys", "")))
        term_lanes: list[str] = []
        display_names: list[str] = []
        keys: list[str] = []
        for display_name, signal_key in term_pairs:
            override = TERM_OVERRIDES.get(signal_key)
            if not override:
                continue
            term_lane = override["term_lane"]
            term_lanes.append(term_lane)
            display_names.append(display_name)
            keys.append(signal_key)

            context = term_context.setdefault(
                signal_key,
                {
                    "signal_key": signal_key,
                    "display_signal_name": display_name,
                    "signal_bucket": normalize_text(signal_by_key.get(signal_key, {}).get("signal_bucket")),
                    "term_lane": override["term_lane"],
                    "recommended_master_action": override["recommended_master_action"],
                    "recommended_target_display_name": override["recommended_target_display_name"],
                    "recommended_target_inci_name": override["recommended_target_inci_name"],
                    "reference_match_status": override["reference_match_status"],
                    "followup_priority": override["followup_priority"],
                    "decision_rationale": override["decision_rationale"],
                    "brands": set(),
                    "products": set(),
                    "sku_rows": set(),
                },
            )
            context["display_signal_name"] = display_name
            context["brands"].add(normalize_text(row.get("brand_name")))
            context["products"].add(normalize_text(row.get("product_name")))
            context["sku_rows"].add(normalize_text(row.get("sku_row_key")))

        row_outputs.append(
            {
                "sku_row_key": normalize_text(row.get("sku_row_key")),
                "brand_name": normalize_text(row.get("brand_name")),
                "product_name": normalize_text(row.get("product_name")),
                "official_product_url": normalize_text(row.get("official_product_url")),
                "market": normalize_text(row.get("market")),
                "category": normalize_text(row.get("category")),
                "term_count": str(len(term_lanes)),
                "existing_bridge_term_count": str(sum(1 for lane in term_lanes if lane == "existing_canonical_bridge")),
                "new_canonical_term_count": str(sum(1 for lane in term_lanes if lane == "new_canonical_candidate")),
                "hold_signal_term_count": str(sum(1 for lane in term_lanes if lane == "hold_signal_only_pending_confirmation")),
                "needs_species_confirmation_term_count": str(sum(1 for lane in term_lanes if lane == "needs_species_or_form_confirmation")),
                "recommended_row_action": recommended_row_action(term_lanes),
                "term_keys": "; ".join(keys),
                "term_display_names": "; ".join(display_names),
                "term_lanes": "; ".join(term_lanes),
                "source_followup_lane": normalize_text(row.get("followup_lane")),
                "source_decision": normalize_text(row.get("approved_decision")),
                "source_review_status": normalize_text(row.get("review_status")),
                "source_trust_tier": normalize_text(row.get("trust_tier")),
                "decision_rationale": normalize_text(row.get("decision_rationale")),
            }
        )

    term_outputs: list[dict[str, str]] = []
    for signal_key, context in sorted(term_context.items(), key=lambda item: (item[1]["display_signal_name"].casefold(), item[0].casefold())):
        match_display, match_inci, match_aliases = pick_reference_match(
            signal_key=signal_key,
            term_display_name=context["display_signal_name"],
            reference_rows=reference_rows,
            recommended_target_display_name=context["recommended_target_display_name"],
            recommended_target_inci_name=context["recommended_target_inci_name"],
        )
        term_outputs.append(
            {
                "signal_key": signal_key,
                "display_signal_name": context["display_signal_name"],
                "signal_bucket": context["signal_bucket"],
                "term_lane": context["term_lane"],
                "recommended_master_action": context["recommended_master_action"],
                "recommended_target_display_name": context["recommended_target_display_name"],
                "recommended_target_inci_name": context["recommended_target_inci_name"],
                "reference_match_status": context["reference_match_status"],
                "reference_match_display_name": match_display,
                "reference_match_inci_name": match_inci,
                "reference_match_aliases_common": match_aliases,
                "followup_priority": context["followup_priority"],
                "sku_row_count": str(len(context["sku_rows"])),
                "example_brands": "; ".join(sorted(context["brands"])),
                "example_products": "; ".join(sorted(context["products"])),
                "decision_rationale": context["decision_rationale"],
            }
        )

    row_outputs.sort(key=lambda row: (row["brand_name"].casefold(), row["product_name"].casefold()))
    term_outputs.sort(key=lambda row: (row["followup_priority"], row["display_signal_name"].casefold(), row["signal_key"].casefold()))

    out_rows = Path(args.out_rows_csv).expanduser().resolve()
    out_terms = Path(args.out_terms_csv).expanduser().resolve()
    write_csv(out_rows, ROW_FIELDS, row_outputs)
    write_csv(out_terms, TERM_FIELDS, term_outputs)

    summary = {
        "followup_row_count": len(row_outputs),
        "unique_term_count": len(term_outputs),
        "row_action_counts": dict(Counter(row["recommended_row_action"] for row in row_outputs)),
        "term_lane_counts": dict(Counter(row["term_lane"] for row in term_outputs)),
        "priority_counts": dict(Counter(row["followup_priority"] for row in term_outputs)),
        "out_rows_csv": str(out_rows),
        "out_terms_csv": str(out_terms),
    }

    if args.out_xlsx:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        append_summary_sheet(workbook, summary)
        append_sheet(workbook, "Row_Followup", ROW_FIELDS, row_outputs)
        append_sheet(workbook, "Unique_Terms", TERM_FIELDS, term_outputs)
        out_xlsx = Path(args.out_xlsx).expanduser().resolve()
        out_xlsx.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(out_xlsx)
        summary["out_xlsx"] = str(out_xlsx)

    out_summary = Path(args.out_summary_json).expanduser().resolve()
    out_summary.parent.mkdir(parents=True, exist_ok=True)
    out_summary.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
