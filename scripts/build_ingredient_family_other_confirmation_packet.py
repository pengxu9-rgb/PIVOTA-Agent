#!/usr/bin/env python3
"""Build a decision-ready packet for confirming `ingredient_family=other` rows."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build the ingredient-family-other confirmation packet. Install it in the local Python environment first."
    ) from exc


PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]


FAMILY_OTHER_MARKER = "confirmed_ingredient_family_other"


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def load_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def resolve_sheet_name(path: Path, requested_sheet: str | None = None) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    if requested_sheet:
        if requested_sheet not in sheet_names:
            raise SystemExit(f"Workbook sheet not found: {requested_sheet}")
        return requested_sheet
    for candidate in PREFERRED_INGREDIENT_SHEETS:
        if candidate in sheet_names:
            return candidate
    raise SystemExit(
        "Workbook is missing a supported ingredient sheet. Expected one of: "
        + ", ".join(PREFERRED_INGREDIENT_SHEETS)
    )


def read_workbook_context(path: Path, sheet_name: str) -> dict[str, dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]

    required = {"record_id", "ingredient_family", "review_notes"}
    missing = [column for column in required if column not in header]
    if missing:
        raise SystemExit(f"Workbook target sheet must include: {', '.join(missing)}")

    result: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        record = {
            key: normalize_text(row[index] if index < len(row) else "")
            for index, key in enumerate(header)
        }
        record_id = record.get("record_id")
        if record_id:
            result[record_id] = record
    return result


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a decision-ready packet for ingredient-family-other confirmation.")
    parser.add_argument("--ingredient-xlsx", required=True, help="Current ingredient workbook")
    parser.add_argument("--sheet-name", help="Optional workbook sheet name")
    parser.add_argument("--family-review-csv", required=True, help="CSV of remaining ingredient_family_review rows")
    parser.add_argument("--out-csv", required=True, help="Where to write the decision packet CSV")
    parser.add_argument("--out-json", help="Optional path for a JSON copy of the packet")
    args = parser.parse_args()

    workbook_path = Path(args.ingredient_xlsx).expanduser().resolve()
    family_review_path = Path(args.family_review_csv).expanduser().resolve()
    rows = load_rows(family_review_path)
    sheet_name = resolve_sheet_name(workbook_path, args.sheet_name)
    workbook_context = read_workbook_context(workbook_path, sheet_name)

    packet_rows: list[dict[str, str]] = []
    for row in rows:
        record_id = normalize_text(row.get("record_id"))
        record = workbook_context.get(record_id, {})
        packet_rows.append(
            {
                "record_id": record_id,
                "canonical_inci_name": normalize_text(row.get("canonical_inci_name")) or record.get("canonical_inci_name", ""),
                "primary_bucket": record.get("primary_bucket", ""),
                "function_tags": record.get("function_tags", ""),
                "benefit_tags": record.get("benefit_tags", ""),
                "existing_ingredient_family": record.get("ingredient_family", ""),
                "existing_review_notes": record.get("review_notes", ""),
                "suggested_marker": FAMILY_OTHER_MARKER,
                "suggested_resolution": "confirm_ingredient_family_other",
                "suggested_rationale": "current controlled vocabulary does not provide a stronger low-risk family fit",
                "decision": "",
                "approved_marker": FAMILY_OTHER_MARKER,
                "reviewer_notes": "",
            }
        )

    fieldnames = [
        "record_id",
        "canonical_inci_name",
        "primary_bucket",
        "function_tags",
        "benefit_tags",
        "existing_ingredient_family",
        "existing_review_notes",
        "suggested_marker",
        "suggested_resolution",
        "suggested_rationale",
        "decision",
        "approved_marker",
        "reviewer_notes",
    ]

    out_csv = Path(args.out_csv).expanduser().resolve()
    write_csv(out_csv, packet_rows, fieldnames)

    out_json = None
    if args.out_json:
        out_json = Path(args.out_json).expanduser().resolve()
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(
            json.dumps(
                {
                    "source_workbook": str(workbook_path),
                    "source_sheet": sheet_name,
                    "source_family_review_csv": str(family_review_path),
                    "row_count": len(packet_rows),
                    "decision_values": [
                        "confirm_ingredient_family_other",
                        "keep_open",
                        "needs_research",
                    ],
                    "rows": packet_rows,
                },
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

    print(
        json.dumps(
            {
                "source_workbook": str(workbook_path),
                "source_sheet": sheet_name,
                "source_family_review_csv": str(family_review_path),
                "row_count": len(packet_rows),
                "out_csv": str(out_csv),
                "out_json": str(out_json) if out_json else None,
            },
            ensure_ascii=True,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
