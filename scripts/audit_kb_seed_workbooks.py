#!/usr/bin/env python3
"""Read-only audit for ingredient/SKU workbook seed files.

This script is intentionally conservative:
- it never writes to runtime databases
- it audits workbook structure and readiness
- it emits a mapping recommendation for staging/import layers
"""

from __future__ import annotations

import argparse
import json
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to audit workbook seeds. Install it in the local Python environment first."
    ) from exc


INGREDIENT_REQUIRED_COLUMNS = [
    "record_id",
    "canonical_inci_name",
    "us_label_name",
    "eu_label_name",
    "normalized_key",
    "parser_variants",
    "primary_bucket",
    "function_tags",
    "benefit_tags",
    "source_urls",
    "kb_version",
]

SKU_REQUIRED_COLUMNS = [
    "brand_name",
    "product_name",
    "official_product_url",
    "market",
    "category",
    "ingredient_granularity",
    "ingredients_or_key_ingredients",
    "extraction_status",
]

BRAND_REQUIRED_COLUMNS = [
    "brand_name",
    "official_url",
    "market_focus_guess",
    "priority_tier",
    "seed_inventory_status",
]

QUEUE_REQUIRED_COLUMNS = [
    "brand_name",
    "official_url",
    "market_focus_guess",
    "priority_tier",
    "seed_inventory_status",
    "phase",
    "suggested_next_step",
]

FLAG_COLUMNS = [
    "is_humectant",
    "is_barrier_support",
    "is_retinoid",
    "is_exfoliant",
    "is_uv_filter",
    "is_preservative",
    "is_surfactant",
    "is_fragrance_or_eo",
]


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def is_nonempty(value: Any) -> bool:
    return bool(normalize_text(value))


def read_sheet_records(workbook_path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(cell is not None and normalize_text(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        records.append(dict(zip(header, padded)))
    return header, records


def missing_columns(header: list[str], required: list[str]) -> list[str]:
    present = set(header)
    return [column for column in required if column not in present]


def count_missing(records: list[dict[str, Any]], field: str) -> int:
    return sum(not is_nonempty(record.get(field)) for record in records)


def duplicate_key_count(records: list[dict[str, Any]], field: str) -> int:
    counts = Counter(normalize_text(record.get(field)) for record in records if is_nonempty(record.get(field)))
    return sum(count > 1 for count in counts.values())


def ingredient_audit(path: Path) -> dict[str, Any]:
    header, records = read_sheet_records(path, "Dictionary")
    bucket_counts = Counter(normalize_text(record.get("primary_bucket")) for record in records)
    flag_yes_counts = {
        field: sum(normalize_text(record.get(field)).lower() == "yes" for record in records)
        for field in FLAG_COLUMNS
    }
    return {
        "workbook": str(path),
        "sheet": "Dictionary",
        "required_columns_missing": missing_columns(header, INGREDIENT_REQUIRED_COLUMNS),
        "row_count": len(records),
        "unique_record_id_count": len({normalize_text(record.get("record_id")) for record in records}),
        "unique_canonical_inci_name_count": len(
            {normalize_text(record.get("canonical_inci_name")) for record in records}
        ),
        "unique_normalized_key_count": len({normalize_text(record.get("normalized_key")) for record in records}),
        "duplicate_record_id_count": duplicate_key_count(records, "record_id"),
        "duplicate_canonical_inci_name_count": duplicate_key_count(records, "canonical_inci_name"),
        "duplicate_normalized_key_count": duplicate_key_count(records, "normalized_key"),
        "missing_aliases_common_count": count_missing(records, "aliases_common"),
        "missing_parser_variants_count": count_missing(records, "parser_variants"),
        "missing_source_urls_count": count_missing(records, "source_urls"),
        "bucket_counts": dict(bucket_counts.most_common()),
        "flag_yes_counts": flag_yes_counts,
        "recommended_target": {
            "table": "seed_ingest.ingredient_reference_seed",
            "role": "parser/reference dictionary seed only",
            "ready_for_direct_runtime_use": False,
        },
        "gating": {
            "seed_ready": (
                not missing_columns(header, INGREDIENT_REQUIRED_COLUMNS)
                and duplicate_key_count(records, "record_id") == 0
                and duplicate_key_count(records, "normalized_key") == 0
            ),
            "needs_manual_alias_enrichment": count_missing(records, "aliases_common") > 0,
        },
    }


def sku_inventory_audit(path: Path) -> dict[str, Any]:
    header, records = read_sheet_records(path, "SKU_Seed_Inventory")
    status_counts = Counter(normalize_text(record.get("extraction_status")) for record in records)
    granularity_counts = Counter(normalize_text(record.get("ingredient_granularity")) for record in records)
    category_counts = Counter(normalize_text(record.get("category")) for record in records)
    ready_for_harvest = sum(
        normalize_text(record.get("extraction_status")) in {"done", "partial_key_ingredients", "pending_full_inci"}
        and is_nonempty(record.get("official_product_url"))
        for record in records
    )
    full_inci_ready = sum(
        normalize_text(record.get("ingredient_granularity")) == "full_inci_official"
        and normalize_text(record.get("extraction_status")) == "done"
        for record in records
    )
    return {
        "sheet": "SKU_Seed_Inventory",
        "required_columns_missing": missing_columns(header, SKU_REQUIRED_COLUMNS),
        "row_count": len(records),
        "unique_brand_count": len({normalize_text(record.get("brand_name")) for record in records}),
        "unique_product_url_count": len(
            {
                normalize_text(record.get("official_product_url"))
                for record in records
                if is_nonempty(record.get("official_product_url"))
            }
        ),
        "duplicate_product_url_count": duplicate_key_count(records, "official_product_url"),
        "missing_sku_code_count": count_missing(records, "sku_code"),
        "missing_size_options_count": count_missing(records, "size_options"),
        "missing_ingredient_text_count": count_missing(records, "ingredients_or_key_ingredients"),
        "status_counts": dict(status_counts),
        "granularity_counts": dict(granularity_counts),
        "category_counts": dict(category_counts.most_common()),
        "harvestable_seed_count": ready_for_harvest,
        "full_inci_seed_count": full_inci_ready,
        "recommended_target": {
            "table": "seed_ingest.sku_seed_inventory_seed",
            "role": "brand/SKU seed inventory and harvest queue input",
            "ready_for_direct_runtime_use": False,
        },
    }


def brand_roster_audit(path: Path) -> dict[str, Any]:
    header, records = read_sheet_records(path, "Brand_Roster_Cleaned")
    return {
        "sheet": "Brand_Roster_Cleaned",
        "required_columns_missing": missing_columns(header, BRAND_REQUIRED_COLUMNS),
        "row_count": len(records),
        "priority_tier_1_count": sum(record.get("priority_tier") == 1 for record in records),
        "seeded_count": sum(normalize_text(record.get("seed_inventory_status")).lower() == "seeded" for record in records),
        "pending_count": sum(normalize_text(record.get("seed_inventory_status")).lower() == "pending" for record in records),
        "recommended_target": {
            "table": "seed_ingest.brand_roster_seed",
            "role": "brand expansion queue seed",
            "ready_for_direct_runtime_use": False,
        },
    }


def extraction_queue_audit(path: Path) -> dict[str, Any]:
    header, records = read_sheet_records(path, "Extraction_Queue")
    phase_counts = Counter(normalize_text(record.get("phase")) for record in records)
    return {
        "sheet": "Extraction_Queue",
        "required_columns_missing": missing_columns(header, QUEUE_REQUIRED_COLUMNS),
        "row_count": len(records),
        "phase_counts": dict(phase_counts),
        "recommended_target": {
            "table": "seed_ingest.extraction_queue_seed",
            "role": "operator-facing extraction queue seed",
            "ready_for_direct_runtime_use": False,
        },
    }


def build_summary(
    ingredient: dict[str, Any],
    sku_inventory: dict[str, Any],
    brand_roster: dict[str, Any],
    queue: dict[str, Any],
) -> dict[str, Any]:
    return {
        "recommendation": {
            "use_as_seed_files": True,
            "use_as_final_runtime_kb": False,
            "reason": [
                "ingredient workbook is a strong parser/reference dictionary seed",
                "sku workbook is a seed inventory and extraction queue sample, not a full SKU KB",
                "runtime ranking should consume harvested/reviewed SKU ingredient evidence rather than raw workbook rows",
            ],
        },
        "next_tables": [
            "seed_ingest.ingredient_reference_seed",
            "seed_ingest.brand_roster_seed",
            "seed_ingest.sku_seed_inventory_seed",
            "seed_ingest.extraction_queue_seed",
        ],
        "promotion_path": [
            "ingest workbook rows into seed_ingest staging tables",
            "validate required columns, unique keys, and market/url hygiene",
            "drive official PDP extraction / ingredient harvest from sku_seed_inventory_seed",
            "review harvested ingredient rows",
            "ingest reviewed rows into pci_kb.sku_ingredients",
            "wire guidance ranking to ingredient evidence from reviewed SKU KB only",
        ],
        "current_seed_health": {
            "ingredient_seed_ready": ingredient["gating"]["seed_ready"],
            "sku_seed_sample_only": sku_inventory["row_count"] < 100,
            "sku_full_inci_rows": sku_inventory["full_inci_seed_count"],
            "brand_roster_rows": brand_roster["row_count"],
            "queue_rows": queue["row_count"],
        },
    }


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit workbook seeds for ingredient/SKU KB rebuild.")
    parser.add_argument("--ingredient-xlsx", required=True, help="Path to the ingredient reference workbook")
    parser.add_argument("--sku-xlsx", required=True, help="Path to the brand/SKU seed workbook")
    parser.add_argument("--out-json", help="Optional path to write the audit JSON")
    args = parser.parse_args()

    ingredient_path = Path(args.ingredient_xlsx).expanduser().resolve()
    sku_path = Path(args.sku_xlsx).expanduser().resolve()

    ingredient = ingredient_audit(ingredient_path)
    sku_inventory = sku_inventory_audit(sku_path)
    brand_roster = brand_roster_audit(sku_path)
    queue = extraction_queue_audit(sku_path)

    payload = {
        "ingredient_workbook_audit": ingredient,
        "sku_workbook_audit": {
            "sku_seed_inventory": sku_inventory,
            "brand_roster": brand_roster,
            "extraction_queue": queue,
        },
        "summary": build_summary(ingredient, sku_inventory, brand_roster, queue),
    }

    rendered = json.dumps(payload, ensure_ascii=True, indent=2)
    if args.out_json:
        out_path = Path(args.out_json).expanduser().resolve()
        out_path.parent.mkdir(parents=True, exist_ok=True)
        out_path.write_text(rendered + "\n", encoding="utf-8")
    print(rendered)


if __name__ == "__main__":
    main()
