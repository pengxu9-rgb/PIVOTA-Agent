#!/usr/bin/env python3
"""Export an ingredient reference workbook into a seed_ingest-ready CSV bundle."""

from __future__ import annotations

import argparse
import csv
import hashlib
import json
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to export ingredient reference seed bundles. Install it in the local Python environment first."
    ) from exc


REQUIRED_COLUMNS = [
    "record_id",
    "canonical_inci_name",
    "us_label_name",
    "eu_label_name",
    "normalized_key",
    "parser_variants",
    "primary_bucket",
    "function_tags",
    "benefit_tags",
    "source_urls",
    "kb_version",
]

METADATA_COLUMNS = [
    "source_file",
    "source_sheet",
    "source_row_number",
    "ingested_at",
]


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        while True:
            chunk = handle.read(1024 * 1024)
            if not chunk:
                break
            digest.update(chunk)
    return digest.hexdigest()


def read_sheet(path: Path, sheet_name: str) -> tuple[list[str], list[tuple[int, dict[str, str]]]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Workbook missing required '{sheet_name}' sheet.")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        raise SystemExit(f"Sheet '{sheet_name}' is empty.")
    header = [normalize_text(cell) for cell in rows[0]]
    records: list[tuple[int, dict[str, str]]] = []
    for row_number, row in enumerate(rows[1:], start=2):
        if not any(normalize_text(cell) for cell in row):
            continue
        record = {
            key: normalize_text(row[index] if index < len(row) else "")
            for index, key in enumerate(header)
        }
        records.append((row_number, record))
    return header, records


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def build_copy_sql(target_table: str, csv_path: Path, fieldnames: list[str]) -> str:
    columns_sql = ",\n  ".join(fieldnames)
    csv_literal = str(csv_path).replace("'", "''")
    return (
        f"COPY {target_table} (\n"
        f"  {columns_sql}\n"
        f")\n"
        f"FROM '{csv_literal}'\n"
        "WITH (\n"
        "  FORMAT csv,\n"
        "  HEADER true,\n"
        "  ENCODING 'UTF8'\n"
        ");\n"
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Export an ingredient reference workbook into a seed_ingest-ready CSV bundle.")
    parser.add_argument("--ingredient-xlsx", required=True, help="Current ingredient workbook")
    parser.add_argument("--sheet-name", default="Dictionary", help="Workbook sheet to export (default: Dictionary)")
    parser.add_argument("--out-csv", required=True, help="Where to write the export CSV")
    parser.add_argument("--out-manifest-json", required=True, help="Where to write the export manifest JSON")
    parser.add_argument("--out-copy-sql", help="Optional path to write a COPY SQL template")
    parser.add_argument("--target-table", default="seed_ingest.ingredient_reference_seed", help="Target staging table")
    args = parser.parse_args()

    workbook_path = Path(args.ingredient_xlsx).expanduser().resolve()
    out_csv = Path(args.out_csv).expanduser().resolve()
    out_manifest = Path(args.out_manifest_json).expanduser().resolve()
    out_copy_sql = Path(args.out_copy_sql).expanduser().resolve() if args.out_copy_sql else None

    header, records = read_sheet(workbook_path, args.sheet_name)
    missing = [column for column in REQUIRED_COLUMNS if column not in header]
    if missing:
        raise SystemExit(f"Workbook sheet is missing required columns: {', '.join(missing)}")

    exported_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat().replace("+00:00", "Z")
    fieldnames = METADATA_COLUMNS + header

    csv_rows: list[dict[str, str]] = []
    for source_row_number, record in records:
        row = {
            "source_file": workbook_path.name,
            "source_sheet": args.sheet_name,
            "source_row_number": str(source_row_number),
            "ingested_at": exported_at,
        }
        row.update(record)
        csv_rows.append(row)

    write_csv(out_csv, fieldnames, csv_rows)

    if out_copy_sql:
        out_copy_sql.parent.mkdir(parents=True, exist_ok=True)
        out_copy_sql.write_text(build_copy_sql(args.target_table, out_csv, fieldnames), encoding="utf-8")

    manifest = {
        "source_workbook": str(workbook_path),
        "source_workbook_sha256": sha256_file(workbook_path),
        "source_sheet": args.sheet_name,
        "target_table": args.target_table,
        "row_count": len(csv_rows),
        "required_columns_validated": REQUIRED_COLUMNS,
        "exported_columns": fieldnames,
        "recommended_primary_key": ["record_id"],
        "recommended_secondary_unique_keys": [
            "normalized_key",
            "canonical_inci_name",
        ],
        "out_csv": str(out_csv),
        "out_copy_sql": str(out_copy_sql) if out_copy_sql else None,
        "exported_at": exported_at,
    }

    out_manifest.parent.mkdir(parents=True, exist_ok=True)
    out_manifest.write_text(json.dumps(manifest, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")

    print(json.dumps(manifest, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
