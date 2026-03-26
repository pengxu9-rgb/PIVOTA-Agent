#!/usr/bin/env python3
"""Export reviewed alias manual-mapping workbench rows into alias patches or new-canonical patches."""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to export alias manual-mapping resolutions. Install it in the local Python environment first."
    ) from exc


PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]

RECORD_ID_PATTERN = re.compile(r"^ing_patch_v13_(\d+)$", re.IGNORECASE)
PAREN_RE = re.compile(r"\([^)]*\)")


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    return "".join(ch for ch in raw if ch.isalnum())


def semantic_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    raw = PAREN_RE.sub(" ", raw)
    raw = " ".join(raw.split())
    return "".join(ch for ch in raw if ch.isalnum())


def canonical_name_key(value: Any) -> str:
    return " ".join(normalize_text(value).casefold().split())


def split_semicolon_values(value: Any) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    return [part.strip() for part in raw.split(";") if part and part.strip()]


def dedupe_join(values: list[str]) -> str:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        text = normalize_text(value)
        key = text.casefold()
        if not text or key in seen:
            continue
        seen.add(key)
        out.append(text)
    return "; ".join(out)


def dedupe_join_exact(values: list[str]) -> str:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        text = normalize_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        out.append(text)
    return "; ".join(out)


def load_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def resolve_ingredient_sheet_name(path: Path, requested_sheet: str | None) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise SystemExit(f"Ingredient workbook missing requested sheet '{requested_sheet}'.")
        return requested_sheet
    for sheet_name in PREFERRED_INGREDIENT_SHEETS:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    raise SystemExit(
        f"Ingredient workbook did not contain any supported sheet names ({', '.join(PREFERRED_INGREDIENT_SHEETS)})."
    )


def read_ingredient_header_and_max_patch(path: Path, sheet_name: str) -> tuple[list[str], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    max_patch = 0
    record_id_index = header.index("record_id")
    for row in rows[1:]:
        record_id = normalize_text(row[record_id_index] if record_id_index < len(row) else "")
        match = RECORD_ID_PATTERN.match(record_id)
        if match:
            max_patch = max(max_patch, int(match.group(1)))
    return header, max_patch


def set_if_present(row: dict[str, str], header: list[str], key: str, value: Any) -> None:
    if key in header:
        row[key] = normalize_text(value)


def build_new_canonical_patch_row(header: list[str], workbench_rows: list[dict[str, str]], record_id: str) -> dict[str, str]:
    canonical_name = normalize_text(workbench_rows[0].get("approved_new_canonical_inci_name"))
    raw_tokens = dedupe_join([row.get("raw_token", "") for row in workbench_rows])
    parser_variants = dedupe_join_exact(
        [canonical_name]
        + [
            value
            for row in workbench_rows
            for value in split_semicolon_values(row.get("approved_parser_variants_addition"))
        ]
        + split_semicolon_values(raw_tokens)
    )
    source_urls = dedupe_join(
        [
            value
            for row in workbench_rows
            for value in split_semicolon_values(row.get("example_urls"))
        ]
    )
    reviewer_notes = dedupe_join([row.get("reviewer_notes", "") for row in workbench_rows])
    subtype_notes = dedupe_join([row.get("manual_mapping_subtype", "") for row in workbench_rows])
    example_brands = dedupe_join([row.get("example_brands", "") for row in workbench_rows])
    example_products = dedupe_join([row.get("example_products", "") for row in workbench_rows])
    priority_scores = [int(normalize_text(row.get("priority_score")) or "0") for row in workbench_rows]
    row = {column: "" for column in header}
    set_if_present(row, header, "record_id", record_id)
    set_if_present(row, header, "canonical_inci_name", canonical_name)
    set_if_present(row, header, "canonical_display_name", canonical_name)
    set_if_present(row, header, "ingredient_family", "other")
    set_if_present(row, header, "us_label_name", canonical_name)
    set_if_present(row, header, "eu_label_name", canonical_name)
    set_if_present(row, header, "us_label_variants", canonical_name)
    set_if_present(row, header, "eu_label_variants", canonical_name)
    set_if_present(row, header, "normalized_key", normalize_key(canonical_name))
    set_if_present(row, header, "aliases_common", "")
    set_if_present(row, header, "parser_variants", parser_variants)
    set_if_present(row, header, "regulatory_bucket", "patch_candidate_review")
    set_if_present(row, header, "source_urls", source_urls)
    set_if_present(row, header, "source_authorities", "brand_official_pdp")
    set_if_present(row, header, "source_types", "official_brand_site")
    set_if_present(row, header, "review_status", "draft")
    set_if_present(row, header, "confidence", "low")
    set_if_present(
        row,
        header,
        "review_notes",
        dedupe_join(
            [
                "Generated from alias manual-mapping workbench; reviewer chose create_new_canonical.",
                reviewer_notes,
            ]
        ),
    )
    set_if_present(row, header, "notes", f"manual_mapping_subtype={subtype_notes}")
    set_if_present(row, header, "kb_version", "spec_v1_patch_v13_alias_manual_mapping")
    for boolean_column in [
        "is_humectant",
        "is_barrier_support",
        "is_retinoid",
        "is_exfoliant",
        "is_uv_filter",
        "is_preservative",
        "is_surfactant",
        "is_fragrance_or_eo",
    ]:
        set_if_present(row, header, boolean_column, "no")

    row["queue_priority_score"] = str(max(priority_scores) if priority_scores else 0)
    row["queue_example_brands"] = example_brands
    row["queue_example_products"] = example_products
    row["queue_example_urls"] = source_urls
    row["source_packet_resolution"] = "manual_mapping_create_new_canonical"
    row["source_packet_confidence"] = "reviewed_manual_mapping"
    row["source_packet_raw_token"] = raw_tokens
    row["semantic_match_key"] = semantic_key(canonical_name)
    return row


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Export reviewed alias manual-mapping workbench rows into alias patches or new-canonical patches.")
    parser.add_argument("--decision-csv", required=True, help="Reviewed alias manual-mapping workbench CSV")
    parser.add_argument("--ingredient-xlsx", required=True, help="Current ingredient workbook")
    parser.add_argument("--ingredient-sheet", help="Optional ingredient workbook sheet")
    parser.add_argument("--out-alias-apply-csv", required=True, help="Where to write apply-ready alias patch CSV")
    parser.add_argument("--out-new-canonical-csv", required=True, help="Where to write apply-ready new-canonical patch CSV")
    parser.add_argument("--out-remainder-csv", required=True, help="Where to write non-exported remainder CSV")
    parser.add_argument("--out-summary-json", required=True, help="Where to write summary JSON")
    args = parser.parse_args()

    decision_csv = Path(args.decision_csv).expanduser().resolve()
    ingredient_xlsx = Path(args.ingredient_xlsx).expanduser().resolve()
    ingredient_sheet = resolve_ingredient_sheet_name(ingredient_xlsx, args.ingredient_sheet)
    header, max_patch = read_ingredient_header_and_max_patch(ingredient_xlsx, ingredient_sheet)

    rows = load_rows(decision_csv)
    alias_groups: dict[tuple[str, str, str], list[dict[str, str]]] = {}
    new_canonical_groups: dict[str, list[dict[str, str]]] = {}
    alias_rows: list[dict[str, str]] = []
    new_canonical_rows: list[dict[str, str]] = []
    remainder_rows: list[dict[str, str]] = []
    next_patch_number = max_patch + 1

    for row in rows:
        decision = normalize_text(row.get("decision"))

        if decision == "map_to_existing_canonical":
            target_record_id = normalize_text(row.get("approved_existing_target_record_id"))
            target_canonical = normalize_text(row.get("approved_existing_target_canonical_inci_name"))
            alias_quality = normalize_text(row.get("approved_alias_quality")) or "exact_label_alias"
            if target_record_id and target_canonical and normalize_text(row.get("raw_token")):
                alias_groups.setdefault((target_record_id, target_canonical, alias_quality), []).append(row)
                continue

        if decision == "create_new_canonical":
            canonical_name = normalize_text(row.get("approved_new_canonical_inci_name"))
            if canonical_name:
                new_canonical_groups.setdefault(canonical_name_key(canonical_name), []).append(row)
                continue

        remainder_rows.append(row)

    for (target_record_id, target_canonical, alias_quality), grouped_rows in alias_groups.items():
        existing_aliases = normalize_text(grouped_rows[0].get("suggested_existing_aliases_common"))
        existing_parser_variants = normalize_text(grouped_rows[0].get("suggested_existing_parser_variants"))
        existing_alias_quality = normalize_text(grouped_rows[0].get("suggested_existing_alias_quality"))
        raw_tokens = [normalize_text(row.get("raw_token")) for row in grouped_rows]
        approved_variants = [
            value
            for row in grouped_rows
            for value in split_semicolon_values(row.get("approved_parser_variants_addition"))
        ]
        alias_rows.append(
            {
                "record_id": target_record_id,
                "canonical_inci_name": target_canonical,
                "existing_aliases_common": existing_aliases,
                "existing_parser_variants": existing_parser_variants,
                "existing_alias_quality": existing_alias_quality,
                "patch_aliases_common": dedupe_join(split_semicolon_values(existing_aliases) + raw_tokens),
                "patch_parser_variants": dedupe_join_exact(split_semicolon_values(existing_parser_variants) + approved_variants + raw_tokens),
                "patch_alias_quality": alias_quality,
                "proposal_sources": "alias_manual_mapping_workbench",
                "quality_reason": dedupe_join(
                    [
                        row.get("reviewer_notes", "") or row.get("suggestion_rationale", "")
                        for row in grouped_rows
                    ]
                ),
            }
        )

    for grouped_rows in new_canonical_groups.values():
        record_id = f"ing_patch_v13_{next_patch_number:03d}"
        next_patch_number += 1
        new_canonical_rows.append(build_new_canonical_patch_row(header, grouped_rows, record_id))

    alias_fieldnames = [
        "record_id",
        "canonical_inci_name",
        "existing_aliases_common",
        "existing_parser_variants",
        "existing_alias_quality",
        "patch_aliases_common",
        "patch_parser_variants",
        "patch_alias_quality",
        "proposal_sources",
        "quality_reason",
    ]
    new_canonical_fieldnames = header + [
        "queue_priority_score",
        "queue_example_brands",
        "queue_example_products",
        "queue_example_urls",
        "source_packet_resolution",
        "source_packet_confidence",
        "source_packet_raw_token",
        "semantic_match_key",
    ]
    remainder_fieldnames = list(rows[0].keys()) if rows else []

    out_alias_apply_csv = Path(args.out_alias_apply_csv).expanduser().resolve()
    out_new_canonical_csv = Path(args.out_new_canonical_csv).expanduser().resolve()
    out_remainder_csv = Path(args.out_remainder_csv).expanduser().resolve()
    out_summary_json = Path(args.out_summary_json).expanduser().resolve()

    write_csv(out_alias_apply_csv, alias_fieldnames, alias_rows)
    write_csv(out_new_canonical_csv, new_canonical_fieldnames, new_canonical_rows)
    write_csv(out_remainder_csv, remainder_fieldnames, remainder_rows)

    summary = {
        "decision_csv": str(decision_csv),
        "ingredient_workbook": str(ingredient_xlsx),
        "ingredient_sheet": ingredient_sheet,
        "alias_apply_ready_count": len(alias_rows),
        "new_canonical_apply_ready_count": len(new_canonical_rows),
        "remainder_count": len(remainder_rows),
        "out_alias_apply_csv": str(out_alias_apply_csv),
        "out_new_canonical_csv": str(out_new_canonical_csv),
        "out_remainder_csv": str(out_remainder_csv),
    }
    out_summary_json.parent.mkdir(parents=True, exist_ok=True)
    out_summary_json.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
