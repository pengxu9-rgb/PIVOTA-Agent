#!/usr/bin/env python3
"""Build a decision-ready review packet from Alias_Normalization queue rows."""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build ingredient alias normalization packets. Install it in the local Python environment first."
    ) from exc


PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]

TERM_FIELDS = [
    "canonical_inci_name",
    "canonical_display_name",
    "us_label_name",
    "eu_label_name",
    "parser_variants",
    "aliases_common",
    "us_label_variants",
    "eu_label_variants",
    "deprecated_aliases",
]

QUEUE_FIELDS = [
    "priority_score",
    "recommended_bucket",
    "recommended_action",
    "triage_reason",
    "raw_token",
    "normalized_token",
    "unmatched_count",
    "sku_row_count",
    "full_inci_count",
    "key_count",
    "product_only_count",
    "top_categories",
    "example_brands",
    "example_products",
    "example_urls",
    "in_current_master_like",
]

PACKET_FIELDS = [
    "priority_score",
    "raw_token",
    "normalized_token",
    "unmatched_count",
    "sku_row_count",
    "full_inci_count",
    "key_count",
    "top_categories",
    "example_brands",
    "example_products",
    "example_urls",
    "suggested_resolution",
    "suggestion_confidence",
    "existing_target_record_id",
    "existing_target_canonical_inci_name",
    "existing_aliases_common",
    "existing_parser_variants",
    "suggested_new_canonical_inci_name",
    "suggested_aliases_common_addition",
    "suggested_alias_quality",
    "suggested_parser_variants_addition",
    "resolution_rationale",
    "decision",
    "approved_target_record_id",
    "approved_new_canonical_inci_name",
    "approved_aliases_common_addition",
    "approved_alias_quality",
    "approved_parser_variants_addition",
    "reviewer_notes",
]

MANUAL_PATTERN = re.compile(r"(?:^|[^a-z])(ci\s*\d+|edta|vitamin\s+[a-z0-9]+|beeswax\s*/|/|%|\(|\)|\b(?:hcl|hci)\b)", re.IGNORECASE)
SHORT_ACRONYM_PATTERN = re.compile(r"^[A-Z0-9-]{2,4}$")
PAREN_RE = re.compile(r"\([^)]*\)")


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    return "".join(ch for ch in raw if ch.isalnum())


def semantic_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    raw = PAREN_RE.sub(" ", raw)
    raw = " ".join(raw.split())
    return "".join(ch for ch in raw if ch.isalnum())


def split_semicolon_values(value: Any) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    return [part.strip() for part in raw.split(";") if part and part.strip()]


def dedupe_join(values: list[str]) -> str:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        text = normalize_text(value)
        key = text.casefold()
        if not text or key in seen:
            continue
        seen.add(key)
        out.append(text)
    return "; ".join(out)


def resolve_ingredient_sheet_name(path: Path, requested_sheet: str | None) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise SystemExit(f"Ingredient workbook missing requested sheet '{requested_sheet}'.")
        return requested_sheet
    for sheet_name in PREFERRED_INGREDIENT_SHEETS:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    raise SystemExit(
        f"Ingredient workbook did not contain any supported sheet names ({', '.join(PREFERRED_INGREDIENT_SHEETS)})."
    )


def read_queue_rows(path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Queue workbook missing required sheet '{sheet_name}'.")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    missing = [field for field in QUEUE_FIELDS if field not in header]
    if missing:
        raise SystemExit(f"Queue sheet missing required columns: {', '.join(missing)}")
    out: list[dict[str, str]] = []
    for row in rows[1:]:
        if not any(normalize_text(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        out.append({header[index]: normalize_text(padded[index]) for index in range(len(header))})
    return out


def build_ingredient_indexes(path: Path, sheet_name: str) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]

    normalized_index: dict[str, list[dict[str, str]]] = defaultdict(list)
    semantic_index: dict[str, list[dict[str, str]]] = defaultdict(list)

    for row in rows[1:]:
        if not any(normalize_text(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        record = {header[index]: normalize_text(padded[index]) for index in range(len(header))}
        base = {
            "record_id": record.get("record_id", ""),
            "canonical_inci_name": record.get("canonical_inci_name", ""),
            "aliases_common": record.get("aliases_common", ""),
            "parser_variants": record.get("parser_variants", ""),
        }
        for field in TERM_FIELDS:
            values = split_semicolon_values(record.get(field)) if field.endswith("variants") or field in {"parser_variants", "aliases_common", "deprecated_aliases"} else [record.get(field, "")]
            for value in values:
                normalized = normalize_key(value)
                if normalized:
                    normalized_index[normalized].append(base)
                semantic = semantic_key(value)
                if semantic:
                    semantic_index[semantic].append(base)

    return {
        "normalized_index": normalized_index,
        "semantic_index": semantic_index,
    }


def smart_title_case_token(token: str) -> str:
    special_tokens = {
        "CI": "CI",
        "PEG": "PEG",
        "PPG": "PPG",
        "PCA": "PCA",
        "DNA": "DNA",
        "RNA": "RNA",
        "AHA": "AHA",
        "BHA": "BHA",
        "PHA": "PHA",
        "HCL": "HCl",
    }

    def transform_piece(piece: str) -> str:
        if not piece:
            return piece
        upper = piece.upper()
        if upper in special_tokens:
            return special_tokens[upper]
        if piece.isdigit():
            return piece
        if re.fullmatch(r"[A-Z0-9]+", piece):
            return piece[0].upper() + piece[1:].lower()
        return piece[0].upper() + piece[1:]

    words: list[str] = []
    for word in token.split():
        if "/" in word:
            words.append("/".join(transform_piece(part) for part in word.split("/")))
            continue
        if "-" in word:
            words.append("-".join(transform_piece(part) for part in word.split("-")))
            continue
        words.append(transform_piece(word))
    return " ".join(words)


def should_route_manual(token: str) -> bool:
    text = normalize_text(token)
    if not text:
        return True
    if MANUAL_PATTERN.search(text):
        return True
    if SHORT_ACRONYM_PATTERN.fullmatch(text):
        return True
    return False


def build_packet_row(queue_row: dict[str, str], ingredient_indexes: dict[str, Any]) -> dict[str, str]:
    raw_token = normalize_text(queue_row.get("raw_token"))
    normalized = normalize_key(queue_row.get("normalized_token") or raw_token)
    semantic = semantic_key(raw_token)

    normalized_hits = ingredient_indexes["normalized_index"].get(normalized, [])
    semantic_hits = ingredient_indexes["semantic_index"].get(semantic, [])
    exact_target = normalized_hits[0] if len({row["record_id"] for row in normalized_hits}) == 1 and normalized_hits else None
    semantic_target = semantic_hits[0] if len({row["record_id"] for row in semantic_hits}) == 1 and semantic_hits else None
    target = exact_target or semantic_target

    suggested_new_canonical = ""
    suggested_aliases_common = ""
    suggested_alias_quality = ""
    suggested_parser_variants = ""
    suggested_resolution = ""
    suggestion_confidence = ""
    resolution_rationale = ""

    if target:
        suggested_resolution = "attach_parser_variant_to_existing_canonical"
        suggestion_confidence = "high"
        suggested_parser_variants = raw_token
        resolution_rationale = "Deterministic normalized or semantic lookup found a single existing canonical target."
    else:
        suggested_new_canonical = smart_title_case_token(raw_token)
        suggested_parser_variants = dedupe_join([suggested_new_canonical, raw_token, suggested_new_canonical.lower()])
        if should_route_manual(raw_token):
            suggested_resolution = "needs_manual_mapping"
            suggestion_confidence = "low"
            resolution_rationale = "Token still lacks a deterministic existing target and contains abbreviation, slash, CI code, parentheses, or other risky normalization cues."
        else:
            suggested_resolution = "new_canonical_candidate_with_parser_variants"
            suggestion_confidence = "medium" if raw_token != suggested_new_canonical else "high"
            resolution_rationale = "Token looks like a label-form or case variant but no deterministic current canonical target exists in v2.2."

    return {
        "priority_score": normalize_text(queue_row.get("priority_score")),
        "raw_token": raw_token,
        "normalized_token": normalized,
        "unmatched_count": normalize_text(queue_row.get("unmatched_count")),
        "sku_row_count": normalize_text(queue_row.get("sku_row_count")),
        "full_inci_count": normalize_text(queue_row.get("full_inci_count")),
        "key_count": normalize_text(queue_row.get("key_count")),
        "top_categories": normalize_text(queue_row.get("top_categories")),
        "example_brands": normalize_text(queue_row.get("example_brands")),
        "example_products": normalize_text(queue_row.get("example_products")),
        "example_urls": normalize_text(queue_row.get("example_urls")),
        "suggested_resolution": suggested_resolution,
        "suggestion_confidence": suggestion_confidence,
        "existing_target_record_id": normalize_text((target or {}).get("record_id")),
        "existing_target_canonical_inci_name": normalize_text((target or {}).get("canonical_inci_name")),
        "existing_aliases_common": normalize_text((target or {}).get("aliases_common")),
        "existing_parser_variants": normalize_text((target or {}).get("parser_variants")),
        "suggested_new_canonical_inci_name": suggested_new_canonical,
        "suggested_aliases_common_addition": suggested_aliases_common,
        "suggested_alias_quality": suggested_alias_quality,
        "suggested_parser_variants_addition": suggested_parser_variants,
        "resolution_rationale": resolution_rationale,
        "decision": "",
        "approved_target_record_id": normalize_text((target or {}).get("record_id")),
        "approved_new_canonical_inci_name": suggested_new_canonical,
        "approved_aliases_common_addition": suggested_aliases_common,
        "approved_alias_quality": suggested_alias_quality,
        "approved_parser_variants_addition": suggested_parser_variants,
        "reviewer_notes": "",
    }


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=PACKET_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(path: Path, rows: list[dict[str, str]]) -> None:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Alias_Normalization_Packet"
    sheet.append(PACKET_FIELDS)
    for row in rows:
        sheet.append([row.get(field, "") for field in PACKET_FIELDS])
    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a decision-ready packet from Alias_Normalization queue rows.")
    parser.add_argument("--queue-xlsx", required=True, help="Alias normalization queue workbook")
    parser.add_argument("--ingredient-xlsx", required=True, help="Current ingredient workbook")
    parser.add_argument("--queue-sheet", default="Alias_Normalization_Queue", help="Queue sheet to read")
    parser.add_argument("--ingredient-sheet", help="Optional ingredient sheet")
    parser.add_argument("--out-csv", required=True, help="Where to write the decision-ready packet CSV")
    parser.add_argument("--out-json", required=True, help="Where to write summary JSON")
    parser.add_argument("--out-xlsx", help="Optional workbook copy of the packet")
    args = parser.parse_args()

    queue_xlsx = Path(args.queue_xlsx).expanduser().resolve()
    ingredient_xlsx = Path(args.ingredient_xlsx).expanduser().resolve()
    out_csv = Path(args.out_csv).expanduser().resolve()
    out_json = Path(args.out_json).expanduser().resolve()
    out_xlsx = Path(args.out_xlsx).expanduser().resolve() if args.out_xlsx else None

    ingredient_sheet = resolve_ingredient_sheet_name(ingredient_xlsx, args.ingredient_sheet)
    queue_rows = read_queue_rows(queue_xlsx, args.queue_sheet)
    ingredient_indexes = build_ingredient_indexes(ingredient_xlsx, ingredient_sheet)

    packet_rows = [build_packet_row(row, ingredient_indexes) for row in queue_rows]
    write_csv(out_csv, packet_rows)
    if out_xlsx:
        write_xlsx(out_xlsx, packet_rows)

    resolution_counts = Counter(row["suggested_resolution"] for row in packet_rows)
    confidence_counts = Counter(row["suggestion_confidence"] for row in packet_rows)
    summary = {
        "queue_workbook": str(queue_xlsx),
        "queue_sheet": args.queue_sheet,
        "ingredient_workbook": str(ingredient_xlsx),
        "ingredient_sheet": ingredient_sheet,
        "row_count": len(packet_rows),
        "resolution_counts": dict(sorted(resolution_counts.items())),
        "confidence_counts": dict(sorted(confidence_counts.items())),
        "out_csv": str(out_csv),
        "out_xlsx": str(out_xlsx) if out_xlsx else None,
    }

    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
