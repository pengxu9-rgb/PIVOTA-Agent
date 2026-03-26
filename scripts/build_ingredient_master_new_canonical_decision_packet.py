#!/usr/bin/env python3
"""Build a researched decision packet for new ingredient-master candidates."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build the ingredient-master new-canonical decision packet. Install it in the local Python environment first."
    ) from exc


OUTPUT_FIELDS = [
    "signal_key",
    "display_signal_name",
    "decision_status",
    "recommended_master_action",
    "recommended_candidate_display_name",
    "recommended_candidate_inci_name",
    "ingredient_master_ready",
    "canonical_writeback_ready",
    "runtime_evidence_eligible",
    "decision_basis",
    "source_urls",
    "example_brands",
    "example_products",
    "reviewer_notes",
]


DECISIONS: dict[str, dict[str, str]] = {
    "viniferine": {
        "decision_status": "approved_rewrite_to_underlying_inci_candidate",
        "recommended_master_action": "open_new_canonical_candidate__palmitoyl_grapevine_shoot_extract",
        "recommended_candidate_display_name": "Palmitoyl Grapevine Shoot Extract",
        "recommended_candidate_inci_name": "Palmitoyl Grapevine Shoot Extract",
        "ingredient_master_ready": "candidate_only",
        "canonical_writeback_ready": "no",
        "runtime_evidence_eligible": "no",
        "decision_basis": "Official Caudalie PDP markets Viniferine but the ingredient list uses Palmitoyl Grapevine Shoot Extract.",
        "source_urls": "https://us.caudalie.com/p/324C/324c-radiance-dark-spot-serum-vitamin-c-alternative.html",
        "reviewer_notes": "Do not open a canonical row called Viniferine. If promoted, promote the underlying INCI candidate instead and keep Viniferine as a branded-active alias clue only.",
    },
    "chlorella": {
        "decision_status": "approved_rewrite_to_underlying_inci_candidate",
        "recommended_master_action": "open_new_canonical_candidate__chlorella_vulgaris_extract",
        "recommended_candidate_display_name": "Chlorella Vulgaris Extract",
        "recommended_candidate_inci_name": "Chlorella Vulgaris Extract",
        "ingredient_master_ready": "candidate_only",
        "canonical_writeback_ready": "no",
        "runtime_evidence_eligible": "no",
        "decision_basis": "Official Elemis PDP ingredient naming points to Chlorella Vulgaris Extract rather than a generic Chlorella canonical row.",
        "source_urls": "https://au.elemis.com/c/procollagen/marine-cream/",
        "reviewer_notes": "Promote the likely extract form, not the generic organism name.",
    },
    "padina_pavonica": {
        "decision_status": "approved_rewrite_to_underlying_inci_candidate",
        "recommended_master_action": "open_new_canonical_candidate__padina_pavonica_thallus_extract",
        "recommended_candidate_display_name": "Padina Pavonica Thallus Extract",
        "recommended_candidate_inci_name": "Padina Pavonica Thallus Extract",
        "ingredient_master_ready": "candidate_only",
        "canonical_writeback_ready": "no",
        "runtime_evidence_eligible": "no",
        "decision_basis": "Official Elemis PDP ingredient naming points to Padina Pavonica Thallus Extract rather than a generic Padina Pavonica row.",
        "source_urls": "https://au.elemis.com/c/procollagen/marine-cream/",
        "reviewer_notes": "Promote the thallus-extract form, not the shorthand marine botanical phrase.",
    },
    "thiamidol": {
        "decision_status": "approved_hold_named_active_pending_exact_label_confirmation",
        "recommended_master_action": "keep_signal_only_until_exact_inci_confirmed",
        "recommended_candidate_display_name": "Thiamidol",
        "recommended_candidate_inci_name": "",
        "ingredient_master_ready": "hold",
        "canonical_writeback_ready": "no",
        "runtime_evidence_eligible": "no",
        "decision_basis": "Official Eucerin ingredient pages confirm Thiamidol as a real branded active, but this packet does not yet lock a safe ingredient-master INCI row.",
        "source_urls": "https://int.eucerin.com/our-research/ingredients/thiamidol; https://int.eucerin.com/products/anti-pigment/day-cream-spf-30; https://int.eucerin.com/products/anti-pigment/spot-corrector",
        "reviewer_notes": "Keep as signal/named-active until exact label-form mapping is explicitly confirmed from source ingredients.",
    },
    "licochalcone_a": {
        "decision_status": "approved_hold_named_active_pending_exact_label_confirmation",
        "recommended_master_action": "keep_signal_only_until_exact_inci_confirmed",
        "recommended_candidate_display_name": "Licochalcone A",
        "recommended_candidate_inci_name": "",
        "ingredient_master_ready": "hold",
        "canonical_writeback_ready": "no",
        "runtime_evidence_eligible": "no",
        "decision_basis": "Official Eucerin ingredient page confirms Licochalcone A as a named active, but this packet does not yet lock a safe canonical ingredient row.",
        "source_urls": "https://int.eucerin.com/our-research/ingredients/licochalcone-a; https://int.eucerin.com/products/anti-pigment/day-cream-spf-30",
        "reviewer_notes": "Keep as signal/named-active until exact label-form mapping is explicitly confirmed from source ingredients.",
    },
    "barbados_cherry": {
        "decision_status": "approved_rewrite_to_underlying_inci_candidate",
        "recommended_master_action": "open_new_canonical_candidate__malpighia_punicifolia_fruit_extract",
        "recommended_candidate_display_name": "Malpighia Punicifolia Fruit Extract",
        "recommended_candidate_inci_name": "Malpighia Punicifolia Fruit Extract",
        "ingredient_master_ready": "candidate_only",
        "canonical_writeback_ready": "no",
        "runtime_evidence_eligible": "no",
        "decision_basis": "Official Fenty product page markets Barbados Cherry and lists MALPIGHIA PUNICIFOLIA FRUIT EXTRACT in ingredients.",
        "source_urls": "https://fentybeauty.com/en-gm/products/total-cleansr-remove-it-all-cleanser-with-barbados-cherry",
        "reviewer_notes": "Promote the underlying fruit-extract INCI candidate, not the consumer-facing botanical nickname.",
    },
    "pansy": {
        "decision_status": "approved_rewrite_to_underlying_inci_candidate",
        "recommended_master_action": "open_new_canonical_candidate__viola_tricolor_extract",
        "recommended_candidate_display_name": "Viola Tricolor Extract",
        "recommended_candidate_inci_name": "Viola Tricolor Extract",
        "ingredient_master_ready": "candidate_only",
        "canonical_writeback_ready": "no",
        "runtime_evidence_eligible": "no",
        "decision_basis": "Official Weleda Skin Food ingredient list uses Viola Tricolor Extract rather than a generic Pansy canonical row.",
        "source_urls": "https://www.weleda.com/product/skin-food-original-ultra-rich-cream-g009398",
        "reviewer_notes": "Promote the botanical extract form, not the common-language plant name.",
    },
}


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def split_multi(value: str) -> list[str]:
    return [part.strip() for part in normalize_text(value).split(";") if part.strip()]


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return [{key: normalize_text(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def append_sheet(workbook: Workbook, title: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])


def append_summary_sheet(workbook: Workbook, summary: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Summary")
    sheet.append(["metric", "value"])
    for key, value in summary.items():
        if isinstance(value, (dict, list)):
            sheet.append([key, json.dumps(value, ensure_ascii=True, sort_keys=True)])
        else:
            sheet.append([key, value])


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a researched decision packet for new ingredient-master candidates.")
    parser.add_argument("--terms-csv", required=True, help="New canonical candidate term CSV")
    parser.add_argument("--row-followup-csv", required=True, help="Row-level followup CSV for URL context")
    parser.add_argument("--out-csv", required=True, help="Where to write the decision packet CSV")
    parser.add_argument("--out-summary-json", required=True, help="Where to write summary JSON")
    parser.add_argument("--out-xlsx", help="Optional workbook output")
    args = parser.parse_args()

    term_rows = read_csv_rows(Path(args.terms_csv).expanduser().resolve())
    row_rows = read_csv_rows(Path(args.row_followup_csv).expanduser().resolve())

    row_urls_by_key: dict[str, list[str]] = {}
    for row in row_rows:
        keys = split_multi(row.get("term_keys", ""))
        for key in keys:
            row_urls_by_key.setdefault(key, [])
            url = normalize_text(row.get("official_product_url"))
            if url and url not in row_urls_by_key[key]:
                row_urls_by_key[key].append(url)

    outputs: list[dict[str, str]] = []
    missing_decisions: list[str] = []
    for term in term_rows:
        signal_key = normalize_text(term.get("signal_key"))
        decision = DECISIONS.get(signal_key)
        if not decision:
            missing_decisions.append(signal_key)
            continue
        source_urls = split_multi(decision["source_urls"]) + row_urls_by_key.get(signal_key, [])
        unique_urls: list[str] = []
        for url in source_urls:
            if url not in unique_urls:
                unique_urls.append(url)
        outputs.append(
            {
                "signal_key": signal_key,
                "display_signal_name": normalize_text(term.get("display_signal_name")),
                "decision_status": decision["decision_status"],
                "recommended_master_action": decision["recommended_master_action"],
                "recommended_candidate_display_name": decision["recommended_candidate_display_name"],
                "recommended_candidate_inci_name": decision["recommended_candidate_inci_name"],
                "ingredient_master_ready": decision["ingredient_master_ready"],
                "canonical_writeback_ready": decision["canonical_writeback_ready"],
                "runtime_evidence_eligible": decision["runtime_evidence_eligible"],
                "decision_basis": decision["decision_basis"],
                "source_urls": "; ".join(unique_urls),
                "example_brands": normalize_text(term.get("example_brands")),
                "example_products": normalize_text(term.get("example_products")),
                "reviewer_notes": decision["reviewer_notes"],
            }
        )

    outputs.sort(key=lambda row: (row["decision_status"], row["display_signal_name"].casefold()))
    out_csv = Path(args.out_csv).expanduser().resolve()
    write_csv(out_csv, OUTPUT_FIELDS, outputs)

    summary = {
        "candidate_count": len(term_rows),
        "decision_count": len(outputs),
        "missing_decision_count": len(missing_decisions),
        "missing_decision_keys": missing_decisions,
        "decision_status_counts": dict(Counter(row["decision_status"] for row in outputs)),
        "ingredient_master_ready_counts": dict(Counter(row["ingredient_master_ready"] for row in outputs)),
        "out_csv": str(out_csv),
    }

    if args.out_xlsx:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        append_summary_sheet(workbook, summary)
        append_sheet(workbook, "Decision_Packet", OUTPUT_FIELDS, outputs)
        out_xlsx = Path(args.out_xlsx).expanduser().resolve()
        out_xlsx.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(out_xlsx)
        summary["out_xlsx"] = str(out_xlsx)

    out_summary = Path(args.out_summary_json).expanduser().resolve()
    out_summary.parent.mkdir(parents=True, exist_ok=True)
    out_summary.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
