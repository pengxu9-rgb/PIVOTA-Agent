#!/usr/bin/env python3
"""Carry forward reviewed manual queue decisions onto a newer manual queue workbook."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to prefill manual review workbooks. Install it in the local Python environment first."
    ) from exc


OLD_REVIEWED_SHEET = "Manual_Queue_Reviewed"

REVIEWED_FIELDS = [
    "source_packet",
    "review_priority_score",
    "brand_name",
    "product_name",
    "official_product_url",
    "market",
    "category",
    "sku_row_key",
    "recommended_review_action",
    "recommended_review_reason",
    "token_count",
    "ingredient_match_count",
    "signal_match_count",
    "parser_cleanup_count",
    "curated_signal_tail_count",
    "parser_fragment_exclusion_count",
    "canonical_ingredients",
    "signal_display_names",
    "signal_keys",
    "suggested_decision",
    "suggested_follow_up",
    "suggestion_confidence",
    "decision_rationale",
    "decision",
    "approved_follow_up",
    "reviewer_notes",
    "approved_decision",
    "trust_tier",
    "reviewer_notes_auto",
    "review_status",
]

REVIEW_ONLY_FIELDS = [
    "decision",
    "approved_follow_up",
    "reviewer_notes",
    "approved_decision",
    "trust_tier",
    "reviewer_notes_auto",
    "review_status",
]


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def read_reviewed_rows(path: Path, sheet_name: str) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Workbook missing required sheet: {sheet_name}")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = [normalize_text(cell) for cell in rows[0]]
    out: list[dict[str, str]] = []
    for row in rows[1:]:
        if not any(normalize_text(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        out.append({header[index]: normalize_text(padded[index]) for index in range(len(header))})
    return out


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return [{key: normalize_text(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def semantic_key(row: dict[str, str]) -> tuple[str, str, str, str, str]:
    return (
        normalize_text(row.get("sku_row_key")),
        normalize_text(row.get("recommended_review_action")),
        normalize_text(row.get("canonical_ingredients")),
        normalize_text(row.get("signal_display_names")),
        normalize_text(row.get("signal_keys")),
    )


def sku_action_key(row: dict[str, str]) -> tuple[str, str]:
    return (
        normalize_text(row.get("sku_row_key")),
        normalize_text(row.get("recommended_review_action")),
    )


def build_row(base_row: dict[str, str], carried_row: dict[str, str] | None) -> dict[str, str]:
    out = {field: normalize_text(base_row.get(field)) for field in REVIEWED_FIELDS if field not in REVIEW_ONLY_FIELDS}
    for field in REVIEW_ONLY_FIELDS:
        out[field] = normalize_text(carried_row.get(field)) if carried_row else ""
    return out


def append_sheet(workbook: Workbook, title: str, rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(REVIEWED_FIELDS)
    for row in rows:
        sheet.append([row.get(field, "") for field in REVIEWED_FIELDS])


def append_summary_sheet(workbook: Workbook, summary_rows: list[tuple[str, str]]) -> None:
    sheet = workbook.create_sheet("Summary_Prefilled")
    sheet.append(["metric", "value"])
    for row in summary_rows:
        sheet.append(list(row))


def main() -> None:
    parser = argparse.ArgumentParser(description="Carry forward reviewed manual queue decisions onto a newer manual queue workbook.")
    parser.add_argument("--old-reviewed-xlsx", required=True, help="Older reviewed manual queue workbook")
    parser.add_argument("--new-manual-queue-csv", required=True, help="New manual handoff queue CSV")
    parser.add_argument("--out-xlsx", required=True, help="Where to write prefilled workbook")
    parser.add_argument("--out-summary-json", required=True, help="Where to write summary JSON")
    args = parser.parse_args()

    old_reviewed_rows = read_reviewed_rows(Path(args.old_reviewed_xlsx).expanduser().resolve(), OLD_REVIEWED_SHEET)
    new_queue_rows = read_csv_rows(Path(args.new_manual_queue_csv).expanduser().resolve())

    exact_lookup = {semantic_key(row): row for row in old_reviewed_rows}
    sku_action_lookup = {sku_action_key(row): row for row in old_reviewed_rows}

    carried_rows: list[dict[str, str]] = []
    needs_review_rows: list[dict[str, str]] = []
    carry_action_counts: Counter[str] = Counter()
    review_reason_counts: Counter[str] = Counter()

    for row in new_queue_rows:
      exact = exact_lookup.get(semantic_key(row))
      if exact:
          out_row = build_row(row, exact)
          carried_rows.append(out_row)
          carry_action_counts[normalize_text(row.get("recommended_review_action"))] += 1
          continue

      out_row = build_row(row, None)
      if sku_action_key(row) in sku_action_lookup:
          out_row["reviewer_notes_auto"] = "needs_manual_review_context_changed"
          review_reason_counts["context_changed_same_sku_action"] += 1
      else:
          out_row["reviewer_notes_auto"] = "needs_manual_review_new_or_reclassified_row"
          review_reason_counts["new_or_reclassified_row"] += 1
      needs_review_rows.append(out_row)

    carried_rows.sort(key=lambda row: (row["recommended_review_action"], row["brand_name"].casefold(), row["product_name"].casefold()))
    needs_review_rows.sort(key=lambda row: (row["recommended_review_action"], row["brand_name"].casefold(), row["product_name"].casefold()))

    workbook = Workbook()
    workbook.remove(workbook.active)
    append_summary_sheet(
        workbook,
        [
            ("old_reviewed_row_count", str(len(old_reviewed_rows))),
            ("new_manual_queue_row_count", str(len(new_queue_rows))),
            ("carried_forward_count", str(len(carried_rows))),
            ("needs_manual_review_count", str(len(needs_review_rows))),
        ],
    )
    append_sheet(workbook, "Manual_Queue_Reviewed", carried_rows)
    append_sheet(workbook, "Manual_Queue_Needs_Review", needs_review_rows)
    append_sheet(
        workbook,
        "Hybrid_Reviewed",
        [row for row in carried_rows if normalize_text(row.get("recommended_review_action")) == "review_hybrid_ingredient_signal_sku"],
    )
    append_sheet(
        workbook,
        "Signal_Led_Reviewed",
        [row for row in carried_rows if normalize_text(row.get("recommended_review_action")) == "review_signal_led_sku"],
    )
    append_sheet(
        workbook,
        "Hybrid_Needs_Review",
        [row for row in needs_review_rows if normalize_text(row.get("recommended_review_action")) == "review_hybrid_ingredient_signal_sku"],
    )
    append_sheet(
        workbook,
        "Signal_Led_Needs_Review",
        [row for row in needs_review_rows if normalize_text(row.get("recommended_review_action")) == "review_signal_led_sku"],
    )

    out_xlsx = Path(args.out_xlsx).expanduser().resolve()
    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_xlsx)

    summary = {
        "old_reviewed_xlsx": str(Path(args.old_reviewed_xlsx).expanduser().resolve()),
        "new_manual_queue_csv": str(Path(args.new_manual_queue_csv).expanduser().resolve()),
        "old_reviewed_row_count": len(old_reviewed_rows),
        "new_manual_queue_row_count": len(new_queue_rows),
        "carried_forward_count": len(carried_rows),
        "needs_manual_review_count": len(needs_review_rows),
        "carried_action_counts": dict(carry_action_counts),
        "needs_review_reason_counts": dict(review_reason_counts),
        "out_xlsx": str(out_xlsx),
    }
    out_summary_json = Path(args.out_summary_json).expanduser().resolve()
    out_summary_json.parent.mkdir(parents=True, exist_ok=True)
    out_summary_json.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
