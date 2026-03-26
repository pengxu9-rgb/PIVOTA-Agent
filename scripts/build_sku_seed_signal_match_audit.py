#!/usr/bin/env python3
"""Overlay unmatched SKU ingredient tokens onto the approved signal dictionary.

This script is intentionally read-only:
- it never writes to a database
- it only audits unmatched ingredient tokens against the reviewed signal dictionary
- it mirrors the runtime signal matching order conservatively
"""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to read signal dictionary XLSX files. Install it in the local Python environment first."
    ) from exc


MATCH_REQUIRED_COLUMNS = [
    "candidate_match_key",
    "sku_row_key",
    "source_row_number",
    "brand_name",
    "product_name",
    "official_product_url",
    "category",
    "ingredient_granularity",
    "raw_token",
    "token_index",
    "token_normalized",
    "match_status",
]

SIGNAL_REQUIRED_COLUMNS = [
    "signal_bucket",
    "signal_key",
    "display_signal_name",
    "raw_token_variants",
    "normalized_token_variants",
    "source_packets",
    "source_decisions",
    "confidence_levels",
]

OUTPUT_FIELDS = [
    "candidate_match_key",
    "sku_row_key",
    "source_row_number",
    "brand_name",
    "product_name",
    "official_product_url",
    "category",
    "ingredient_granularity",
    "raw_token",
    "token_index",
    "token_normalized",
    "ingredient_match_status",
    "signal_match_status",
    "signal_match_score",
    "signal_match_method",
    "signal_bucket",
    "signal_key",
    "display_signal_name",
    "signal_confidence_levels",
    "signal_source_packets",
    "signal_source_decisions",
    "signal_raw_token_variants",
    "signal_normalized_token_variants",
    "ambiguity_signal_keys",
]

TM_RE = re.compile(r"[™®©]")


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def split_semicolon(value: Any) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    return [part.strip() for part in raw.split(";") if part.strip()]


def dedupe_keep_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        item = normalize_text(value)
        if not item:
            continue
        marker = item.lower()
        if marker in seen:
            continue
        seen.add(marker)
        ordered.append(item)
    return ordered


def normalize_signal_key(value: Any) -> str:
    raw = TM_RE.sub(" ", str(value or ""))
    raw = unicodedata.normalize("NFKC", raw).lower()
    return "".join(ch for ch in raw if ch.isalnum())[:240]


def normalize_signal_text(value: Any) -> str:
    raw = TM_RE.sub(" ", str(value or "")).strip()
    if not raw:
        return ""
    return " ".join(unicodedata.normalize("NFKC", raw).split())


def load_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader)


def load_signal_rows(path: Path, sheet_name: str) -> list[dict[str, str]]:
    if path.suffix.lower() == ".csv":
        return load_csv_rows(path)

    workbook = load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"Signal dictionary workbook missing required sheet '{sheet_name}': {path}")
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    header = [normalize_text(cell) for cell in rows[0]]
    out: list[dict[str, str]] = []
    for row in rows[1:]:
        if not row or not any(normalize_text(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        out.append(
            {
                header[index]: normalize_text(padded[index] if index < len(padded) else "")
                for index in range(len(header))
            }
        )
    return out


def assert_required_columns(rows: list[dict[str, str]], required: list[str], label: str) -> None:
    available = list(rows[0].keys()) if rows else []
    missing = [field for field in required if field not in available]
    if missing:
        raise SystemExit(f"{label} missing required columns: {', '.join(missing)}")


def score_signal_row(row: dict[str, str], normalized_key: str, normalized_text: str) -> tuple[int, str]:
    signal_key = normalize_signal_key(row.get("signal_key"))
    normalized_variants = {normalize_text(value) for value in split_semicolon(row.get("normalized_token_variants"))}
    display_name = normalize_text(row.get("display_signal_name"))
    raw_variants = split_semicolon(row.get("raw_token_variants"))

    if normalized_key and signal_key == normalized_key:
        return 100, "signal_key"
    if normalized_key and normalized_key in normalized_variants:
        return 90, "normalized_token_variants"
    if normalized_text and display_name.lower() == normalized_text.lower():
        return 80, "display_signal_name"
    if normalized_text and any(value.lower() == normalized_text.lower() for value in raw_variants):
        return 75, "raw_token_variants"
    return 0, ""


def choose_best_signal(rows: list[dict[str, str]], normalized_key: str, normalized_text: str) -> tuple[dict[str, str] | None, bool, int, str]:
    candidates: list[tuple[int, str, dict[str, str]]] = []
    for row in rows:
        score, method = score_signal_row(row, normalized_key, normalized_text)
        if score <= 0:
            continue
        candidates.append((score, method, row))

    if not candidates:
        return None, False, 0, ""

    candidates.sort(key=lambda item: (-item[0], item[2].get("signal_bucket", ""), item[2].get("signal_key", "")))
    top_score = candidates[0][0]
    top_rows = [item for item in candidates if item[0] == top_score]
    top_keys = {item[2].get("signal_key", "") for item in top_rows}
    is_ambiguous = len(top_keys) > 1
    return candidates[0][2], is_ambiguous, candidates[0][0], candidates[0][1]


def write_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def write_filtered_csv(path: Path, rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_FIELDS)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Audit unmatched SKU ingredient tokens against the approved signal dictionary.")
    parser.add_argument("--match-csv", required=True, help="SKU x ingredient match candidate CSV")
    parser.add_argument("--signal-dictionary", required=True, help="Signal dictionary CSV or XLSX path")
    parser.add_argument("--signal-sheet", default="Signal_Dictionary", help="Signal dictionary sheet name when XLSX is used")
    parser.add_argument("--out-csv", required=True, help="Where to write row-level signal audit CSV")
    parser.add_argument("--out-summary-json", required=True, help="Where to write summary JSON")
    parser.add_argument("--out-hit-csv", help="Optional CSV for signal-matched rows only")
    parser.add_argument("--out-remainder-csv", help="Optional CSV for rows still unresolved after signal audit")
    args = parser.parse_args()

    match_rows = load_csv_rows(Path(args.match_csv).expanduser().resolve())
    signal_rows = load_signal_rows(Path(args.signal_dictionary).expanduser().resolve(), args.signal_sheet)
    assert_required_columns(match_rows, MATCH_REQUIRED_COLUMNS, "Match CSV")
    assert_required_columns(signal_rows, SIGNAL_REQUIRED_COLUMNS, "Signal dictionary")

    unmatched_rows = [row for row in match_rows if normalize_text(row.get("match_status")) == "unmatched"]
    audit_rows: list[dict[str, str]] = []
    hit_rows: list[dict[str, str]] = []
    remainder_rows: list[dict[str, str]] = []
    signal_bucket_counts: Counter[str] = Counter()
    signal_key_counts: Counter[str] = Counter()
    status_counts: Counter[str] = Counter()

    for row in unmatched_rows:
        normalized_key = normalize_signal_key(row.get("raw_token"))
        normalized_text = normalize_signal_text(row.get("raw_token"))
        best, is_ambiguous, score, method = choose_best_signal(signal_rows, normalized_key, normalized_text)

        if best and not is_ambiguous:
            signal_match_status = "matched"
            signal_bucket_counts[normalize_text(best.get("signal_bucket"))] += 1
            signal_key_counts[normalize_text(best.get("signal_key"))] += 1
        elif is_ambiguous:
            signal_match_status = "ambiguous"
        else:
            signal_match_status = "unmatched"

        status_counts[signal_match_status] += 1

        output_row = {
            "candidate_match_key": normalize_text(row.get("candidate_match_key")),
            "sku_row_key": normalize_text(row.get("sku_row_key")),
            "source_row_number": normalize_text(row.get("source_row_number")),
            "brand_name": normalize_text(row.get("brand_name")),
            "product_name": normalize_text(row.get("product_name")),
            "official_product_url": normalize_text(row.get("official_product_url")),
            "category": normalize_text(row.get("category")),
            "ingredient_granularity": normalize_text(row.get("ingredient_granularity")),
            "raw_token": normalize_text(row.get("raw_token")),
            "token_index": normalize_text(row.get("token_index")),
            "token_normalized": normalize_text(row.get("token_normalized")),
            "ingredient_match_status": "unmatched",
            "signal_match_status": signal_match_status,
            "signal_match_score": str(score or ""),
            "signal_match_method": method,
            "signal_bucket": normalize_text(best.get("signal_bucket")) if best and not is_ambiguous else "",
            "signal_key": normalize_text(best.get("signal_key")) if best and not is_ambiguous else "",
            "display_signal_name": normalize_text(best.get("display_signal_name")) if best and not is_ambiguous else "",
            "signal_confidence_levels": normalize_text(best.get("confidence_levels")) if best and not is_ambiguous else "",
            "signal_source_packets": normalize_text(best.get("source_packets")) if best and not is_ambiguous else "",
            "signal_source_decisions": normalize_text(best.get("source_decisions")) if best and not is_ambiguous else "",
            "signal_raw_token_variants": normalize_text(best.get("raw_token_variants")) if best and not is_ambiguous else "",
            "signal_normalized_token_variants": normalize_text(best.get("normalized_token_variants")) if best and not is_ambiguous else "",
            "ambiguity_signal_keys": ";".join(
                sorted(
                    {
                        normalize_text(candidate.get("signal_key"))
                        for candidate in signal_rows
                        if score_signal_row(candidate, normalized_key, normalized_text)[0] == score and score > 0
                    }
                )
            )
            if is_ambiguous
            else "",
        }
        audit_rows.append(output_row)
        if signal_match_status == "matched":
            hit_rows.append(output_row)
        else:
            remainder_rows.append(output_row)

    out_csv = Path(args.out_csv).expanduser().resolve()
    write_csv(out_csv, audit_rows)

    if args.out_hit_csv:
        write_filtered_csv(Path(args.out_hit_csv).expanduser().resolve(), hit_rows)
    if args.out_remainder_csv:
        write_filtered_csv(Path(args.out_remainder_csv).expanduser().resolve(), remainder_rows)

    summary = {
        "match_csv": str(Path(args.match_csv).expanduser().resolve()),
        "signal_dictionary": str(Path(args.signal_dictionary).expanduser().resolve()),
        "signal_sheet": args.signal_sheet,
        "signal_dictionary_row_count": len(signal_rows),
        "unmatched_input_row_count": len(unmatched_rows),
        "signal_match_status_counts": dict(status_counts),
        "signal_bucket_hit_counts": dict(signal_bucket_counts),
        "top_signal_keys": dict(signal_key_counts.most_common(25)),
        "out_csv": str(out_csv),
        "out_hit_csv": str(Path(args.out_hit_csv).expanduser().resolve()) if args.out_hit_csv else "",
        "out_remainder_csv": str(Path(args.out_remainder_csv).expanduser().resolve()) if args.out_remainder_csv else "",
    }

    out_summary_json = Path(args.out_summary_json).expanduser().resolve()
    out_summary_json.parent.mkdir(parents=True, exist_ok=True)
    out_summary_json.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
