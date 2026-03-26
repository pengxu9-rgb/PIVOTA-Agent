#!/usr/bin/env python3
"""Build an enriched workbench for remaining ingredient alias manual review rows."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build the ingredient alias manual review workbench. Install it in the local Python environment first."
    ) from exc


PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]


HIGH_CONFIDENCE_ALIASES: dict[str, tuple[str, str, str]] = {
    "PCA": ("Pyroglutamic Acid; Pyrrolidone Carboxylic Acid", "exact_label_alias", "expanded abbreviation"),
    "Petrolatum": ("Petroleum Jelly", "common_alias", "consumer common name"),
    "Cyclopentasiloxane": ("D5", "legacy_alias", "industry shorthand"),
    "Cyclohexasiloxane": ("D6", "legacy_alias", "industry shorthand"),
    "PABA": ("Para-Aminobenzoic Acid", "exact_label_alias", "expanded abbreviation"),
    "BHA": ("Butylated Hydroxyanisole", "exact_label_alias", "expanded abbreviation"),
    "BHT": ("Butylated Hydroxytoluene", "exact_label_alias", "expanded abbreviation"),
    "Retinal": ("Retinaldehyde", "common_alias", "common ingredient shorthand"),
    "Sulfur": ("Sulphur", "exact_label_alias", "cross-market spelling"),
}

MEDIUM_CONFIDENCE_ALIASES: dict[str, tuple[str, str, str]] = {
    "Betaine": ("Trimethylglycine", "legacy_alias", "scientific synonym"),
    "Bisabolol": ("Alpha-Bisabolol", "legacy_alias", "ingredient shorthand commonly seen on PDPs"),
    "Propanediol": ("1,3-Propanediol", "exact_label_alias", "expanded chemical form"),
    "Kaolin": ("China Clay", "common_alias", "common material name"),
}


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def parse_multi(value: Any) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split(";") if part and part.strip()]


def resolve_sheet_name(workbook_path: Path, requested_sheet: str | None = None) -> str:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    if requested_sheet:
        if requested_sheet not in sheet_names:
            raise SystemExit(f"Workbook sheet not found: {requested_sheet}")
        return requested_sheet
    for candidate in PREFERRED_INGREDIENT_SHEETS:
        if candidate in sheet_names:
            return candidate
    raise SystemExit(
        "Workbook is missing a supported ingredient sheet. Expected one of: "
        + ", ".join(PREFERRED_INGREDIENT_SHEETS)
    )


def read_records(workbook_path: Path, sheet_name: str) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    by_id: dict[str, dict[str, Any]] = {}
    for row in rows[1:]:
        if not any(cell is not None and normalize_text(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        record = dict(zip(header, padded))
        record_id = normalize_text(record.get("record_id"))
        if record_id:
            by_id[record_id] = record
    return by_id


def read_manual_rows(path: Path) -> list[dict[str, str]]:
    with path.open(encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def dedupe_keep_order(values: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for value in values:
        text = normalize_text(value)
        if not text:
            continue
        key = text.casefold()
        if key in seen:
            continue
        seen.add(key)
        out.append(text)
    return out


def build_resolution(record: dict[str, Any]) -> tuple[str, str, str, str]:
    canonical = normalize_text(record.get("canonical_inci_name"))
    if canonical in HIGH_CONFIDENCE_ALIASES:
        aliases, quality, rationale = HIGH_CONFIDENCE_ALIASES[canonical]
        return "candidate_alias_high_confidence", aliases, quality, rationale
    if canonical in MEDIUM_CONFIDENCE_ALIASES:
        aliases, quality, rationale = MEDIUM_CONFIDENCE_ALIASES[canonical]
        return "candidate_alias_manual_confirmation", aliases, quality, rationale

    parser_variants = dedupe_keep_order(parse_multi(record.get("parser_variants")))
    display_name = normalize_text(record.get("canonical_display_name"))
    us_name = normalize_text(record.get("us_label_name"))
    eu_name = normalize_text(record.get("eu_label_name"))
    if parser_variants in ([canonical], [canonical, canonical.lower()]) and display_name == canonical and us_name == canonical and eu_name == canonical:
        return "likely_no_safe_common_alias", "", "", "workbook context shows only exact self-name variants"

    return "needs_manual_research", "", "", "workbook context did not surface a safe deterministic alias"


def build_rows(workbook_records: dict[str, dict[str, Any]], manual_rows: list[dict[str, str]]) -> list[dict[str, Any]]:
    out: list[dict[str, Any]] = []
    for manual_row in manual_rows:
        record_id = normalize_text(manual_row.get("record_id"))
        record = workbook_records.get(record_id, {})
        resolution, suggested_aliases, suggested_quality, rationale = build_resolution(record)
        out.append(
            {
                "record_id": record_id,
                "canonical_inci_name": normalize_text(record.get("canonical_inci_name")),
                "canonical_display_name": normalize_text(record.get("canonical_display_name")),
                "ingredient_family": normalize_text(record.get("ingredient_family")),
                "us_label_name": normalize_text(record.get("us_label_name")),
                "eu_label_name": normalize_text(record.get("eu_label_name")),
                "us_label_variants": normalize_text(record.get("us_label_variants")),
                "eu_label_variants": normalize_text(record.get("eu_label_variants")),
                "parser_variants": normalize_text(record.get("parser_variants")),
                "suggested_resolution": resolution,
                "suggested_aliases_common": suggested_aliases,
                "suggested_alias_quality": suggested_quality,
                "resolution_rationale": rationale,
            }
        )
    return out


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "record_id",
        "canonical_inci_name",
        "canonical_display_name",
        "ingredient_family",
        "us_label_name",
        "eu_label_name",
        "us_label_variants",
        "eu_label_variants",
        "parser_variants",
        "suggested_resolution",
        "suggested_aliases_common",
        "suggested_alias_quality",
        "resolution_rationale",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a manual-review workbench for remaining ingredient alias rows.")
    parser.add_argument("--ingredient-xlsx", required=True)
    parser.add_argument("--sheet-name", help="Optional workbook sheet name")
    parser.add_argument("--manual-csv", required=True)
    parser.add_argument("--out-json", help="Optional path to write JSON summary")
    parser.add_argument("--out-csv", help="Optional path to write CSV workbench")
    args = parser.parse_args()

    workbook_path = Path(args.ingredient_xlsx).expanduser().resolve()
    manual_path = Path(args.manual_csv).expanduser().resolve()
    sheet_name = resolve_sheet_name(workbook_path, args.sheet_name)
    workbook_records = read_records(workbook_path, sheet_name)
    manual_rows = read_manual_rows(manual_path)
    rows = build_rows(workbook_records, manual_rows)

    summary = {
        "workbook": str(workbook_path),
        "sheet_name": sheet_name,
        "manual_csv": str(manual_path),
        "row_count": len(rows),
        "resolution_counts": {
            "candidate_alias_high_confidence": sum(r["suggested_resolution"] == "candidate_alias_high_confidence" for r in rows),
            "candidate_alias_manual_confirmation": sum(r["suggested_resolution"] == "candidate_alias_manual_confirmation" for r in rows),
            "likely_no_safe_common_alias": sum(r["suggested_resolution"] == "likely_no_safe_common_alias" for r in rows),
            "needs_manual_research": sum(r["suggested_resolution"] == "needs_manual_research" for r in rows),
        },
        "rows": rows,
    }

    rendered = json.dumps(summary, ensure_ascii=True, indent=2)
    if args.out_json:
        out_json = Path(args.out_json).expanduser().resolve()
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(rendered + "\n", encoding="utf-8")
    if args.out_csv:
        write_csv(Path(args.out_csv).expanduser().resolve(), rows)
    print(rendered)


if __name__ == "__main__":
    main()
