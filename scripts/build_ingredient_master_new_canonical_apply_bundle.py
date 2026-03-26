#!/usr/bin/env python3
"""Build apply-ready and hold bundles from researched new-canonical decisions."""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build the ingredient-master new-canonical apply bundle. Install it in the local Python environment first."
    ) from exc


TARGET_SHEET = "Ingredient_Reference_Merged_v2"
PATCH_ID_RE = re.compile(r"^ing_patch_v13_(\d+)$", re.IGNORECASE)

APPLY_FIELDS = [
    "record_id",
    "canonical_inci_name",
    "canonical_display_name",
    "ingredient_family",
    "us_label_name",
    "eu_label_name",
    "us_label_variants",
    "eu_label_variants",
    "cross_market_notes",
    "normalized_key",
    "aliases_common",
    "parser_variants",
    "deprecated_aliases",
    "alias_quality",
    "notes_for_parser",
    "primary_bucket",
    "all_buckets",
    "function_tags",
    "benefit_tags",
    "risk_flags",
    "is_humectant",
    "is_barrier_support",
    "is_retinoid",
    "is_exfoliant",
    "is_uv_filter",
    "is_preservative",
    "is_surfactant",
    "is_fragrance_or_eo",
    "regulatory_bucket",
    "source_urls",
    "source_authorities",
    "source_types",
    "review_status",
    "confidence",
    "last_reviewed_at",
    "review_notes",
    "notes",
    "kb_version",
    "queue_priority_score",
    "queue_example_brands",
    "queue_example_products",
    "queue_example_urls",
    "source_packet_resolution",
    "source_packet_confidence",
    "source_packet_raw_token",
    "semantic_match_key",
]

HOLD_FIELDS = [
    "signal_key",
    "display_signal_name",
    "decision_status",
    "recommended_master_action",
    "decision_basis",
    "source_urls",
    "example_brands",
    "example_products",
    "reviewer_notes",
]


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: str) -> str:
    return "".join(ch.lower() for ch in normalize_text(value) if ch.isalnum())


def split_multi(value: str) -> list[str]:
    return [part.strip() for part in normalize_text(value).split(";") if part.strip()]


def join_multi(values: list[str]) -> str:
    return "; ".join([normalize_text(value) for value in values if normalize_text(value)])


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return [{key: normalize_text(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def append_sheet(workbook: Workbook, title: str, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    sheet = workbook.create_sheet(title)
    sheet.append(fieldnames)
    for row in rows:
        sheet.append([row.get(field, "") for field in fieldnames])


def append_summary_sheet(workbook: Workbook, summary: dict[str, Any]) -> None:
    sheet = workbook.create_sheet("Summary")
    sheet.append(["metric", "value"])
    for key, value in summary.items():
        if isinstance(value, (dict, list)):
            sheet.append([key, json.dumps(value, ensure_ascii=True, sort_keys=True)])
        else:
            sheet.append([key, value])


def next_patch_numbers(workbook_path: Path, count: int) -> list[int]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if TARGET_SHEET not in workbook.sheetnames:
        raise SystemExit(f"Workbook missing target sheet: {TARGET_SHEET}")
    sheet = workbook[TARGET_SHEET]
    header = [normalize_text(cell.value) for cell in sheet[1]]
    try:
        record_id_index = header.index("record_id")
    except ValueError as exc:
        raise SystemExit("Workbook target sheet missing record_id column.") from exc

    max_seen = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        record_id = normalize_text(row[record_id_index])
        match = PATCH_ID_RE.match(record_id)
        if match:
            max_seen = max(max_seen, int(match.group(1)))
    return list(range(max_seen + 1, max_seen + count + 1))


def main() -> None:
    parser = argparse.ArgumentParser(description="Build apply-ready and hold bundles from researched new-canonical decisions.")
    parser.add_argument("--decision-csv", required=True, help="Researched new-canonical decision packet CSV")
    parser.add_argument("--terms-csv", required=True, help="Original new-canonical term CSV to recover followup priority")
    parser.add_argument("--ingredient-xlsx", required=True, help="Current ingredient workbook used for record-id allocation")
    parser.add_argument("--out-apply-csv", required=True, help="Where to write apply-ready new-canonical patch CSV")
    parser.add_argument("--out-hold-csv", required=True, help="Where to write hold review CSV")
    parser.add_argument("--out-summary-json", required=True, help="Where to write summary JSON")
    parser.add_argument("--out-xlsx", help="Optional workbook output")
    args = parser.parse_args()

    decision_rows = read_csv_rows(Path(args.decision_csv).expanduser().resolve())
    term_rows = read_csv_rows(Path(args.terms_csv).expanduser().resolve())
    term_by_key = {row["signal_key"]: row for row in term_rows}

    apply_candidates = [
        row for row in decision_rows if row["decision_status"] == "approved_rewrite_to_underlying_inci_candidate"
    ]
    hold_rows = [
        {
            key: row[key]
            for key in HOLD_FIELDS
        }
        for row in decision_rows
        if row["decision_status"] == "approved_hold_named_active_pending_exact_label_confirmation"
    ]

    patch_numbers = next_patch_numbers(Path(args.ingredient_xlsx).expanduser().resolve(), len(apply_candidates))

    apply_rows: list[dict[str, str]] = []
    for patch_number, row in zip(patch_numbers, apply_candidates):
        signal_key = row["signal_key"]
        term = term_by_key.get(signal_key, {})
        candidate_inci = normalize_text(row["recommended_candidate_inci_name"])
        candidate_display = normalize_text(row["recommended_candidate_display_name"]) or candidate_inci
        source_urls = split_multi(row.get("source_urls", ""))
        priority = normalize_text(term.get("followup_priority"))
        queue_priority_score = {"high": "22", "medium": "18", "low": "14"}.get(priority, "18")
        display_signal = normalize_text(row.get("display_signal_name"))

        apply_rows.append(
            {
                "record_id": f"ing_patch_v13_{patch_number}",
                "canonical_inci_name": candidate_inci,
                "canonical_display_name": candidate_display,
                "ingredient_family": "plant_extract",
                "us_label_name": candidate_inci,
                "eu_label_name": candidate_inci,
                "us_label_variants": candidate_inci,
                "eu_label_variants": candidate_inci,
                "cross_market_notes": f"Promoted from reviewed signal-led followup term '{display_signal}'.",
                "normalized_key": normalize_key(candidate_inci),
                "aliases_common": "",
                "parser_variants": join_multi([candidate_inci, display_signal]),
                "deprecated_aliases": "",
                "alias_quality": "",
                "notes_for_parser": f"Reviewed signal-led followup rewrite from '{display_signal}'; keep signal term in parser variants until canonical coverage is verified.",
                "primary_bucket": "",
                "all_buckets": "",
                "function_tags": "",
                "benefit_tags": "",
                "risk_flags": "",
                "is_humectant": "no",
                "is_barrier_support": "no",
                "is_retinoid": "no",
                "is_exfoliant": "no",
                "is_uv_filter": "no",
                "is_preservative": "no",
                "is_surfactant": "no",
                "is_fragrance_or_eo": "no",
                "regulatory_bucket": "patch_candidate_review",
                "source_urls": join_multi(source_urls),
                "source_authorities": "brand_official_pdp",
                "source_types": "official_brand_site",
                "review_status": "draft",
                "confidence": "low",
                "last_reviewed_at": "",
                "review_notes": normalize_text(row.get("decision_basis")),
                "notes": f"rewritten_from_signal_key={signal_key}; rewritten_from_signal_display={display_signal}",
                "kb_version": "v13_merged_v2_24_signal_followup_patch",
                "queue_priority_score": queue_priority_score,
                "queue_example_brands": normalize_text(row.get("example_brands")),
                "queue_example_products": normalize_text(row.get("example_products")),
                "queue_example_urls": join_multi(source_urls),
                "source_packet_resolution": "reviewed_signal_followup_rewrite_to_inci_candidate",
                "source_packet_confidence": "reviewed_manual_signal_followup",
                "source_packet_raw_token": display_signal,
                "semantic_match_key": normalize_key(candidate_inci),
            }
        )

    apply_rows.sort(key=lambda row: normalize_text(row["canonical_inci_name"]).casefold())
    hold_rows.sort(key=lambda row: normalize_text(row["display_signal_name"]).casefold())

    out_apply = Path(args.out_apply_csv).expanduser().resolve()
    out_hold = Path(args.out_hold_csv).expanduser().resolve()
    write_csv(out_apply, APPLY_FIELDS, apply_rows)
    write_csv(out_hold, HOLD_FIELDS, hold_rows)

    summary = {
        "decision_row_count": len(decision_rows),
        "apply_ready_count": len(apply_rows),
        "hold_count": len(hold_rows),
        "apply_priority_counts": dict(Counter(row["queue_priority_score"] for row in apply_rows)),
        "hold_signal_names": [row["display_signal_name"] for row in hold_rows],
        "out_apply_csv": str(out_apply),
        "out_hold_csv": str(out_hold),
    }

    if args.out_xlsx:
        workbook = Workbook()
        default_sheet = workbook.active
        workbook.remove(default_sheet)
        append_summary_sheet(workbook, summary)
        append_sheet(workbook, "Apply_Ready", APPLY_FIELDS, apply_rows)
        append_sheet(workbook, "Hold", HOLD_FIELDS, hold_rows)
        out_xlsx = Path(args.out_xlsx).expanduser().resolve()
        out_xlsx.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(out_xlsx)
        summary["out_xlsx"] = str(out_xlsx)

    out_summary = Path(args.out_summary_json).expanduser().resolve()
    out_summary.parent.mkdir(parents=True, exist_ok=True)
    out_summary.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
