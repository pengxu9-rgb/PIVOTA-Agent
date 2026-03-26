#!/usr/bin/env python3
"""Propose a narrow curated ingredient-family overlay for `other` rows."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to propose ingredient family overlays. Install it in the local Python environment first."
    ) from exc


PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]


CURATED_FAMILY_MAP: dict[str, tuple[str, str]] = {
    "Ammonium Lactate": ("acid_exfoliant", "curated_exfoliant_acid_family"),
    "Avena Sativa Kernel Flour": ("plant_extract", "curated_botanical_family"),
    "Asiatic Acid": ("plant_extract", "curated_botanical_family"),
    "Asiaticoside": ("plant_extract", "curated_botanical_family"),
    "Bakuchiol": ("plant_extract", "curated_botanical_family"),
    "Bisabolol": ("plant_extract", "curated_botanical_family"),
    "Gluconolactone": ("acid_exfoliant", "curated_exfoliant_acid_family"),
    "Hamamelis Virginiana Water": ("plant_extract", "curated_botanical_family"),
    "Madecassic Acid": ("plant_extract", "curated_botanical_family"),
    "Madecassoside": ("plant_extract", "curated_botanical_family"),
    "Sodium Ascorbate": ("vitamin", "curated_vitamin_family"),
}


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def load_queue_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def resolve_sheet_name(path: Path, requested_sheet: str | None = None) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    if requested_sheet:
        if requested_sheet not in sheet_names:
            raise SystemExit(f"Workbook sheet not found: {requested_sheet}")
        return requested_sheet
    for candidate in PREFERRED_INGREDIENT_SHEETS:
        if candidate in sheet_names:
            return candidate
    raise SystemExit(
        "Workbook is missing a supported ingredient sheet. Expected one of: "
        + ", ".join(PREFERRED_INGREDIENT_SHEETS)
    )


def read_workbook_records(path: Path, sheet_name: str) -> dict[str, dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    records: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        if not any(row):
            continue
        rec = {
            key: normalize_text(row[index] if index < len(row) else "")
            for index, key in enumerate(header)
        }
        record_id = rec.get("record_id")
        if record_id:
            records[record_id] = rec
    return records


def propose_family(record: dict[str, str]) -> tuple[str, str] | None:
    canonical = record.get("canonical_inci_name", "")
    if canonical in CURATED_FAMILY_MAP:
        return CURATED_FAMILY_MAP[canonical]

    benefit_tags = record.get("benefit_tags", "").lower()
    function_tags = record.get("function_tags", "").lower()
    primary_bucket = record.get("primary_bucket", "").lower()
    if "emollient" in benefit_tags and "skin conditioning" in function_tags and primary_bucket == "repair":
        return ("emollient", "heuristic_emollient_benefit")

    return None


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Propose a narrow curated ingredient-family overlay for `other` rows.")
    parser.add_argument("--ingredient-xlsx", required=True, help="Current ingredient workbook")
    parser.add_argument("--sheet-name", help="Optional workbook sheet name")
    parser.add_argument("--family-review-csv", required=True, help="ingredient_family_review CSV")
    parser.add_argument("--out-apply-csv", required=True, help="Where to write the apply-ready family patch CSV")
    parser.add_argument("--out-remainder-csv", help="Optional path for unmatched remainder rows")
    parser.add_argument("--out-json", help="Optional path for a JSON summary")
    args = parser.parse_args()

    workbook_path = Path(args.ingredient_xlsx).expanduser().resolve()
    review_csv_path = Path(args.family_review_csv).expanduser().resolve()
    queue_rows = load_queue_rows(review_csv_path)
    sheet_name = resolve_sheet_name(workbook_path, args.sheet_name)
    workbook_records = read_workbook_records(workbook_path, sheet_name)

    apply_rows: list[dict[str, str]] = []
    remainder_rows: list[dict[str, str]] = []

    for row in queue_rows:
        record_id = normalize_text(row.get("record_id"))
        record = workbook_records.get(record_id)
        if not record:
            remainder_rows.append(row)
            continue
        proposed = propose_family(record)
        if not proposed:
            remainder_rows.append(row)
            continue
        patch_family, quality_reason = proposed
        apply_rows.append(
            {
                "record_id": record_id,
                "canonical_inci_name": record.get("canonical_inci_name", ""),
                "existing_ingredient_family": record.get("ingredient_family", ""),
                "patch_ingredient_family": patch_family,
                "proposal_sources": "curated_family_overlay",
                "quality_reason": quality_reason,
            }
        )

    apply_fieldnames = [
        "record_id",
        "canonical_inci_name",
        "existing_ingredient_family",
        "patch_ingredient_family",
        "proposal_sources",
        "quality_reason",
    ]
    remainder_fieldnames = list(queue_rows[0].keys()) if queue_rows else []

    out_apply_csv = Path(args.out_apply_csv).expanduser().resolve()
    write_csv(out_apply_csv, apply_rows, apply_fieldnames)

    remainder_path = None
    if args.out_remainder_csv:
        remainder_path = Path(args.out_remainder_csv).expanduser().resolve()
        write_csv(remainder_path, remainder_rows, remainder_fieldnames)

    payload = {
        "source_workbook": str(workbook_path),
        "source_sheet": sheet_name,
        "source_family_review_csv": str(review_csv_path),
        "curated_family_map_size": len(CURATED_FAMILY_MAP),
        "apply_ready_count": len(apply_rows),
        "remainder_count": len(remainder_rows),
        "matched_rows": apply_rows,
        "out_apply_csv": str(out_apply_csv),
        "out_remainder_csv": str(remainder_path) if remainder_path else None,
    }

    if args.out_json:
        out_json = Path(args.out_json).expanduser().resolve()
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(json.dumps(payload, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")

    print(json.dumps(payload, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
