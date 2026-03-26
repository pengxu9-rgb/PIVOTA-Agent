#!/usr/bin/env python3
"""Apply alias writeback patch CSV to an ingredient workbook copy.

This helper is conservative:
- it writes to a new workbook path only
- it matches rows by `record_id`
- it only updates alias fields when the current workbook value still matches the
  expected existing value from the patch row
- conflicting rows are skipped and reported
"""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to apply workbook alias patches. Install it in the local Python environment first."
    ) from exc


PREFERRED_TARGET_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def load_patch_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def build_header_index(sheet) -> dict[str, int]:
    header = [normalize_text(cell.value) for cell in sheet[1]]
    return {name: index + 1 for index, name in enumerate(header)}


def resolve_target_sheet_name(workbook, requested_sheet: str | None) -> str:
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise SystemExit(f"Workbook missing requested target sheet '{requested_sheet}'.")
        return requested_sheet
    for sheet_name in PREFERRED_TARGET_SHEETS:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    raise SystemExit(
        f"Workbook did not contain any supported target sheet names ({', '.join(PREFERRED_TARGET_SHEETS)})."
    )


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply ingredient alias writeback patch CSV to a workbook copy.")
    parser.add_argument("--ingredient-xlsx", required=True, help="Source ingredient workbook")
    parser.add_argument("--patch-csv", required=True, help="Apply-ready patch CSV produced by export_ingredient_alias_writeback_patch.py")
    parser.add_argument("--out-xlsx", required=True, help="Path to write the patched workbook copy")
    parser.add_argument("--target-sheet", help="Optional target sheet name")
    parser.add_argument("--out-report-json", help="Optional path for apply report JSON")
    args = parser.parse_args()

    workbook_path = Path(args.ingredient_xlsx).expanduser().resolve()
    patch_path = Path(args.patch_csv).expanduser().resolve()
    out_xlsx = Path(args.out_xlsx).expanduser().resolve()

    workbook = load_workbook(workbook_path)
    target_sheet_name = resolve_target_sheet_name(workbook, args.target_sheet)
    sheet = workbook[target_sheet_name]
    header_index = build_header_index(sheet)

    required_columns = {
        "record_id",
        "aliases_common",
        "alias_quality",
    }
    missing_columns = [column for column in required_columns if column not in header_index]
    if missing_columns:
        raise SystemExit(f"Workbook Dictionary sheet is missing required columns: {', '.join(missing_columns)}")

    patch_rows = load_patch_rows(patch_path)
    patch_has_parser_variants = any(
        normalize_text(row.get("existing_parser_variants")) or normalize_text(row.get("patch_parser_variants"))
        for row in patch_rows
    )
    if patch_has_parser_variants and "parser_variants" not in header_index:
        raise SystemExit("Workbook target sheet is missing required column: parser_variants")

    row_by_record_id: dict[str, int] = {}
    for row_number in range(2, sheet.max_row + 1):
        record_id = normalize_text(sheet.cell(row=row_number, column=header_index["record_id"]).value)
        if record_id:
            row_by_record_id[record_id] = row_number

    applied: list[str] = []
    skipped_missing_record: list[str] = []
    skipped_conflicts: list[dict[str, str]] = []

    for patch in patch_rows:
        record_id = normalize_text(patch.get("record_id"))
        if not record_id:
            continue
        row_number = row_by_record_id.get(record_id)
        if not row_number:
            skipped_missing_record.append(record_id)
            continue

        alias_cell = sheet.cell(row=row_number, column=header_index["aliases_common"])
        quality_cell = sheet.cell(row=row_number, column=header_index["alias_quality"])
        parser_variants_cell = (
            sheet.cell(row=row_number, column=header_index["parser_variants"])
            if patch_has_parser_variants
            else None
        )

        current_aliases = normalize_text(alias_cell.value)
        current_quality = normalize_text(quality_cell.value)
        current_parser_variants = normalize_text(parser_variants_cell.value) if parser_variants_cell else ""
        expected_aliases = normalize_text(patch.get("existing_aliases_common"))
        expected_parser_variants = normalize_text(patch.get("existing_parser_variants"))
        expected_quality = normalize_text(patch.get("existing_alias_quality"))

        if (
            current_aliases != expected_aliases
            or current_quality != expected_quality
            or (patch_has_parser_variants and current_parser_variants != expected_parser_variants)
        ):
            skipped_conflicts.append(
                {
                    "record_id": record_id,
                    "current_aliases_common": current_aliases,
                    "expected_aliases_common": expected_aliases,
                    "current_parser_variants": current_parser_variants,
                    "expected_parser_variants": expected_parser_variants,
                    "current_alias_quality": current_quality,
                    "expected_alias_quality": expected_quality,
                }
            )
            continue

        alias_cell.value = normalize_text(patch.get("patch_aliases_common"))
        quality_cell.value = normalize_text(patch.get("patch_alias_quality"))
        if parser_variants_cell is not None:
            parser_variants_cell.value = normalize_text(patch.get("patch_parser_variants"))
        applied.append(record_id)

    out_xlsx.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(out_xlsx)

    report = {
        "source_workbook": str(workbook_path),
        "patch_csv": str(patch_path),
        "out_workbook": str(out_xlsx),
        "target_sheet": target_sheet_name,
        "patch_row_count": len(patch_rows),
        "applied_count": len(applied),
        "skipped_missing_record_count": len(skipped_missing_record),
        "skipped_conflict_count": len(skipped_conflicts),
        "skipped_missing_record_ids": skipped_missing_record,
        "skipped_conflicts": skipped_conflicts,
    }

    rendered = json.dumps(report, ensure_ascii=True, indent=2)
    if args.out_report_json:
        report_path = Path(args.out_report_json).expanduser().resolve()
        report_path.parent.mkdir(parents=True, exist_ok=True)
        report_path.write_text(rendered + "\n", encoding="utf-8")
    print(rendered)


if __name__ == "__main__":
    main()
