#!/usr/bin/env python3
"""Build a triage workbook for unmatched ingredient tokens."""

from __future__ import annotations

import argparse
import csv
import json
import re
import unicodedata
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

try:
    from openpyxl import Workbook, load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build ingredient unmatched priority queues. Install it in the local Python environment first."
    ) from exc


MATCH_REQUIRED_COLUMNS = [
    "sku_row_key",
    "brand_name",
    "product_name",
    "official_product_url",
    "category",
    "ingredient_granularity",
    "raw_token",
    "token_normalized",
    "match_status",
]

TERM_FIELDS = [
    "canonical_inci_name",
    "canonical_display_name",
    "us_label_name",
    "eu_label_name",
    "parser_variants",
    "aliases_common",
    "us_label_variants",
    "eu_label_variants",
    "deprecated_aliases",
]

QUEUE_COLUMNS = [
    "priority_score",
    "recommended_bucket",
    "recommended_action",
    "triage_reason",
    "raw_token",
    "normalized_token",
    "unmatched_count",
    "sku_row_count",
    "full_inci_count",
    "key_count",
    "product_only_count",
    "top_categories",
    "example_brands",
    "example_products",
    "example_urls",
    "in_current_master_like",
]

SIGNAL_PATTERNS = [
    re.compile(pattern, re.IGNORECASE)
    for pattern in [
        r"\baha\b",
        r"\bbha\b",
        r"\bpha\b",
        r"\bceramides?\b",
        r"\bpeptides?\b",
        r"\bcollagen\b",
        r"\bantioxidants?\b",
        r"\bcomplex\b",
        r"\bblend\b",
        r"\blipid\b",
        r"\bretinoid\b",
    ]
]

MANUAL_PATTERNS = [
    re.compile(pattern, re.IGNORECASE)
    for pattern in [
        r"active ingredient",
        r"inactive ingredient",
        r"ingredients:",
        r"ingredient:",
    ]
]

BUCKET_META = {
    "candidate_canonical_full_inci": {
        "sheet": "Full_INCI_Priority_Queue",
        "action": "append_new_canonical_candidate",
        "reason": "Appears in full INCI rows and is not represented in current master-like keys.",
        "adjustment": 20,
    },
    "alias_or_normalization_gap": {
        "sheet": "Alias_Normalization_Queue",
        "action": "append_alias_or_parser_variant",
        "reason": "Looks like label/format/common-name variation; prefer alias or parser normalization before new canonical.",
        "adjustment": 8,
    },
    "signal_or_family_term": {
        "sheet": "Signal_Review_Queue",
        "action": "do_not_add_to_canonical__route_to_signal_dict",
        "reason": "Looks like family, marketing, or broad ingredient-group term; avoid canonical row.",
        "adjustment": -40,
    },
    "verify_parser_or_export": {
        "sheet": "Verify_Parser_Queue",
        "action": "verify_current_match_logic_or_export_filter",
        "reason": "Seems already covered by current ingredient master-like keys; verify matching logic/token export.",
        "adjustment": 10,
    },
    "manual_review": {
        "sheet": "Manual_Review_Queue",
        "action": "manual_triage_required",
        "reason": "Needs manual review before deciding between parser fix, alias normalization, or new canonical.",
        "adjustment": 12,
    },
}

PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]

PAREN_RE = re.compile(r"\([^)]*\)")


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def normalize_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    return "".join(ch for ch in raw if ch.isalnum())


def semantic_key(value: Any) -> str:
    raw = unicodedata.normalize("NFKC", normalize_text(value)).lower()
    raw = PAREN_RE.sub(" ", raw)
    raw = " ".join(raw.split())
    return "".join(ch for ch in raw if ch.isalnum())


def split_semicolon_values(value: Any) -> list[str]:
    raw = normalize_text(value)
    if not raw:
        return []
    return [part.strip() for part in raw.split(";") if part and part.strip()]


def read_match_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        rows = list(csv.DictReader(handle))
    missing = [column for column in MATCH_REQUIRED_COLUMNS if column not in (rows[0].keys() if rows else MATCH_REQUIRED_COLUMNS)]
    if rows and missing:
        raise SystemExit(f"Match candidate CSV missing required columns: {', '.join(missing)}")
    return rows


def resolve_ingredient_sheet_name(path: Path, requested_sheet: str | None) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise SystemExit(f"Ingredient workbook missing requested sheet '{requested_sheet}'.")
        return requested_sheet
    for sheet_name in PREFERRED_INGREDIENT_SHEETS:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    raise SystemExit(
        f"Ingredient workbook did not contain any supported sheet names ({', '.join(PREFERRED_INGREDIENT_SHEETS)})."
    )


def build_master_like_keys(path: Path, sheet_name: str) -> tuple[set[str], set[str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return set(), set()

    header = [normalize_text(cell) for cell in rows[0]]
    header_index = {name: index for index, name in enumerate(header)}
    exact_keys: set[str] = set()
    semantic_keys: set[str] = set()

    for row in rows[1:]:
        if not row:
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        record = {
            key: normalize_text(padded[index] if index < len(padded) else "")
            for index, key in enumerate(header)
        }
        for field in TERM_FIELDS:
            if field not in header_index:
                continue
            values = split_semicolon_values(record.get(field)) if field.endswith("variants") or field in {"parser_variants", "aliases_common", "deprecated_aliases"} else [record.get(field, "")]
            for value in values:
                normalized = normalize_key(value)
                if normalized:
                    exact_keys.add(normalized)
                semantic = semantic_key(value)
                if semantic:
                    semantic_keys.add(semantic)

    return exact_keys, semantic_keys


def top_join(counter: Counter[str], limit: int = 5) -> str:
    return "; ".join(f"{value}:{count}" for value, count in counter.most_common(limit) if value)


def ordered_unique(values: list[str], limit: int = 5) -> str:
    seen: set[str] = set()
    ordered: list[str] = []
    for value in values:
        text = normalize_text(value)
        if not text or text in seen:
            continue
        seen.add(text)
        ordered.append(text)
        if len(ordered) >= limit:
            break
    return "; ".join(ordered)


def is_mostly_uppercase(token: str) -> bool:
    letters = [ch for ch in token if ch.isalpha()]
    if len(letters) < 3:
        return False
    uppercase_count = sum(1 for ch in letters if ch.isupper())
    return uppercase_count / len(letters) >= 0.7


def is_signal_term(token: str, full_count: int, key_count: int) -> bool:
    lowered = normalize_text(token)
    if key_count > 0 and full_count == 0:
        return True
    if "%" in lowered:
        return True
    if "™" in lowered or "®" in lowered:
        return True
    return any(pattern.search(lowered) for pattern in SIGNAL_PATTERNS)


def is_manual_term(token: str) -> bool:
    lowered = normalize_text(token)
    if "%" in lowered and ":" in lowered:
        return True
    return any(pattern.search(lowered) for pattern in MANUAL_PATTERNS)


def is_alias_like(token: str) -> bool:
    text = normalize_text(token)
    if not text:
        return False
    if is_mostly_uppercase(text):
        return True
    if any(marker in text for marker in ["/", "(", ")", "'", '"']):
        return True
    return False


def classify_bucket(token: str, full_count: int, key_count: int, in_master_like: bool) -> str:
    if in_master_like:
        return "verify_parser_or_export"
    if is_manual_term(token):
        return "manual_review"
    if is_signal_term(token, full_count, key_count):
        return "signal_or_family_term"
    if is_alias_like(token):
        return "alias_or_normalization_gap"
    if full_count > 0:
        return "candidate_canonical_full_inci"
    return "manual_review"


def compute_priority_score(bucket: str, unmatched_count: int, sku_row_count: int, full_count: int, key_count: int) -> int:
    base = unmatched_count * 8 + sku_row_count * 4 + full_count * 2 + key_count
    return base + int(BUCKET_META[bucket]["adjustment"])


def build_aggregate_rows(
    match_rows: list[dict[str, str]],
    master_like_exact: set[str],
    master_like_semantic: set[str],
) -> list[dict[str, Any]]:
    grouped: dict[str, dict[str, Any]] = {}

    for row in match_rows:
        if normalize_text(row.get("match_status")) != "unmatched":
            continue
        raw_token = normalize_text(row.get("raw_token"))
        if not raw_token:
            continue

        entry = grouped.setdefault(
            raw_token,
            {
                "raw_token": raw_token,
                "normalized_token": normalize_key(row.get("token_normalized") or raw_token),
                "unmatched_count": 0,
                "sku_row_keys": set(),
                "full_inci_count": 0,
                "key_count": 0,
                "product_only_count": 0,
                "categories": Counter(),
                "brands": [],
                "products": [],
                "urls": [],
            },
        )

        entry["unmatched_count"] += 1
        entry["sku_row_keys"].add(normalize_text(row.get("sku_row_key")))

        granularity = normalize_text(row.get("ingredient_granularity"))
        if granularity == "full_inci_official":
            entry["full_inci_count"] += 1
        elif granularity == "key_ingredients_official":
            entry["key_count"] += 1
        elif granularity == "product_page_only":
            entry["product_only_count"] += 1

        category = normalize_text(row.get("category"))
        if category:
            entry["categories"][category] += 1
        entry["brands"].append(normalize_text(row.get("brand_name")))
        entry["products"].append(normalize_text(row.get("product_name")))
        entry["urls"].append(normalize_text(row.get("official_product_url")))

    triage_rows: list[dict[str, Any]] = []
    for raw_token, entry in grouped.items():
        normalized_token = normalize_key(entry["normalized_token"] or raw_token)
        semantic_token = semantic_key(raw_token)
        in_master_like = normalized_token in master_like_exact or semantic_token in master_like_semantic
        bucket = classify_bucket(raw_token, entry["full_inci_count"], entry["key_count"], in_master_like)
        priority_score = compute_priority_score(
            bucket=bucket,
            unmatched_count=int(entry["unmatched_count"]),
            sku_row_count=len(entry["sku_row_keys"]),
            full_count=int(entry["full_inci_count"]),
            key_count=int(entry["key_count"]),
        )

        triage_rows.append(
            {
                "priority_score": priority_score,
                "recommended_bucket": bucket,
                "recommended_action": BUCKET_META[bucket]["action"],
                "triage_reason": BUCKET_META[bucket]["reason"],
                "raw_token": raw_token,
                "normalized_token": normalized_token,
                "unmatched_count": int(entry["unmatched_count"]),
                "sku_row_count": len(entry["sku_row_keys"]),
                "full_inci_count": int(entry["full_inci_count"]),
                "key_count": int(entry["key_count"]),
                "product_only_count": int(entry["product_only_count"]),
                "top_categories": top_join(entry["categories"]),
                "example_brands": ordered_unique(entry["brands"]),
                "example_products": ordered_unique(entry["products"]),
                "example_urls": ordered_unique(entry["urls"]),
                "in_current_master_like": bool(in_master_like),
            }
        )

    triage_rows.sort(
        key=lambda row: (
            -int(row["priority_score"]),
            -int(row["unmatched_count"]),
            -int(row["full_inci_count"]),
            row["raw_token"].lower(),
        )
    )
    return triage_rows


def write_workbook(path: Path, triage_rows: list[dict[str, Any]], summary: dict[str, Any]) -> None:
    workbook = Workbook()
    readme = workbook.active
    readme.title = "README"
    readme.append(["metric", "value"])
    for metric, value in [
        ("current_match_summary_token_count", summary.get("token_count")),
        ("current_match_summary_matched_token_count", summary.get("matched_token_count")),
        ("current_match_summary_unmatched_token_count", summary.get("unmatched_token_count")),
        ("current_match_summary_ambiguous_token_count", summary.get("ambiguous_token_count")),
        ("full_inci_priority_rows", summary.get("full_inci_priority_rows")),
        ("alias_normalization_rows", summary.get("alias_normalization_rows")),
        ("signal_or_family_rows", summary.get("signal_or_family_rows")),
        ("verify_parser_rows", summary.get("verify_parser_rows")),
        ("manual_review_rows", summary.get("manual_review_rows")),
        ("all_triage_rows", summary.get("all_triage_rows")),
        ("queue_source_match_csv", summary.get("match_csv")),
        ("queue_source_ingredient_workbook", summary.get("ingredient_workbook")),
        ("queue_note", "v2.2 queue rebuild from latest unmatched candidate rows; use queue buckets before any further canonical backfill."),
    ]:
        readme.append([metric, value])

    bucket_to_rows: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for row in triage_rows:
        bucket_to_rows[row["recommended_bucket"]].append(row)

    for bucket in [
        "candidate_canonical_full_inci",
        "alias_or_normalization_gap",
        "signal_or_family_term",
        "verify_parser_or_export",
        "manual_review",
    ]:
        sheet = workbook.create_sheet(BUCKET_META[bucket]["sheet"])
        sheet.append(QUEUE_COLUMNS)
        for row in bucket_to_rows.get(bucket, []):
            sheet.append([row.get(column, "") for column in QUEUE_COLUMNS])

    all_triage = workbook.create_sheet("All_Triage")
    all_triage.append(QUEUE_COLUMNS)
    for row in triage_rows:
        all_triage.append([row.get(column, "") for column in QUEUE_COLUMNS])

    path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(path)


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=QUEUE_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def load_summary(path: Path | None) -> dict[str, Any]:
    if not path:
        return {}
    with path.open("r", encoding="utf-8") as handle:
        return json.load(handle)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build triage workbook for unmatched ingredient tokens.")
    parser.add_argument("--match-csv", required=True, help="Candidate match CSV with matched/unmatched rows")
    parser.add_argument("--ingredient-xlsx", required=True, help="Current ingredient workbook path")
    parser.add_argument("--ingredient-sheet", help="Optional ingredient workbook sheet")
    parser.add_argument("--match-summary-json", help="Optional latest match summary JSON for README metrics")
    parser.add_argument("--out-xlsx", required=True, help="Where to write the triage workbook")
    parser.add_argument("--out-json", required=True, help="Where to write the queue summary JSON")
    parser.add_argument("--out-all-triage-csv", help="Optional CSV export for the All_Triage sheet")
    args = parser.parse_args()

    match_csv = Path(args.match_csv).expanduser().resolve()
    ingredient_xlsx = Path(args.ingredient_xlsx).expanduser().resolve()
    out_xlsx = Path(args.out_xlsx).expanduser().resolve()
    out_json = Path(args.out_json).expanduser().resolve()
    out_all_triage_csv = Path(args.out_all_triage_csv).expanduser().resolve() if args.out_all_triage_csv else None
    summary_json_path = Path(args.match_summary_json).expanduser().resolve() if args.match_summary_json else None

    match_rows = read_match_rows(match_csv)
    ingredient_sheet = resolve_ingredient_sheet_name(ingredient_xlsx, args.ingredient_sheet)
    master_like_exact, master_like_semantic = build_master_like_keys(ingredient_xlsx, ingredient_sheet)
    triage_rows = build_aggregate_rows(match_rows, master_like_exact, master_like_semantic)
    match_summary = load_summary(summary_json_path)

    bucket_counts = Counter(row["recommended_bucket"] for row in triage_rows)
    summary = {
        "match_csv": str(match_csv),
        "ingredient_workbook": str(ingredient_xlsx),
        "ingredient_sheet": ingredient_sheet,
        "token_count": match_summary.get("token_count"),
        "matched_token_count": match_summary.get("matched_token_count"),
        "unmatched_token_count": match_summary.get("unmatched_token_count"),
        "ambiguous_token_count": match_summary.get("ambiguous_token_count"),
        "full_inci_priority_rows": bucket_counts.get("candidate_canonical_full_inci", 0),
        "alias_normalization_rows": bucket_counts.get("alias_or_normalization_gap", 0),
        "signal_or_family_rows": bucket_counts.get("signal_or_family_term", 0),
        "verify_parser_rows": bucket_counts.get("verify_parser_or_export", 0),
        "manual_review_rows": bucket_counts.get("manual_review", 0),
        "all_triage_rows": len(triage_rows),
        "out_xlsx": str(out_xlsx),
        "out_all_triage_csv": str(out_all_triage_csv) if out_all_triage_csv else None,
    }

    write_workbook(out_xlsx, triage_rows, summary)
    if out_all_triage_csv:
        write_csv(out_all_triage_csv, triage_rows)

    out_json.parent.mkdir(parents=True, exist_ok=True)
    out_json.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
