#!/usr/bin/env python3
"""Apply a curated overlay to resolve the safest SKU ingredient/signal tail rows."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to read ingredient workbooks. Install it in the local Python environment first."
    ) from exc


PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]

PARSER_INGREDIENT_OVERLAYS = {
    "Active Ingredient(s) & Concentration: Octinoxate 7.5%": {
        "target_canonical_inci_name": "Ethylhexyl Methoxycinnamate",
        "matched_reference_term": "Octinoxate",
    },
    "Zinc Oxide 9.0%. Inactive Ingredients: Water": {
        "target_canonical_inci_name": "Zinc Oxide",
        "matched_reference_term": "Zinc Oxide",
    },
    "Slaai ingredients: Ethylhexyl Palmitate": {
        "target_canonical_inci_name": "Ethylhexyl Palmitate",
        "matched_reference_term": "Ethylhexyl Palmitate",
    },
    "Active Ingredients: Petrolatum 60%": {
        "target_canonical_inci_name": "Petrolatum",
        "matched_reference_term": "Petrolatum",
    },
    "Inactive Ingredients: Microcrystalline Wax": {
        "target_canonical_inci_name": "Microcrystalline Wax",
        "matched_reference_term": "Microcrystalline Wax",
    },
    "Active Ingredients: Titanium Dioxide 11.6%": {
        "target_canonical_inci_name": "Titanium Dioxide",
        "matched_reference_term": "Titanium Dioxide",
    },
    "Zinc Oxide 8.6%. Inactive Ingredients: Cyclopentasiloxane": {
        "target_canonical_inci_name": "Zinc Oxide",
        "matched_reference_term": "Zinc Oxide",
    },
    "Active Ingredients: Zinc Oxide 12%. Inactive Ingredients: Water": {
        "target_canonical_inci_name": "Zinc Oxide",
        "matched_reference_term": "Zinc Oxide",
    },
}

CURATED_SIGNAL_OVERLAYS = {
    "22 plant extracts": {
        "signal_bucket": "marketing_or_blend_signal",
        "signal_key": "22_plant_extracts",
        "display_signal_name": "22 plant extracts",
    },
    "advanced age-defense technology": {
        "signal_bucket": "claim_phrase_signal",
        "signal_key": "advanced_age_defense_technology",
        "display_signal_name": "advanced age-defense technology",
    },
    "LIPOCHROMAN™": {
        "signal_bucket": "named_active_signal",
        "signal_key": "lipochroman",
        "display_signal_name": "LIPOCHROMAN™",
    },
    "PHYSAVIE™": {
        "signal_bucket": "named_active_signal",
        "signal_key": "physavie",
        "display_signal_name": "PHYSAVIE™",
    },
    "Vital ET™": {
        "signal_bucket": "marketing_or_blend_signal",
        "signal_key": "vital_et",
        "display_signal_name": "Vital ET™",
    },
    "Venuceane™": {
        "signal_bucket": "named_active_signal",
        "signal_key": "venuceane",
        "display_signal_name": "Venuceane™",
    },
    "Revinage™": {
        "signal_bucket": "named_active_signal",
        "signal_key": "revinage",
        "display_signal_name": "Revinage™",
    },
    "ZPOLY™ Complex": {
        "signal_bucket": "marketing_or_blend_signal",
        "signal_key": "zpoly_complex",
        "display_signal_name": "ZPOLY™ Complex",
    },
}

PARSER_FRAGMENT_EXCLUSION_OVERLAYS = {
    (
        "Bobbi Brown",
        "Vitamin Enriched Face Base",
        "Vitamins B, C, and E; Hyaluronic Acid; Squalane; Shea Butter",
        "C",
    ): {
        "exclusion_reason": "parser_fragment_from_key_ingredient_series",
    },
    (
        "Bobbi Brown",
        "Vitamin Enriched Face Base",
        "Vitamins B, C, and E; Hyaluronic Acid; Squalane; Shea Butter",
        "and E",
    ): {
        "exclusion_reason": "parser_fragment_from_key_ingredient_series",
    },
    (
        "Dr.Jart+",
        "Ceramidin Skin Barrier Moisturizing Cream",
        "Ceramides NP, NG, NS, AS, AP; Panthenol; Glycerin",
        "NG",
    ): {
        "exclusion_reason": "parser_fragment_from_ceramide_suffix_series",
    },
    (
        "Dr.Jart+",
        "Ceramidin Skin Barrier Moisturizing Cream",
        "Ceramides NP, NG, NS, AS, AP; Panthenol; Glycerin",
        "NS",
    ): {
        "exclusion_reason": "parser_fragment_from_ceramide_suffix_series",
    },
    (
        "Dr.Jart+",
        "Ceramidin Skin Barrier Moisturizing Cream",
        "Ceramides NP, NG, NS, AS, AP; Panthenol; Glycerin",
        "AS",
    ): {
        "exclusion_reason": "parser_fragment_from_ceramide_suffix_series",
    },
    (
        "Dr.Jart+",
        "Ceramidin Skin Barrier Moisturizing Cream",
        "Ceramides NP, NG, NS, AS, AP; Panthenol; Glycerin",
        "AP",
    ): {
        "exclusion_reason": "parser_fragment_from_ceramide_suffix_series",
    },
}


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def resolve_sheet_name(path: Path, requested_sheet: str | None) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise SystemExit(f"Workbook missing required '{requested_sheet}' sheet: {path}")
        return requested_sheet
    for sheet_name in PREFERRED_INGREDIENT_SHEETS:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    raise SystemExit(f"Workbook missing any supported sheet ({', '.join(PREFERRED_INGREDIENT_SHEETS)}): {path}")


def build_ingredient_lookup(path: Path, sheet_name: str) -> dict[str, dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    out: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        if not row or not any(normalize_text(cell) for cell in row):
            continue
        record = {
            header[index]: normalize_text(row[index] if index < len(row) else "")
            for index in range(len(header))
        }
        canonical = normalize_text(record.get("canonical_inci_name"))
        if not canonical:
            continue
        out[canonical] = {
            "ingredient_record_id": normalize_text(record.get("record_id")),
            "canonical_inci_name": canonical,
            "canonical_display_name": normalize_text(record.get("canonical_display_name")) or canonical,
            "ingredient_normalized_key": normalize_text(record.get("normalized_key")),
            "ingredient_family": normalize_text(record.get("ingredient_family")),
            "primary_bucket": normalize_text(record.get("primary_bucket")),
        }
    return out


def read_csv(path: Path) -> tuple[list[str], list[dict[str, str]]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        reader = csv.DictReader(handle)
        return list(reader.fieldnames or []), list(reader)


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Apply a curated overlay to resolve the safest SKU ingredient/signal tail rows.")
    parser.add_argument("--candidate-csv", required=True, help="Combined audit candidate CSV")
    parser.add_argument("--ingredient-xlsx", required=True, help="Ingredient workbook for canonical lookups")
    parser.add_argument("--ingredient-sheet", help="Optional ingredient workbook sheet")
    parser.add_argument("--out-csv", required=True, help="Where to write updated candidate CSV")
    parser.add_argument("--out-summary-json", required=True, help="Where to write overlay summary JSON")
    parser.add_argument("--out-remainder-csv", help="Optional CSV for rows still unresolved after overlay")
    args = parser.parse_args()

    fieldnames, rows = read_csv(Path(args.candidate_csv).expanduser().resolve())
    ingredient_path = Path(args.ingredient_xlsx).expanduser().resolve()
    ingredient_sheet = resolve_sheet_name(ingredient_path, args.ingredient_sheet)
    ingredient_lookup = build_ingredient_lookup(ingredient_path, ingredient_sheet)

    parser_overlay_count = 0
    signal_overlay_count = 0
    parser_fragment_exclusion_count = 0
    unresolved_rows: list[dict[str, str]] = []
    resolution_counts: Counter[str] = Counter()

    out_rows: list[dict[str, str]] = []
    for row in rows:
        raw_token = normalize_text(row.get("raw_token"))
        resolution_type = normalize_text(row.get("audit_resolution_type"))
        overlay_key = (
            normalize_text(row.get("brand_name")),
            normalize_text(row.get("product_name")),
            normalize_text(row.get("raw_ingredient_text")),
            raw_token,
        )

        if resolution_type == "no_deterministic_match" and raw_token in PARSER_INGREDIENT_OVERLAYS:
            spec = PARSER_INGREDIENT_OVERLAYS[raw_token]
            target = ingredient_lookup.get(spec["target_canonical_inci_name"])
            if not target:
                raise SystemExit(f"Missing target ingredient canonical in workbook: {spec['target_canonical_inci_name']}")
            row.update(target)
            row["ingredient_match_status"] = "matched"
            row["ingredient_match_method"] = "parser_cleanup_curated"
            row["ingredient_match_confidence"] = "high"
            row["matched_reference_term"] = spec["matched_reference_term"]
            row["signal_match_status"] = ""
            row["signal_match_score"] = ""
            row["signal_match_method"] = ""
            row["signal_bucket"] = ""
            row["signal_key"] = ""
            row["display_signal_name"] = ""
            row["signal_confidence_levels"] = ""
            row["signal_source_packets"] = ""
            row["signal_source_decisions"] = ""
            row["audit_resolution_status"] = "covered"
            row["audit_resolution_type"] = "parser_cleanup_ingredient_match"
            row["audit_resolution_rank"] = "95"
            parser_overlay_count += 1
        elif resolution_type == "no_deterministic_match" and overlay_key in PARSER_FRAGMENT_EXCLUSION_OVERLAYS:
            spec = PARSER_FRAGMENT_EXCLUSION_OVERLAYS[overlay_key]
            row["ingredient_match_status"] = "excluded"
            row["ingredient_match_method"] = "parser_fragment_exclusion"
            row["ingredient_match_confidence"] = "high"
            row["ingredient_record_id"] = ""
            row["canonical_inci_name"] = ""
            row["canonical_display_name"] = ""
            row["ingredient_normalized_key"] = ""
            row["ingredient_family"] = ""
            row["primary_bucket"] = ""
            row["matched_reference_term"] = ""
            row["signal_match_status"] = ""
            row["signal_match_score"] = ""
            row["signal_match_method"] = ""
            row["signal_bucket"] = ""
            row["signal_key"] = ""
            row["display_signal_name"] = ""
            row["signal_confidence_levels"] = ""
            row["signal_source_packets"] = ""
            row["signal_source_decisions"] = spec["exclusion_reason"]
            row["audit_resolution_status"] = "covered"
            row["audit_resolution_type"] = "parser_fragment_excluded"
            row["audit_resolution_rank"] = "60"
            parser_fragment_exclusion_count += 1
        elif resolution_type == "no_deterministic_match" and raw_token in CURATED_SIGNAL_OVERLAYS:
            spec = CURATED_SIGNAL_OVERLAYS[raw_token]
            row["signal_match_status"] = "matched"
            row["signal_match_score"] = "95"
            row["signal_match_method"] = "curated_signal_overlay"
            row["signal_bucket"] = spec["signal_bucket"]
            row["signal_key"] = spec["signal_key"]
            row["display_signal_name"] = spec["display_signal_name"]
            row["signal_confidence_levels"] = "high"
            row["signal_source_packets"] = "curated_tail_overlay"
            row["signal_source_decisions"] = "approve_curated_overlay"
            row["audit_resolution_status"] = "covered"
            row["audit_resolution_type"] = "curated_signal_tail_overlay"
            row["audit_resolution_rank"] = "70"
            signal_overlay_count += 1

        resolution_counts[normalize_text(row.get("audit_resolution_type"))] += 1
        if normalize_text(row.get("audit_resolution_status")) != "covered":
            unresolved_rows.append(row)
        out_rows.append(row)

    out_csv = Path(args.out_csv).expanduser().resolve()
    write_csv(out_csv, fieldnames, out_rows)

    if args.out_remainder_csv:
        write_csv(Path(args.out_remainder_csv).expanduser().resolve(), fieldnames, unresolved_rows)

    summary = {
        "candidate_csv": str(Path(args.candidate_csv).expanduser().resolve()),
        "ingredient_workbook": str(ingredient_path),
        "ingredient_sheet": ingredient_sheet,
        "row_count": len(out_rows),
        "parser_overlay_count": parser_overlay_count,
        "signal_overlay_count": signal_overlay_count,
        "parser_fragment_exclusion_count": parser_fragment_exclusion_count,
        "total_overlay_count": parser_overlay_count + signal_overlay_count + parser_fragment_exclusion_count,
        "remaining_unresolved_count": len(unresolved_rows),
        "resolution_type_counts": dict(resolution_counts),
        "out_csv": str(out_csv),
        "out_remainder_csv": str(Path(args.out_remainder_csv).expanduser().resolve()) if args.out_remainder_csv else "",
    }

    out_summary_json = Path(args.out_summary_json).expanduser().resolve()
    out_summary_json.parent.mkdir(parents=True, exist_ok=True)
    out_summary_json.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
