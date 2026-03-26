#!/usr/bin/env python3
"""Build a decision-ready packet for confirming no-safe-common-alias rows."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build the no-safe-common-alias confirmation packet. Install it in the local Python environment first."
    ) from exc


PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]


NO_SAFE_ALIAS_MARKER = "confirmed_no_safe_common_alias"


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def load_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def resolve_sheet_name(path: Path, requested_sheet: str | None = None) -> str:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet_names = workbook.sheetnames
    workbook.close()
    if requested_sheet:
        if requested_sheet not in sheet_names:
            raise SystemExit(f"Workbook sheet not found: {requested_sheet}")
        return requested_sheet
    for candidate in PREFERRED_INGREDIENT_SHEETS:
        if candidate in sheet_names:
            return candidate
    raise SystemExit(
        "Workbook is missing a supported ingredient sheet. Expected one of: "
        + ", ".join(PREFERRED_INGREDIENT_SHEETS)
    )


def read_workbook_review_notes(path: Path, sheet_name: str) -> dict[str, str]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    try:
        record_id_idx = header.index("record_id")
        review_notes_idx = header.index("review_notes")
    except ValueError as exc:
        raise SystemExit("Workbook target sheet must include record_id and review_notes columns.") from exc

    result: dict[str, str] = {}
    for row in rows[1:]:
        record_id = normalize_text(row[record_id_idx] if record_id_idx < len(row) else "")
        if not record_id:
            continue
        review_notes = normalize_text(row[review_notes_idx] if review_notes_idx < len(row) else "")
        result[record_id] = review_notes
    return result


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a decision-ready packet for no-safe-common-alias confirmation.")
    parser.add_argument("--ingredient-xlsx", required=True, help="Current ingredient workbook")
    parser.add_argument("--sheet-name", help="Optional workbook sheet name")
    parser.add_argument("--alias-gap-csv", required=True, help="CSV of remaining alias-gap rows")
    parser.add_argument("--out-csv", required=True, help="Where to write the decision packet CSV")
    parser.add_argument("--out-json", help="Optional path for a JSON copy of the packet")
    args = parser.parse_args()

    workbook_path = Path(args.ingredient_xlsx).expanduser().resolve()
    alias_gap_path = Path(args.alias_gap_csv).expanduser().resolve()
    rows = load_rows(alias_gap_path)
    sheet_name = resolve_sheet_name(workbook_path, args.sheet_name)
    review_notes_by_record = read_workbook_review_notes(workbook_path, sheet_name)

    packet_rows: list[dict[str, str]] = []
    for row in rows:
        record_id = normalize_text(row.get("record_id"))
        canonical = normalize_text(row.get("canonical_inci_name"))
        ingredient_family = normalize_text(row.get("ingredient_family"))
        current_review_notes = review_notes_by_record.get(record_id, "")
        packet_rows.append(
            {
                "record_id": record_id,
                "canonical_inci_name": canonical,
                "ingredient_family": ingredient_family,
                "existing_review_notes": current_review_notes,
                "suggested_marker": NO_SAFE_ALIAS_MARKER,
                "suggested_resolution": "confirm_no_safe_common_alias",
                "suggested_rationale": "workbook context shows only exact self-name variants",
                "decision": "",
                "approved_marker": NO_SAFE_ALIAS_MARKER,
                "reviewer_notes": "",
            }
        )

    fieldnames = [
        "record_id",
        "canonical_inci_name",
        "ingredient_family",
        "existing_review_notes",
        "suggested_marker",
        "suggested_resolution",
        "suggested_rationale",
        "decision",
        "approved_marker",
        "reviewer_notes",
    ]

    out_csv = Path(args.out_csv).expanduser().resolve()
    write_csv(out_csv, packet_rows, fieldnames)

    out_json = None
    if args.out_json:
        out_json = Path(args.out_json).expanduser().resolve()
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(
            json.dumps(
                {
                    "source_workbook": str(workbook_path),
                    "source_sheet": sheet_name,
                    "source_alias_gap_csv": str(alias_gap_path),
                    "row_count": len(packet_rows),
                    "decision_values": [
                        "confirm_no_safe_common_alias",
                        "keep_open",
                        "needs_research",
                    ],
                    "rows": packet_rows,
                },
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

    print(
        json.dumps(
            {
                "source_workbook": str(workbook_path),
                "source_sheet": sheet_name,
                "source_alias_gap_csv": str(alias_gap_path),
                "row_count": len(packet_rows),
                "out_csv": str(out_csv),
                "out_json": str(out_json) if out_json else None,
            },
            ensure_ascii=True,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
