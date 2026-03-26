#!/usr/bin/env python3
"""Build an apply-ready workbook alias/parser bridge patch from reviewed bridge candidates."""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build the ingredient-master bridge patch. Install it in the local Python environment first."
    ) from exc


TARGET_SHEET = "Ingredient_Reference_Merged_v2"

PATCH_FIELDS = [
    "record_id",
    "canonical_inci_name",
    "existing_aliases_common",
    "existing_parser_variants",
    "existing_alias_quality",
    "patch_aliases_common",
    "patch_parser_variants",
    "patch_alias_quality",
    "proposal_sources",
    "quality_reason",
]


BRIDGE_RULES: dict[str, dict[str, Any]] = {
    "calendula": {
        "alias_additions": ["Calendula"],
        "parser_additions": ["Calendula", "calendula"],
        "force_alias_quality": "common_alias",
        "patch_type": "common_alias_and_parser_bridge",
        "quality_reason": "Bridge reviewed common botanical name onto existing calendula extract canonical row.",
    },
    "dexpanthenol": {
        "alias_additions": ["Dexpanthenol"],
        "parser_additions": ["Dexpanthenol", "dexpanthenol"],
        "force_alias_quality": "common_alias",
        "patch_type": "common_alias_and_parser_bridge",
        "quality_reason": "Bridge reviewed Dexpanthenol synonym onto existing Panthenol canonical row.",
    },
    "ginkgo_biloba": {
        "alias_additions": ["Ginkgo Biloba"],
        "parser_additions": ["Ginkgo Biloba", "ginkgo biloba"],
        "force_alias_quality": "common_alias",
        "patch_type": "common_alias_and_parser_bridge",
        "quality_reason": "Bridge reviewed common botanical name onto existing Ginkgo Biloba Leaf Extract canonical row.",
    },
    "glycerin": {
        "alias_additions": [],
        "parser_additions": ["Pure Glycerin", "pure glycerin"],
        "force_alias_quality": None,
        "patch_type": "parser_variant_only",
        "quality_reason": "Route reviewed PDP phrase 'pure glycerin' onto existing Glycerin canonical row without promoting it as a common alias.",
    },
    "urea": {
        "alias_additions": [],
        "parser_additions": ["Pure Urea", "pure urea"],
        "force_alias_quality": None,
        "patch_type": "parser_variant_only",
        "quality_reason": "Route reviewed PDP phrase 'pure urea' onto existing Urea canonical row without promoting it as a common alias.",
    },
    "royal_jelly": {
        "alias_additions": ["Royal Jelly"],
        "parser_additions": ["Royal Jelly", "royal jelly"],
        "force_alias_quality": "common_alias",
        "patch_type": "common_alias_and_parser_bridge",
        "quality_reason": "Bridge reviewed common ingredient name onto existing Royal Jelly Extract canonical row.",
    },
}


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def split_multi(value: str) -> list[str]:
    return [part.strip() for part in normalize_text(value).split(";") if part.strip()]


def join_multi(values: list[str]) -> str:
    return "; ".join(values)


def ordered_union(existing: list[str], additions: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in existing + additions:
        token = normalize_text(value)
        if not token:
            continue
        if token in seen:
            continue
        seen.add(token)
        out.append(token)
    return out


def read_csv_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return [{key: normalize_text(value) for key, value in row.items()} for row in csv.DictReader(handle)]


def write_csv(path: Path, fieldnames: list[str], rows: list[dict[str, str]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def load_workbook_rows(path: Path) -> dict[str, dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if TARGET_SHEET not in workbook.sheetnames:
        raise SystemExit(f"Workbook missing target sheet: {TARGET_SHEET}")
    sheet = workbook[TARGET_SHEET]
    header = [normalize_text(cell.value) for cell in sheet[1]]
    index = {name: idx for idx, name in enumerate(header)}
    required = [
        "record_id",
        "canonical_inci_name",
        "aliases_common",
        "parser_variants",
        "alias_quality",
    ]
    missing = [name for name in required if name not in index]
    if missing:
        raise SystemExit(f"Workbook missing required columns: {', '.join(missing)}")

    rows: dict[str, dict[str, str]] = {}
    for raw in sheet.iter_rows(min_row=2, values_only=True):
        row = {name: normalize_text(raw[idx]) for name, idx in index.items()}
        record_id = row.get("record_id")
        if record_id:
            rows[record_id] = row
    return rows


def main() -> None:
    parser = argparse.ArgumentParser(description="Build an ingredient-master alias/parser bridge patch from reviewed bridge candidates.")
    parser.add_argument("--bridge-csv", required=True, help="Existing-canonical bridge candidate CSV")
    parser.add_argument("--ingredient-xlsx", required=True, help="Source ingredient workbook")
    parser.add_argument("--out-patch-csv", required=True, help="Where to write apply-ready patch CSV")
    parser.add_argument("--out-summary-json", required=True, help="Where to write summary JSON")
    args = parser.parse_args()

    bridge_rows = read_csv_rows(Path(args.bridge_csv).expanduser().resolve())
    workbook_rows = load_workbook_rows(Path(args.ingredient_xlsx).expanduser().resolve())

    patch_rows: list[dict[str, str]] = []
    skipped_unknown_rule: list[str] = []
    skipped_missing_record: list[str] = []

    for candidate in bridge_rows:
        signal_key = normalize_text(candidate.get("signal_key"))
        rule = BRIDGE_RULES.get(signal_key)
        if not rule:
            skipped_unknown_rule.append(signal_key)
            continue

        record_id = normalize_text(candidate.get("reference_match_display_name"))
        row = None
        for workbook_row in workbook_rows.values():
            if (
                normalize_text(workbook_row.get("canonical_display_name")) == normalize_text(candidate.get("reference_match_display_name"))
                or normalize_text(workbook_row.get("canonical_inci_name")) == normalize_text(candidate.get("reference_match_inci_name"))
            ):
                row = workbook_row
                break
        if row is None:
            skipped_missing_record.append(signal_key)
            continue

        existing_aliases = split_multi(row.get("aliases_common", ""))
        existing_parser = split_multi(row.get("parser_variants", ""))

        patch_aliases = ordered_union(existing_aliases, rule["alias_additions"])
        patch_parser = ordered_union(existing_parser, rule["parser_additions"])
        patch_quality = normalize_text(rule["force_alias_quality"] or row.get("alias_quality"))

        patch_rows.append(
            {
                "record_id": normalize_text(row.get("record_id")),
                "canonical_inci_name": normalize_text(row.get("canonical_inci_name")),
                "existing_aliases_common": join_multi(existing_aliases),
                "existing_parser_variants": join_multi(existing_parser),
                "existing_alias_quality": normalize_text(row.get("alias_quality")),
                "patch_aliases_common": join_multi(patch_aliases),
                "patch_parser_variants": join_multi(patch_parser),
                "patch_alias_quality": patch_quality,
                "proposal_sources": "ingredient_master_followup_existing_bridge",
                "quality_reason": normalize_text(rule["quality_reason"]),
            }
        )

    patch_rows.sort(key=lambda row: normalize_text(row["canonical_inci_name"]).casefold())
    out_patch = Path(args.out_patch_csv).expanduser().resolve()
    write_csv(out_patch, PATCH_FIELDS, patch_rows)

    summary = {
        "bridge_candidate_count": len(bridge_rows),
        "patch_row_count": len(patch_rows),
        "patch_type_counts": dict(
            Counter(BRIDGE_RULES[candidate["signal_key"]]["patch_type"] for candidate in bridge_rows if candidate["signal_key"] in BRIDGE_RULES)
        ),
        "skipped_unknown_rule_count": len(skipped_unknown_rule),
        "skipped_unknown_rule_keys": skipped_unknown_rule,
        "skipped_missing_record_count": len(skipped_missing_record),
        "skipped_missing_record_keys": skipped_missing_record,
        "out_patch_csv": str(out_patch),
    }

    out_summary = Path(args.out_summary_json).expanduser().resolve()
    out_summary.parent.mkdir(parents=True, exist_ok=True)
    out_summary.write_text(json.dumps(summary, ensure_ascii=True, indent=2) + "\n", encoding="utf-8")
    print(json.dumps(summary, ensure_ascii=True, indent=2))


if __name__ == "__main__":
    main()
