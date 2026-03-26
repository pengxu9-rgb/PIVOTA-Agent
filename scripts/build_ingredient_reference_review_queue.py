#!/usr/bin/env python3
"""Build a prioritized review queue for an ingredient reference workbook.

This is a read-only helper for workbook curation. It does not write to any
runtime database. It inspects the workbook against the spec and outputs:

- column fill summary
- heuristic consistency warnings
- prioritized row-level review queue
"""

from __future__ import annotations

import argparse
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to review ingredient workbook seeds. Install it in the local Python environment first."
    ) from exc


REVIEW_PRIORITY_SCORES = {
    "p1": 6,
    "p2": 3,
}

PRIORITY_ORDER = {
    "p1": 1,
    "p2": 2,
    "p3": 3,
}

FLAG_EXPECTATIONS = {
    "is_barrier_support": {
        "buckets": {"repair"},
        "benefits": {"barrier", "barrier support", "repair", "barrier repair"},
    },
    "is_uv_filter": {
        "buckets": {"sunscreen"},
        "benefits": {"uv", "sun", "sunscreen", "uva", "uvb"},
    },
    "is_fragrance_or_eo": {
        "buckets": {"fragrance/essential oil"},
        "benefits": {"fragrance", "essential oil", "perfuming", "perfume"},
    },
}

PREFERRED_INGREDIENT_SHEETS = [
    "Ingredient_Reference_Merged_v2",
    "Dictionary",
]


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def is_nonempty(value: Any) -> bool:
    return bool(normalize_text(value))


def parse_multi(value: Any) -> list[str]:
    text = normalize_text(value)
    if not text:
        return []
    return [part.strip() for part in text.split(";") if part and part.strip()]


def resolve_sheet_name(workbook_path: Path, requested_sheet: str | None) -> str:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    if requested_sheet:
        if requested_sheet not in workbook.sheetnames:
            raise SystemExit(f"Worksheet {requested_sheet} does not exist.")
        return requested_sheet
    for sheet_name in PREFERRED_INGREDIENT_SHEETS:
        if sheet_name in workbook.sheetnames:
            return sheet_name
    raise SystemExit(
        f"Workbook did not contain any supported sheet names ({', '.join(PREFERRED_INGREDIENT_SHEETS)})."
    )


def read_records(workbook_path: Path, sheet_name: str) -> tuple[list[str], list[dict[str, Any]]]:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[sheet_name]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    records = []
    for row in rows[1:]:
        if not any(is_nonempty(cell) for cell in row):
            continue
        padded = list(row) + [None] * max(0, len(header) - len(row))
        records.append(dict(zip(header, padded)))
    return header, records


def compute_column_fill(records: list[dict[str, Any]], columns: list[str]) -> dict[str, dict[str, Any]]:
    total = max(len(records), 1)
    result: dict[str, dict[str, Any]] = {}
    for column in columns:
        filled = sum(is_nonempty(record.get(column)) for record in records)
        result[column] = {
            "filled_count": filled,
            "missing_count": len(records) - filled,
            "fill_rate": round(filled / total, 4),
        }
    return result


def bool_consistency_issues(record: dict[str, Any]) -> list[str]:
    issues: list[str] = []
    buckets = {item.lower() for item in parse_multi(record.get("all_buckets"))}
    benefits = {item.lower() for item in parse_multi(record.get("benefit_tags"))}

    for flag, rule in FLAG_EXPECTATIONS.items():
        if normalize_text(record.get(flag)).lower() != "yes":
            continue
        expected_bucket = any(bucket in buckets for bucket in rule["buckets"])
        expected_benefit = any(
            any(token in benefit for token in rule["benefits"]) for benefit in benefits
        )
        if not expected_bucket and not expected_benefit:
            issues.append(f"{flag}_taxonomy_mismatch")
    return issues


def derive_row_reasons(record: dict[str, Any]) -> tuple[int, list[str]]:
    score = 0
    reasons: list[str] = []

    canonical = normalize_text(record.get("canonical_inci_name"))
    display = normalize_text(record.get("canonical_display_name"))
    aliases = normalize_text(record.get("aliases_common"))
    alias_quality = normalize_text(record.get("alias_quality"))
    notes_for_parser = normalize_text(record.get("notes_for_parser"))
    ingredient_family = normalize_text(record.get("ingredient_family"))
    cross_market_notes = normalize_text(record.get("cross_market_notes"))
    deprecated_aliases = normalize_text(record.get("deprecated_aliases"))
    review_status = normalize_text(record.get("review_status")).lower()
    confidence = normalize_text(record.get("confidence")).lower()
    review_notes = normalize_text(record.get("review_notes")).lower()
    us_label = normalize_text(record.get("us_label_name"))
    eu_label = normalize_text(record.get("eu_label_name"))
    parser_variants = parse_multi(record.get("parser_variants"))
    no_safe_common_alias_confirmed = "confirmed_no_safe_common_alias" in review_notes
    ingredient_family_other_confirmed = "confirmed_ingredient_family_other" in review_notes

    if not aliases and not no_safe_common_alias_confirmed:
        if display and display != canonical:
            score += 3
            reasons.append("missing_aliases_common_for_noncanonical_display")
        else:
            score += 1
            reasons.append("missing_aliases_common")

    if (aliases or deprecated_aliases) and not alias_quality:
        score += 2
        reasons.append("missing_alias_quality")

    if not notes_for_parser and (len(parser_variants) >= 4 or (display and display != canonical)):
        score += 2
        reasons.append("missing_notes_for_parser")

    if us_label and eu_label and us_label != eu_label and not cross_market_notes:
        score += 2
        reasons.append("missing_cross_market_notes")

    if ingredient_family == "other" and not ingredient_family_other_confirmed:
        score += 1
        reasons.append("ingredient_family_other_review")

    if review_status == "draft":
        score += 1
        reasons.append("review_status_still_draft")

    medium_confidence_confirmed = "confirmed_confidence_medium" in review_notes
    low_confidence_confirmed = "confirmed_confidence_low" in review_notes

    if confidence == "medium" and not medium_confidence_confirmed:
        score += 1
        reasons.append("confidence_medium")
    if confidence == "low" and not low_confidence_confirmed:
        score += 1
        reasons.append("confidence_low")

    consistency = bool_consistency_issues(record)
    if consistency:
        score += 2
        reasons.extend(consistency)

    return score, reasons


def priority_from_score(score: int) -> str | None:
    if score >= REVIEW_PRIORITY_SCORES["p1"]:
        return "p1"
    if score >= REVIEW_PRIORITY_SCORES["p2"]:
        return "p2"
    if score > 0:
        return "p3"
    return None


def build_review_queue(records: list[dict[str, Any]]) -> list[dict[str, Any]]:
    queue: list[dict[str, Any]] = []
    for record in records:
        score, reasons = derive_row_reasons(record)
        priority = priority_from_score(score)
        if not priority:
            continue
        queue.append(
            {
                "priority": priority,
                "priority_score": score,
                "record_id": normalize_text(record.get("record_id")),
                "canonical_inci_name": normalize_text(record.get("canonical_inci_name")),
                "canonical_display_name": normalize_text(record.get("canonical_display_name")),
                "ingredient_family": normalize_text(record.get("ingredient_family")),
                "review_status": normalize_text(record.get("review_status")),
                "confidence": normalize_text(record.get("confidence")),
                "reasons": reasons,
            }
        )
    return sorted(
        queue,
        key=lambda item: (
            0 if item["priority"] == "p1" else 1 if item["priority"] == "p2" else 2,
            -item["priority_score"],
            item["canonical_inci_name"],
        ),
    )


def top_reason_counts(queue: list[dict[str, Any]]) -> dict[str, int]:
    counts = Counter()
    for item in queue:
        counts.update(item["reasons"])
    return dict(counts.most_common())


def write_csv(path: Path, rows: list[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "priority",
        "priority_score",
        "record_id",
        "canonical_inci_name",
        "canonical_display_name",
        "ingredient_family",
        "review_status",
        "confidence",
        "reasons",
    ]
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        for row in rows:
            out = dict(row)
            out["reasons"] = "; ".join(row["reasons"])
            writer.writerow(out)


def filter_by_min_priority(rows: list[dict[str, Any]], min_priority: str | None) -> list[dict[str, Any]]:
    if not min_priority:
        return rows
    threshold = PRIORITY_ORDER[min_priority]
    return [row for row in rows if PRIORITY_ORDER[row["priority"]] <= threshold]


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a prioritized review queue for ingredient reference workbook curation.")
    parser.add_argument("--ingredient-xlsx", required=True, help="Path to the ingredient reference workbook")
    parser.add_argument("--sheet-name", help="Optional workbook sheet name")
    parser.add_argument("--out-json", help="Optional path to write JSON output")
    parser.add_argument("--out-csv", help="Optional path to write row-level review CSV")
    parser.add_argument("--top", type=int, default=50, help="Number of top review rows to print in JSON summary")
    parser.add_argument(
        "--min-priority",
        choices=["p1", "p2", "p3"],
        help="Only keep rows at or above this priority in JSON and CSV outputs",
    )
    args = parser.parse_args()

    workbook_path = Path(args.ingredient_xlsx).expanduser().resolve()
    sheet_name = resolve_sheet_name(workbook_path, args.sheet_name)
    _, records = read_records(workbook_path, sheet_name)

    tracked_columns = [
        "aliases_common",
        "deprecated_aliases",
        "alias_quality",
        "notes_for_parser",
        "cross_market_notes",
        "review_status",
        "confidence",
        "source_authorities",
        "source_types",
        "review_notes",
    ]
    column_fill = compute_column_fill(records, tracked_columns)
    queue = build_review_queue(records)
    filtered_queue = filter_by_min_priority(queue, args.min_priority)
    priority_counts = Counter(item["priority"] for item in filtered_queue)

    payload = {
        "workbook": str(workbook_path),
        "sheet_name": sheet_name,
        "row_count": len(records),
        "min_priority": args.min_priority or "all",
        "column_fill": column_fill,
        "priority_counts": dict(priority_counts),
        "top_reason_counts": top_reason_counts(filtered_queue),
        "all_review_rows": filtered_queue,
        "top_review_rows": filtered_queue[: max(args.top, 0)],
    }

    rendered = json.dumps(payload, ensure_ascii=True, indent=2)
    if args.out_json:
        out_json = Path(args.out_json).expanduser().resolve()
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(rendered + "\n", encoding="utf-8")
    if args.out_csv:
        write_csv(Path(args.out_csv).expanduser().resolve(), filtered_queue)
    print(rendered)


if __name__ == "__main__":
    main()
