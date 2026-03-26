#!/usr/bin/env python3
"""Build a decision-ready packet for confidence confirmation or override."""

from __future__ import annotations

import argparse
import csv
import json
from pathlib import Path
from typing import Any

try:
    from openpyxl import load_workbook
except Exception as exc:  # pragma: no cover
    raise SystemExit(
        "openpyxl is required to build the ingredient confidence confirmation packet. Install it in the local Python environment first."
    ) from exc


def normalize_text(value: Any) -> str:
    return str(value or "").strip()


def confidence_marker(confidence: str) -> str:
    value = normalize_text(confidence).lower()
    if value == "medium":
        return "confirmed_confidence_medium"
    if value == "low":
        return "confirmed_confidence_low"
    return ""


def load_rows(path: Path) -> list[dict[str, str]]:
    with path.open("r", encoding="utf-8", newline="") as handle:
        return list(csv.DictReader(handle))


def read_workbook_context(path: Path) -> dict[str, dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    if "Dictionary" not in workbook.sheetnames:
        raise SystemExit("Workbook missing required 'Dictionary' sheet.")
    sheet = workbook["Dictionary"]
    rows = list(sheet.iter_rows(values_only=True))
    header = [normalize_text(cell) for cell in rows[0]]
    result: dict[str, dict[str, str]] = {}
    for row in rows[1:]:
        record = {
            key: normalize_text(row[index] if index < len(row) else "")
            for index, key in enumerate(header)
        }
        record_id = record.get("record_id")
        if record_id:
            result[record_id] = record
    return result


def write_csv(path: Path, rows: list[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    parser = argparse.ArgumentParser(description="Build a decision-ready packet for ingredient confidence confirmation or override.")
    parser.add_argument("--ingredient-xlsx", required=True, help="Current ingredient workbook")
    parser.add_argument("--status-confidence-csv", required=True, help="CSV of remaining status/confidence review rows")
    parser.add_argument("--out-csv", required=True, help="Where to write the decision packet CSV")
    parser.add_argument("--out-json", help="Optional path for a JSON copy of the packet")
    args = parser.parse_args()

    workbook_path = Path(args.ingredient_xlsx).expanduser().resolve()
    source_csv = Path(args.status_confidence_csv).expanduser().resolve()
    rows = load_rows(source_csv)
    workbook_context = read_workbook_context(workbook_path)

    packet_rows: list[dict[str, str]] = []
    for row in rows:
        record_id = normalize_text(row.get("record_id"))
        record = workbook_context.get(record_id, {})
        existing_confidence = record.get("confidence", normalize_text(row.get("confidence")))
        packet_rows.append(
            {
                "record_id": record_id,
                "canonical_inci_name": normalize_text(row.get("canonical_inci_name")) or record.get("canonical_inci_name", ""),
                "ingredient_family": record.get("ingredient_family", normalize_text(row.get("ingredient_family"))),
                "review_status": record.get("review_status", normalize_text(row.get("review_status"))),
                "existing_confidence": existing_confidence,
                "existing_review_notes": record.get("review_notes", ""),
                "suggested_marker": confidence_marker(existing_confidence),
                "suggested_resolution": "confirm_current_confidence",
                "suggested_rationale": "current confidence retained after structured workbook review",
                "decision": "",
                "approved_confidence": existing_confidence,
                "approved_marker": confidence_marker(existing_confidence),
                "reviewer_notes": "",
            }
        )

    fieldnames = [
        "record_id",
        "canonical_inci_name",
        "ingredient_family",
        "review_status",
        "existing_confidence",
        "existing_review_notes",
        "suggested_marker",
        "suggested_resolution",
        "suggested_rationale",
        "decision",
        "approved_confidence",
        "approved_marker",
        "reviewer_notes",
    ]

    out_csv = Path(args.out_csv).expanduser().resolve()
    write_csv(out_csv, packet_rows, fieldnames)

    out_json = None
    if args.out_json:
        out_json = Path(args.out_json).expanduser().resolve()
        out_json.parent.mkdir(parents=True, exist_ok=True)
        out_json.write_text(
            json.dumps(
                {
                    "source_workbook": str(workbook_path),
                    "source_status_confidence_csv": str(source_csv),
                    "row_count": len(packet_rows),
                    "decision_values": [
                        "confirm_current_confidence",
                        "set_high",
                        "set_low",
                        "needs_research",
                    ],
                    "rows": packet_rows,
                },
                ensure_ascii=True,
                indent=2,
            )
            + "\n",
            encoding="utf-8",
        )

    print(
        json.dumps(
            {
                "source_workbook": str(workbook_path),
                "source_status_confidence_csv": str(source_csv),
                "row_count": len(packet_rows),
                "out_csv": str(out_csv),
                "out_json": str(out_json) if out_json else None,
            },
            ensure_ascii=True,
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
